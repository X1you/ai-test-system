#!/usr/bin/env python3
"""
Pipeline API 路由。

Endpoints:
  POST /api/pipeline/start              — 上传需求 + 启动
  GET  /api/pipeline/{id}/progress      — 获取进度（JSON / HTMX 片段）
  GET  /api/pipeline/{id}/status        — 详细状态
  GET  /api/pipeline/list               — 任务列表（分页 + 搜索）
  POST /api/pipeline/{id}/cancel        — 取消
  POST /api/pipeline/{id}/resume        — 断点继续
  GET  /api/pipeline/{id}/artifacts     — 产物列表
  GET  /api/pipeline/{id}/artifacts/{name} — 下载产物
  GET  /api/pipeline/{id}/preview/{name}   — 预览产物
"""

import re
import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, JSONResponse

from core.config_loader import load_config
from core.utils import safe_join_path
from web.services.task_manager import get_task_manager

# ─── 常量 ───

# 上传目录（使用绝对路径）
UPLOAD_DIR = Path(__file__).resolve().parents[1] / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# 允许的需求文档后缀
ALLOWED_EXTENSIONS = {".md", ".txt"}

# 上传文件大小上限（字节）— 需求文档与执行结果共用
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Excel 预览最大行数
EXCEL_PREVIEW_MAX_ROWS = 50
# Excel 单元格预览最大字符数
EXCEL_CELL_MAX_CHARS = 100

# 产物文件 → 友好展示名映射
_ARTIFACT_NAMES: dict[str, str] = {
    "requirements_analysis.md": "需求分析",
    "clarification_needed.md": "待确认清单",
    "knowledge-context.md": "知识库上下文",
    "testpoints.md": "测试点清单",
    "testcases.xlsx": "测试用例",
    "testcases.xmind": "XMind 用例",
    "test_case_review_report.md": "评审报告",
    "test_report.md": "测试报告",
}

router = APIRouter(tags=["pipeline"])


# ─── 辅助函数 ───


def _require_task(pipeline_id: str):
    """获取任务，不存在时抛 404。

    消除各 endpoint 中重复的 get_task + HTTPException(404) 样板。
    """
    tm = get_task_manager()
    task = tm.get_task(pipeline_id)
    if not task:
        raise HTTPException(404, "Pipeline 不存在")
    return task


def _detect_file_type(name: str) -> str:
    """根据文件后缀推断展示类型。"""
    if name.endswith(".md"):
        return "markdown"
    if name.endswith(".xlsx"):
        return "excel"
    return "xmind"


# ─── 路由 ───


@router.post("/start")
async def start_pipeline(
    file: UploadFile = File(...),
    mode: str = Form("semi"),
    dimensions: str = Form("basic"),
    formats: str = Form("excel"),
):
    """上传需求文档 + 启动 Pipeline。"""
    # 校验文件后缀
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"仅支持 {', '.join(ALLOWED_EXTENSIONS)} 文件")

    # 读取并校验大小
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "文件过大（>10MB）")

    # 检查并发限制
    tm = get_task_manager()
    if tm.is_full():
        raise HTTPException(429, "并发任务已达上限，请等待现有任务完成")

    # 保存上传文件（安全文件名：仅保留 \w \- _ . 和空格）
    upload_id = uuid.uuid4().hex[:8]
    safe_name = re.sub(r"[^\w\-_. ]", "_", file.filename or "upload.md")
    # 截断超长文件名（防止文件系统目录项溢出 / 磁盘填充攻击）
    # 保留扩展名 + 限制 basename ≤ 100 字符
    name_stem = Path(safe_name).stem[:90]
    name_suffix = Path(safe_name).suffix.lower()
    safe_name = f"{name_stem}{name_suffix}" if name_suffix else name_stem
    upload_path = UPLOAD_DIR / f"{upload_id}_{safe_name}"
    upload_path.write_bytes(content)

    # 加载配置并创建任务
    config = load_config()
    task = tm.create_task(
        config=config,
        requirements_path=str(upload_path),
        mode=mode,
        dimensions=dimensions,
        formats=formats,
    )

    return JSONResponse(
        status_code=201,
        content={
            "pipeline_id": task.pipeline_id,
            "redirect": f"/pipeline/{task.pipeline_id}",
            "status": "running",
        },
    )


def _db_progress(pipeline_id: str) -> dict:
    """从 DB 构建历史任务的只读进度视图（重启后内存无 task 时回退）。"""
    from db.repository import get_repository
    from core.pipeline import STEP_REGISTRY, TOTAL_STEPS

    repo = get_repository()
    p = repo.get_pipeline(pipeline_id)
    if not p:
        raise HTTPException(404, "Pipeline 不存在")

    completed_ids = set(repo.get_completed_step_ids(pipeline_id))
    db_steps = repo.get_steps(pipeline_id)
    step_detail = {s.step_id: s for s in db_steps}

    steps_view = []
    for meta in STEP_REGISTRY:
        st = "done" if meta.id in completed_ids else "pending"
        detail_obj = step_detail.get(meta.id)
        steps_view.append({
            "id": meta.id,
            "name": meta.name,
            "status": st,
            "detail": detail_obj.detail if detail_obj else None,
        })

    return {
        "pipeline_id": p.id,
        "percent": round(len(completed_ids) / TOTAL_STEPS * 100),
        "status": p.status,
        "mode": p.mode,
        "completed_steps": sorted(completed_ids),
        "current_step": max(completed_ids, default=0) + 1 if p.status == "running" else 0,
        "steps": steps_view,
        "logs": [],
        "llm_stats": {},
        "error": p.error,
        "started_at": p.started_at.isoformat() if p.started_at else "",
        "output_dir": p.output_dir,
    }


def _require_task_or_db(pipeline_id: str):
    """获取内存 task；不存在时返回 (None, db_progress_dict)。"""
    tm = get_task_manager()
    task = tm.get_task(pipeline_id)
    if task:
        return task, None
    try:
        return None, _db_progress(pipeline_id)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(404, "Pipeline 不存在")


@router.get("/{pipeline_id}/progress")
async def get_progress(pipeline_id: str):
    """获取进度（Sprint 6.1 起统一返回 JSON，HTMX HTML 分支已移除）。

    优先从内存 TaskManager 获取活跃任务；
    内存中不存在时回退到 DB 构建历史进度（重启后的 interrupted/cancelled/error 等）。
    """
    task, db_data = _require_task_or_db(pipeline_id)
    if task:
        return task.get_progress()
    return db_data


@router.get("/{pipeline_id}/status")
async def get_status(
    pipeline_id: str,
):
    """详细状态。"""
    task, db_data = _require_task_or_db(pipeline_id)
    if task:
        return task.get_progress()
    return db_data


@router.get("/list")
async def list_pipelines(
    page: int = 1,
    page_size: int = 20,
    keyword: str = "",
    status: str = "",
):
    """任务列表（支持分页 + 搜索过滤）。

    - keyword: 模糊匹配 pipeline_id 或 requirements 文件名
    - status: 精确状态过滤（running/pending/done/error/paused/cancelled）
    """
    if page < 1:
        page = 1
    if page_size < 1 or page_size > 100:
        page_size = 20

    tm = get_task_manager()
    all_tasks = tm.list_tasks()

    # —— 统计仪表盘始终反映全量任务（不受过滤/分页影响）——
    s_running = s_done = 0
    for t in all_tasks:
        st0 = t.get("status", "")
        if st0 in ("running", "pending"):
            s_running += 1
        elif st0 == "done":
            s_done += 1
    all_stats = {
        "total": len(all_tasks),
        "running": s_running,
        "done": s_done,
        "other": len(all_tasks) - s_running - s_done,
    }

    # —— 搜索过滤 ——
    kw = (keyword or "").strip().lower()
    st = (status or "").strip().lower()
    if kw:
        all_tasks = [
            t
            for t in all_tasks
            if kw in t.get("pipeline_id", "").lower()
            or kw in t.get("requirements", "").lower()
        ]
    if st:
        all_tasks = [t for t in all_tasks if t.get("status", "").lower() == st]

    # —— 分页 ——
    total = len(all_tasks)
    start = (page - 1) * page_size
    items = all_tasks[start : start + page_size]
    pages = (total + page_size - 1) // page_size if page_size > 0 else 0

    return {
        "items": items,
        "pipelines": items,  # 向后兼容
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": pages,
        "keyword": keyword or "",
        "status": status or "",
        "all_stats": all_stats,
    }


@router.post("/{pipeline_id}/cancel")
async def cancel_pipeline(
    pipeline_id: str,
):
    """取消 Pipeline。"""
    task = _require_task(pipeline_id)

    if task.status not in ("running", "pending"):
        raise HTTPException(400, f"当前状态不允许取消: {task.status}")

    task.cancel()
    return {"pipeline_id": pipeline_id, "status": "cancelled"}


@router.post("/{pipeline_id}/resume")
async def resume_pipeline(
    pipeline_id: str,
    file: UploadFile = None,
):
    """断点继续（可上传已执行结果的 Excel 覆盖 output 目录中的 testcases.xlsx）。"""
    task = _require_task(pipeline_id)

    if task.status not in ("paused", "done"):
        raise HTTPException(400, f"当前状态不允许继续: {task.status}")

    # 如果有新上传的 Excel，覆盖旧文件
    if file and file.filename and file.filename.endswith(".xlsx"):
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            raise HTTPException(400, "文件过大（>10MB）")
        target_path = Path(task.output_dir) / "testcases.xlsx"
        target_path.write_bytes(content)
        task._on_log("OK", f"已接收执行结果文件: {file.filename}")

    # 后台继续执行（从 Step 6 开始）
    task.resume_background()

    return {
        "pipeline_id": pipeline_id,
        "status": "running",
        "redirect": f"/pipeline/{pipeline_id}",
    }


def _get_output_dir(pipeline_id: str) -> Path:
    """获取任务的 output_dir：优先内存 task，回退 DB。"""
    tm = get_task_manager()
    task = tm.get_task(pipeline_id)
    if task:
        return Path(task.output_dir)
    try:
        from db.repository import get_repository
        repo = get_repository()
        p = repo.get_pipeline(pipeline_id)
        if p and p.output_dir:
            return Path(p.output_dir)
    except Exception:
        pass
    raise HTTPException(404, "Pipeline 不存在")


@router.get("/{pipeline_id}/artifacts")
async def list_artifacts(
    pipeline_id: str,
):
    """产物列表。"""
    output_dir = _get_output_dir(pipeline_id)

    artifacts = []
    for fname, display_name in _ARTIFACT_NAMES.items():
        fpath = output_dir / fname
        if fpath.exists():
            artifacts.append(
                {
                    "name": fname,
                    "display_name": display_name,
                    "size": fpath.stat().st_size,
                    "type": _detect_file_type(fname),
                }
            )

    return {"pipeline_id": pipeline_id, "artifacts": artifacts}


@router.get("/{pipeline_id}/artifacts/{name}")
async def download_artifact(
    pipeline_id: str,
    name: str,
):
    """下载产物。"""
    output_dir = _get_output_dir(pipeline_id)

    # 安全校验：防止路径穿越
    try:
        file_path = safe_join_path(str(output_dir), name)
    except ValueError:
        raise HTTPException(400, "非法文件名")

    if not file_path.exists():
        raise HTTPException(404, "文件不存在")

    return FileResponse(
        str(file_path),
        filename=Path(name).name,
        media_type="application/octet-stream",
    )


@router.get("/{pipeline_id}/preview/{name}")
async def preview_artifact(
    pipeline_id: str,
    name: str,
):
    """预览产物（Markdown 渲染 / Excel 表格）。"""
    output_dir = _get_output_dir(pipeline_id)

    # 安全校验：防止路径穿越
    try:
        file_path = safe_join_path(str(output_dir), name)
    except ValueError:
        raise HTTPException(400, "非法文件名")

    if not file_path.exists():
        raise HTTPException(404, "文件不存在")

    if name.endswith(".md"):
        return _preview_markdown(file_path)

    if name.endswith(".xlsx"):
        return _preview_excel(file_path)

    raise HTTPException(400, "不支持预览此格式")


# ─── 预览内部实现 ───


def _preview_markdown(file_path: Path) -> dict:
    """渲染 Markdown 为 HTML（优先用 markdown 库，缺失时转义兜底）。"""
    content = file_path.read_text(encoding="utf-8")
    try:
        import markdown as md

        html = md.markdown(content, extensions=["tables", "fenced_code"])
    except ImportError:
        # fallback: HTML-escape + <pre>（防止 XSS）
        import html as html_module

        escaped = html_module.escape(content)
        html = f"<pre>{escaped}</pre>"
    return {"type": "markdown", "html": html}


def _preview_excel(file_path: Path) -> dict:
    """读取 Excel 前 N 行为二维数组返回。"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise HTTPException(500, "Excel 无活动工作表")
        rows = []
        for i, row in enumerate(
            ws.iter_rows(min_row=1, values_only=True)
        ):
            if i >= EXCEL_PREVIEW_MAX_ROWS:
                break
            rows.append([str(c or "")[:EXCEL_CELL_MAX_CHARS] for c in row])
        wb.close()
        return {"type": "excel", "rows": rows}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Excel 读取失败: {e}")


# ─── v5.0 自动化测试工程 ZIP 动态打包导出 ───


@router.get("/{pipeline_id}/export_pytest_project")
async def export_pytest_project_zip(
    pipeline_id: str,
):
    """将 Pipeline 的 testcases.json 导出为完整 PyTest 沙箱工程 ZIP。

    v5.0 核心路由：动态调用 scripts/export_pytest.py 生成完整工程目录，
    然后内存打包为 ZIP 二进制流返回给浏览器下载。

    流程：
      1. 读取 Pipeline 的 testcases.json
      2. 调用 export_pytest.export_project 生成临时工程目录
      3. 内存 ZIP 打包（io.BytesIO + zipfile）
      4. StreamingResponse 返回二进制流
    """
    import io
    import shutil
    import tempfile
    import zipfile

    from fastapi.responses import StreamingResponse

    output_dir = _get_output_dir(pipeline_id)

    # 1. 检查 testcases.json 是否存在
    json_path = output_dir / "testcases.json"
    if not json_path.exists():
        raise HTTPException(
            404,
            "testcases.json 不存在。请先完成 Step 4（用例生成）。",
        )

    # 2. 在临时目录生成完整 PyTest 工程
    #    延迟导入避免循环依赖
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from scripts.export_pytest import export_project

    with tempfile.TemporaryDirectory() as tmpdir:
        project_dir = Path(tmpdir) / "automated_test_project"
        count = export_project(
            str(json_path),
            str(project_dir),
            module_name=f"Pipeline_{pipeline_id[:8]}_Tests",
        )
        if count == 0:
            raise HTTPException(500, "工程生成失败：无可用用例")

        # 3. 内存 ZIP 打包（排除 .venv 和 __pycache__）
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for file_path in project_dir.rglob("*"):
                if file_path.is_file():
                    # 排除虚拟环境和缓存
                    rel = file_path.relative_to(project_dir)
                    if any(part in (".venv", "__pycache__", ".pytest_cache")
                           for part in rel.parts):
                        continue
                    arcname = f"automated_test_project/{rel}"
                    zf.write(file_path, arcname)

        # 4. 重置指针并返回流
        zip_buffer.seek(0)
        filename = f"automated_test_project_{pipeline_id[:8]}.zip"

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )
