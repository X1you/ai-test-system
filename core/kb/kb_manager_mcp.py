#!/usr/bin/env python3
"""
知识库管理器 - MCP 层集成
通过 MCP 协议访问 Obsidian Vault
"""

import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

# 添加 scripts 目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from mcp_client import KnowledgeItem, MCPClient

# 结构化日志 — structlog 优先，降级到 print（与 core.logger 一致）
try:
    import structlog
    _logger = structlog.get_logger("core.kb.kb_manager_mcp")
except ImportError:
    class _FallbackLogger:
        def _log(self, level, event, **kw):
            parts = [f"[{level}] [core.kb.kb_manager_mcp] {event}"]
            parts.extend(f"{k}={v}" for k, v in kw.items())
            print(" ".join(parts), file=sys.stderr)
        def debug(self, e, **k): self._log("DEBUG", e, **k)
        def info(self, e, **k): self._log("INFO", e, **k)
        def warning(self, e, **k): self._log("WARN", e, **k)
        def error(self, e, **k): self._log("ERROR", e, **k)
    _logger = _FallbackLogger()

# ============================================================================
# 配置
# ============================================================================

PYTHON_EXECUTABLE = sys.executable  # 使用当前 Python 解释器，无需绑定特定环境
# 支持通过环境变量覆盖 vault 路径（Step2KBSearch 通过 env 传递）
OBSIDIAN_VAULT = os.environ.get("OBSIDIAN_VAULT", "") or str(
    Path.home() / "Documents" / "test-interview-kb"
)

# ============================================================================
# 知识库管理器 (MCP 层)
# ============================================================================

class KnowledgeBaseManager:
    """知识库管理器 - 通过 MCP 协议访问 Obsidian Vault"""

    def __init__(self, vault_path: str = OBSIDIAN_VAULT,
                 obsidian_api_base: str = "",
                 obsidian_api_key: str = ""):
        self.mcp_client = MCPClient(
            vault_path,
            obsidian_api_base=obsidian_api_base,
            obsidian_api_key=obsidian_api_key,
        )

    def search(self, query: str, category: str = None, limit: int = 20) -> list[dict]:
        """检索知识库"""
        return self.mcp_client.search(query, category=category, limit=limit)

    def add(self, item: KnowledgeItem) -> bool:
        """添加知识条目"""
        return self.mcp_client.create_file(item)

    def ingest(self, source_file: str, category: str, module: str = "",
               project: str = "", batch: str = "") -> int:
        """回灌知识 (从 Excel 或 Markdown)

        Excel 格式自动识别:
        - testcases.xlsx 标准12列格式（用例编号|模块|功能点|维度|标题|优先级|前置条件|步骤|数据|预期|备注|结果）
        - 通用3列格式（标题|内容|标签）

        归档结构（历史用例）:
          🏆 历史用例/{项目名}/{批次日期}/TC-001 xxx.md
          如未指定 project/batch，则用文件名和当前日期兜底

        其他分类: 仍按平铺存储（业务规则/坑点等天然是独立条目）
        """
        if not os.path.exists(source_file):
            _logger.error("kb_ingest_file_not_found", source_file=source_file)
            return 0

        count = 0

        # 兜底项目名和批次名
        if not project:
            project = Path(source_file).stem  # testcases → testcases
        if not batch:
            batch = datetime.now().strftime("%Y-%m-%d")

        if source_file.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(source_file)
                ws = wb.active

                # 自动检测 Excel 格式
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                is_testcase_format = headers and headers[0] == '用例编号' and '用例标题' in headers

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 3:
                        continue

                    if is_testcase_format:
                        # 标准 testcases.xlsx 12列格式
                        tc_id = str(row[0] or '')
                        tc_module = str(row[1] or '')
                        tc_feature = str(row[2] or '')
                        tc_dimension = str(row[3] or '') if len(row) > 3 else ''
                        tc_title = str(row[4] or '') if len(row) > 4 else ''
                        tc_priority = str(row[5] or '') if len(row) > 5 else ''
                        tc_precondition = str(row[6] or '') if len(row) > 6 else ''
                        tc_steps = str(row[7] or '') if len(row) > 7 else ''
                        tc_test_data = str(row[8] or '') if len(row) > 8 else ''
                        tc_expected = str(row[9] or '') if len(row) > 9 else ''
                        tc_result = str(row[11] or '') if len(row) > 11 else ''

                        if not tc_title:
                            continue

                        # 组装结构化内容
                        content_lines = [
                            f"**用例编号**: {tc_id}",
                            f"**模块**: {tc_module}",
                            f"**功能点**: {tc_feature}",
                            f"**测试维度**: {tc_dimension}",
                            f"**优先级**: {tc_priority}",
                            "",
                            f"**前置条件**: {tc_precondition}",
                            "**测试步骤**:",
                            tc_steps,
                            "",
                            f"**测试数据**: {tc_test_data}",
                            f"**预期结果**: {tc_expected}",
                        ]
                        if tc_result:
                            content_lines.append(f"**执行结果**: {tc_result}")

                        item = KnowledgeItem(
                            id='',
                            title=f"{tc_id} {tc_title}" if tc_id else tc_title,
                            content='\n'.join(content_lines),
                            category=category,
                            module=tc_module or module,
                            tags=[t.strip() for t in [tc_dimension, tc_priority, tc_feature] if t.strip()]
                        )
                        # 历史用例按项目维度归档
                        if category == 'historical-cases':
                            # 把项目和批次信息编码到 module 字段
                            item.module = f"{project}/{batch}/{tc_module or module}"
                        if self.add(item):
                            count += 1
                    else:
                        # 通用3列格式 (title, content, tags)
                        title = str(row[0] or '')
                        content = str(row[1] or '')
                        tags = str(row[2] or '').split(',')

                        if title and content:
                            item = KnowledgeItem(
                                id='',
                                title=title,
                                content=content,
                                category=category,
                                module=module,
                                tags=[t.strip() for t in tags if t.strip()]
                            )
                            if self.add(item):
                                count += 1
            except ImportError:
                _logger.error("kb_ingest_openpyxl_missing", hint="pip install openpyxl")
            except Exception as e:
                _logger.error("kb_ingest_excel_read_failed", source_file=source_file, error=str(e))

        elif source_file.endswith('.md'):
            # Markdown 回灌
            with open(source_file, encoding='utf-8') as f:
                content = f.read()
                title = Path(source_file).stem

                item = KnowledgeItem(
                    id='',
                    title=title,
                    content=content,
                    category=category,
                    module=module,
                    tags=[category, module]
                )
                if self.add(item):
                    count = 1
        else:
            _logger.error("kb_ingest_unsupported_format", source_file=source_file)
            return 0

        _logger.info("kb_ingest_done", source_file=source_file, category=category, count=count)
        return count

    def export(self, query: str, output_file: str = "knowledge-context.md") -> str:
        """导出增强上下文 (Markdown 格式)"""
        results = self.search(query, limit=50)

        # 按分类分组
        grouped = defaultdict(list)
        for item in results:
            grouped[item['category']].append(item)

        # 生成 Markdown
        lines = [
            "# 知识库增强上下文",
            "",
            f"> 检索关键词: {query} | 命中 {len(results)} 条相关知识",
            "",
            "> 来源: Obsidian Vault (MCP 协议)",
            ""
        ]

        for category, items in grouped.items():
            cat_display = {
                'business-rules': '📋 业务规则',
                'historical-cases': '🏆 历史优质用例',
                'pitfalls': '⚠️ 线上坑点',
                'templates': '📝 用例模板',
                'data-dictionary': '📖 数据字典',
                'business-specs': '📘 业务规范',
                'team-standards': '📐 团队规范',
            }.get(category, category)

            lines.append(f"## {cat_display} ({len(items)} 条)")

            for idx, item in enumerate(items, 1):
                lines.append(f"### {idx}. [[{item['filepath']}]]")
                lines.append(f"**模块**: {item.get('module', 'N/A')}")
                # 显示标签（标签增强后新增）
                item_tags = item.get('tags', [])
                if item_tags:
                    lines.append(f"**标签**: {', '.join(f'`{t}`' for t in item_tags)}")
                if item.get('severity'):
                    lines.append(f"**严重级别**: {item['severity']}")

                # 提取内容片段 (前 500 字)
                content_preview = item['content'][:500]
                if '---' in content_preview:
                    content_preview = content_preview.split('---', 2)[-1].strip()

                lines.append(content_preview + ("..." if len(item['content']) > 500 else ""))
                lines.append("")

        # 写入文件
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        _logger.info("kb_export_done", output_path=str(output_path), query=query, results=len(results))
        return str(output_path)

    def status(self) -> dict:
        """知识库概况"""
        return self.mcp_client.status()

    def add_single(self, title: str, content: str, category: str, module: str = "",
                   tags: list[str] = None, severity: str = None) -> bool:
        """添加单条知识"""
        item = KnowledgeItem(
            id='',
            title=title,
            content=content,
            category=category,
            module=module,
            tags=tags or [],
            severity=severity
        )
        return self.add(item)


# ============================================================================
# CLI 入口
# ============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(description="知识库管理器 (MCP 层)")
    subparsers = parser.add_subparsers(dest='command', required=True)

    # search 命令
    search_parser = subparsers.add_parser('search', help='检索知识库')
    search_parser.add_argument('query', help='检索关键词')
    search_parser.add_argument('--category', choices=['business-rules', 'historical-cases', 'pitfalls', 'templates', 'data-dictionary', 'business-specs', 'team-standards'], help='限定分类')
    search_parser.add_argument('--limit', type=int, default=20, help='返回条数')

    # add 命令
    add_parser = subparsers.add_parser('add', help='添加单条知识')
    add_parser.add_argument('--title', required=True, help='标题')
    add_parser.add_argument('--content', required=True, help='内容')
    add_parser.add_argument('--category', required=True, choices=['business-rules', 'historical-cases', 'pitfalls', 'templates', 'data-dictionary', 'business-specs', 'team-standards'])
    add_parser.add_argument('--module', default='', help='所属模块')
    add_parser.add_argument('--tags', default='', help='标签 (逗号分隔)')
    add_parser.add_argument('--severity', choices=['high', 'medium', 'low'], help='严重级别 (仅坑点)')

    # ingest 命令
    ingest_parser = subparsers.add_parser('ingest', help='回灌知识文件')
    ingest_parser.add_argument('source_file', help='源文件 (Excel 或 Markdown)')
    ingest_parser.add_argument('--category', required=True, choices=['business-rules', 'historical-cases', 'pitfalls', 'templates', 'data-dictionary', 'business-specs', 'team-standards'])
    ingest_parser.add_argument('--module', default='', help='所属模块')
    ingest_parser.add_argument('--project', default='', help='项目名 (历史用例归档: 🏆 历史用例/项目名/批次/)')
    ingest_parser.add_argument('--batch', default='', help='批次名/日期 (历史用例归档)')

    # export 命令
    export_parser = subparsers.add_parser('export', help='导出增强上下文')
    export_parser.add_argument('query', help='检索关键词')
    export_parser.add_argument('--output', default='knowledge-context.md', help='输出文件路径')

    # status 命令
    subparsers.add_parser('status', help='知识库概况')

    # tags 命令 — 列出所有标签及其出现次数
    subparsers.add_parser('tags', help='列出所有关键词标签')

    args = parser.parse_args()

    # 创建管理器
    kb = KnowledgeBaseManager()

    # 执行命令
    if args.command == 'search':
        results = kb.search(args.query, category=args.category, limit=args.limit)
        print(json.dumps(results, ensure_ascii=False, indent=2))

    elif args.command == 'add':
        tags = [t.strip() for t in args.tags.split(',')] if args.tags else []
        kb.add_single(args.title, args.content, args.category, args.module, tags, args.severity)

    elif args.command == 'ingest':
        kb.ingest(args.source_file, args.category, args.module,
                  project=getattr(args, 'project', ''),
                  batch=getattr(args, 'batch', ''))

    elif args.command == 'export':
        kb.export(args.query, args.output)

    elif args.command == 'status':
        summary = kb.status()
        print(json.dumps(summary, ensure_ascii=False, indent=2))

    elif args.command == 'tags':
        # 遍历知识库，收集所有标签
        from collections import Counter
        all_files = kb.mcp_client.list_files()
        tag_counter: Counter = Counter()
        tag_files: dict = {}
        for filepath in all_files:
            file_data = kb.mcp_client.read_file(filepath)
            if not file_data:
                continue
            frontmatter = kb.mcp_client._parse_yaml_frontmatter(file_data['content'])
            file_tags = frontmatter.get('tags', [])
            if isinstance(file_tags, str):
                file_tags = [t.strip() for t in file_tags.split(',')]
            for tag in file_tags:
                tag_str = str(tag).strip()
                if tag_str:
                    tag_counter[tag_str] += 1
                    tag_files.setdefault(tag_str, []).append(filepath)
        if not tag_counter:
            print("📭 知识库中暂无标签")
        else:
            print(f"📋 知识库标签（共 {len(tag_counter)} 个）:\n")
            for tag, count in tag_counter.most_common():
                print(f"  `{tag}` ({count})")


if __name__ == '__main__':
    main()
