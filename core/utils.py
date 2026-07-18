#!/usr/bin/env python3
"""
通用工具模块

提供跨模块共享的基础工具函数。
"""

import fcntl
import os
from contextlib import contextmanager
from pathlib import Path


@contextmanager
def file_lock(lock_file: str, timeout: int = 10):
    """
    文件锁上下文管理器 — 防止并发写入冲突

    Args:
        lock_file: 锁文件路径
        timeout: 超时时间（秒），默认10秒

    Raises:
        TimeoutError: 获取锁超时
        IOError: 文件操作失败
    """
    lock_path = Path(lock_file)
    lock_path.parent.mkdir(parents=True, exist_ok=True)

    fd = None
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_WRONLY)

        try:
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
        except BlockingIOError:
            import time
            waited = 0
            while waited < timeout:
                time.sleep(0.1)
                waited += 0.1
                try:
                    fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                    break
                except BlockingIOError:
                    continue
            else:
                raise TimeoutError(f"获取文件锁超时 ({timeout}秒)，可能有其他进程正在写入")

        yield

    finally:
        if fd is not None:
            try:
                fcntl.flock(fd, fcntl.LOCK_UN)
            except Exception:
                pass
            finally:
                try:
                    os.close(fd)
                except Exception:
                    pass


def safe_join_path(base_dir: str, user_path: str) -> Path:
    """安全路径拼接 — 防止路径穿越攻击

    Args:
        base_dir: 基准目录（受信任）
        user_path: 用户输入的相对路径（不受信任）

    Returns:
        解析后的安全绝对路径

    Raises:
        ValueError: 检测到路径穿越（结果路径不在 base_dir 之下）
    """
    base = Path(base_dir).resolve()
    resolved = (base / user_path).resolve()

    # 确保 resolved 在 base 目录之下（防止路径穿越）
    if resolved != base and base not in resolved.parents:
        raise ValueError(f"路径穿越检测：{user_path} 解析后不在允许的目录范围内")

    return resolved


def excel_has_results(xlsx_path) -> bool:
    """检查 Excel 是否已填写执行结果（至少一行的"执行结果"列有值）。

    被 Pipeline._has_results 和 Step6HumanTest._check_has_results 共享，
    消除原先两处独立的重复实现。

    列识别规则：表头完全等于"执行结果"，或同时包含"执行"和"结果"。
    首行视为表头，从第二行开始检测。

    Args:
        xlsx_path: Excel 文件路径（str/Path）

    Returns:
        True 表示至少一行执行结果已填写；False 表示未填写、无执行结果列、
        文件不存在或 openpyxl 未安装（容错降级，不抛异常）。
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(xlsx_path), data_only=True)
        ws = wb.active
        if not ws:
            wb.close()
            return False

        # 定位"执行结果"列
        result_col = None
        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col).value or "").strip()
            if header == "执行结果" or ("执行" in header and "结果" in header):
                result_col = col
                break
        if not result_col:
            wb.close()
            return False

        # 统计已填写行数（首行为表头，从第二行开始）
        filled = any(
            str(ws.cell(row=row, column=result_col).value or "").strip()
            for row in range(2, ws.max_row + 1)
        )
        wb.close()
        return filled
    except Exception:
        # openpyxl 未安装或文件损坏均视为无结果
        return False
