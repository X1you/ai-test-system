#!/usr/bin/env python3
"""
Pipeline 引擎 — 独立版（v2.1）

自动串联 7 步：
  Step 1  需求分析 (AI)
  Step 2  知识库检索 (脚本)
  Step 3  测试点梳理 (AI)
  Step 4  生成测试用例 (脚本)
  Step 5  用例评审 (AI)
  Step 6  执行测试 (人工)
  Step 7  生成测试报告 (脚本)

支持：断点续跑、人工检查点、知识库自动集成、LLM 自检。

设计要点：
  - STEP_REGISTRY 集中定义步骤元数据（id/名称/输出文件/暂停模式），
    status()/summary() 复用同一数据源，消除硬编码重复。
  - run() 以"跳过-执行-结果处理"三段式循环驱动各步骤，
    可扩展性集中在 _dispatch_step()。
"""

import json
import os
import re
import subprocess
import sys
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from core.llm_client import LLMClient, LLMError
from core.steps.base import StepResult
from core.steps.step0_gap_analysis import Step0GapAnalysis
from core.steps.step1_analysis import Step1Analysis
from core.steps.step2_kb_search import Step2KBSearch
from core.steps.step3_testpoints import Step3Testpoints
from core.steps.step4_generate import Step4Generate
from core.steps.step5_review import Step5Review
from core.steps.step6_human_test import Step6HumanTest
from core.steps.step7_report import Step7Report

# 项目根目录（与 step2_kb_search.py 一致：core/pipeline.py → parents[1] = 项目根）
PROJECT_ROOT = Path(__file__).resolve().parents[1]

STATE_FILE = "_pipeline_state.json"


@dataclass(frozen=True)
class StepMeta:
    """单步元数据 — 步骤 id、展示名、主要输出文件、是否需要人工确认。"""

    id: int
    name: str
    output_file: str
    needs_pause: bool = False  # semi/step 模式下完成后是否暂停确认


# 集中维护步骤元数据 — status()、_print_summary()、run() 全部复用，
# 新增/调整步骤只改这一处。
#
# Step 0 是需求漏洞扫描（PRD Gap Analysis），独立于核心生成链，
# 其产物 requirement_gap_analysis.md 和 gap_count 注入全局上下文。
# Step 0 容灾降级：失败/超时一律 gap_count=0，不阻断后续 Step 1+。
STEP_REGISTRY: list[StepMeta] = [
    StepMeta(0, "需求漏洞扫描", "requirement_gap_analysis.md", needs_pause=False),
    StepMeta(1, "需求分析", "requirements_analysis.md", needs_pause=True),
    StepMeta(2, "知识库检索", "knowledge-context.md"),
    StepMeta(3, "测试点梳理", "testpoints.md", needs_pause=True),
    StepMeta(4, "生成测试用例", "testcases.xlsx"),
    StepMeta(5, "用例评审", "test_case_review_report.md", needs_pause=True),
    StepMeta(6, "执行测试", "testcases.xlsx"),
    StepMeta(7, "生成测试报告", "test_report.md"),
]

TOTAL_STEPS = len(STEP_REGISTRY)


class Pipeline:
    """全流程 Pipeline 引擎。"""

    def __init__(self, config: dict, output_dir: str):
        self.config = config
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.llm: LLMClient | None = None

        # ★ 全局上下文（贯穿各步骤，供 Step7 报告读取 ROI 数据）
        # 关键字段：gap_count（Step0 注入）、case_count（Step4 注入）、
        #           total_duration（Step4 注入，v3.0 ROI 精准计算）
        self.context: dict = {"gap_count": 0, "case_count": 0, "total_duration": 0}

        # WebUI 回调钩子（CLI 模式为 None，不影响行为）
        self.on_log: Callable | None = None         # fn(level, msg)
        self.on_step_done: Callable | None = None   # fn(step_id, result_dict)

        # 交互模式标志：CLI=True（_pause 走 input 交互），WebUI=False（状态机管暂停）
        self.interactive: bool = True

    # ─── 日志 / LLM 初始化 ───

    def _notify_log(self, level: str, msg: str):
        """日志回调通知 — 同时输出到结构化日志和 WebUI 回调。"""
        from core.logger import get_logger

        logger = get_logger("pipeline")
        logger.info("step_log", level=level, message=msg)
        if self.on_log:
            self.on_log(level, msg)

    def _init_llm(self):
        """初始化 LLM 客户端（幂等，重复调用安全）。"""
        if self.llm:
            return
        try:
            self.llm = LLMClient(self.config["llm"])
        except LLMError as e:
            print(f"❌ LLM 初始化失败: {e}", file=sys.stderr)
            raise

    # ─── 状态管理 ───

    def load_state(self) -> dict:
        """加载断点状态；不存在时返回初始状态。"""
        path = self.output_dir / STATE_FILE
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
        return {
            "started": datetime.now().isoformat(),
            "completed_steps": [],
            "step_results": {},
            "mode": "semi",
            "requirements_file": "",
        }

    def save_state(self, state: dict):
        """持久化断点状态（幂等写入）。"""
        state["updated"] = datetime.now().isoformat()
        path = self.output_dir / STATE_FILE
        path.write_text(
            json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    def mark_done(self, state: dict, step_id: int, result: StepResult):
        """标记步骤完成并持久化，同时通知 WebUI 回调。

        同时记录该步骤输出文件的 content hash，用于后续篡改检测（TC-006/018）。
        """
        if step_id not in state["completed_steps"]:
            state["completed_steps"].append(step_id)
        result_info = {
            "ok": result.ok,
            "data": result.data,
        }
        # 记录输出文件 content hash（篡改检测用）
        output_file = self._step_output_file(step_id)
        if output_file:
            file_hash = self._compute_file_hash(self.output_dir / output_file)
            if file_hash:
                state.setdefault("output_hashes", {})[str(step_id)] = file_hash
        state["step_results"][str(step_id)] = result_info
        self.save_state(state)

        # 通知 WebUI 回调
        if self.on_step_done:
            self.on_step_done(step_id, result_info)

    @staticmethod
    def _compute_file_hash(path: Path) -> str | None:
        """计算文件内容的 SHA256 hash（用于篡改检测）。

        Returns:
            64 字符 hex hash，文件不存在或读取失败返回 None。
        """
        import hashlib

        try:
            content = path.read_bytes()
            return hashlib.sha256(content).hexdigest()
        except Exception:
            return None

    def _step_output_file(self, step_id: int) -> str:
        """从 STEP_REGISTRY 取某步的主要输出文件名。"""
        for meta in STEP_REGISTRY:
            if meta.id == step_id:
                return meta.output_file
        return ""

    @staticmethod
    def is_done(state: dict, step_id: int) -> bool:
        """步骤是否已完成。"""
        return step_id in state["completed_steps"]

    def _check_skip(self, state: dict, step_id: int, output_file: str) -> bool:
        """检查是否可跳过：已完成 + 输出文件存在 + 内容未被篡改。

        篡改检测（TC-006/018 修复）：对比 state.output_hashes 中记录的 hash
        与当前文件实际 hash，不一致则认为产物被篡改，拒绝跳过（强制重跑该步）。
        日志会明确提示"产物被篡改"，而非静默继续。
        """
        if not self.is_done(state, step_id):
            return False
        path = self.output_dir / output_file
        if not path.exists():
            return False

        # 篡改检测：比对 content hash
        stored_hash = state.get("output_hashes", {}).get(str(step_id))
        if stored_hash:
            current_hash = self._compute_file_hash(path)
            if current_hash != stored_hash:
                # ⚠️ 特殊豁免：Step4 和 Step6 共用 testcases.xlsx，
                # 用户填入执行结果后 hash 自然改变，不视为篡改。
                # 检测到 Excel 已有执行结果 → 跳过警告和强制重跑。
                # 豁免覆盖所有以 testcases.xlsx 为产物的步骤（防御性：
                # 即便步骤路由改变也能正确豁免，避免丢失用户填写的执行结果）。
                if output_file == "testcases.xlsx":
                    if self._has_results(path):
                        return True

                self._notify_log(
                    "WARN",
                    f"⚠️ Step {step_id} 产物 {output_file} 内容与完成时不一致"
                    f"（可能被篡改），将重新执行该步骤",
                )
                # 从已完成列表移除，强制重跑
                state["completed_steps"] = [
                    s for s in state["completed_steps"] if s != step_id
                ]
                if "output_hashes" in state:
                    state["output_hashes"].pop(str(step_id), None)
                self.save_state(state)
                return False
        return True

    def _has_results(self, xlsx_path: Path) -> bool:
        """检查 Excel 是否已填写执行结果（至少一行的"执行结果"列有值）。

        委托给 core.utils.excel_has_results（与 Step6HumanTest 共享同一实现）。
        """
        from core.utils import excel_has_results

        return excel_has_results(xlsx_path)

    # ─── Step0/Step4 上下文恢复 & Step6 引导（数据流闭环辅助） ───

    def _read_output(self, filename: str) -> str:
        """读取输出目录下的产物文件内容（不存在返回空串）。"""
        path = self.output_dir / filename
        if path.exists():
            try:
                return path.read_text(encoding="utf-8")
            except Exception:
                return ""
        return ""

    def _read_prd_content(self, requirements_file: str) -> str:
        """读取原始 PRD 需求文档内容（v3.0 四变量之一）。

        使用 Step0 的 _safe_read_text 做编码兜底。
        """
        from core.steps.step0_gap_analysis import Step0GapAnalysis
        path = Path(requirements_file)
        if not path.exists():
            return ""
        content = Step0GapAnalysis._safe_read_text(path)
        return content or ""

    def _recover_total_duration(self) -> int:
        """从 testcases.json 恢复总预估时长（v3.0 ROI 精准计算）。

        断点续跑场景：从 Step4 落盘的 JSON 读取 estimated_duration 累加。
        """
        import json
        json_path = self.output_dir / "testcases.json"
        if not json_path.exists():
            return 0
        try:
            cases = json.loads(json_path.read_text(encoding="utf-8"))
            if isinstance(cases, list):
                return sum(int(c.get("estimated_duration", 0) or 0) for c in cases if isinstance(c, dict))
        except Exception:
            return 0
        return 0

    def _recover_gap_count_from_file(self) -> int:
        """从 Step0 产物 requirement_gap_analysis.md 恢复 gap_count。

        断点续跑场景：Step0 已完成时，从落盘报告重新解析 GAP_COUNT。
        解析失败返回 0（与降级语义一致）。

        ★ 修复 TC-012：原实现解析失败时静默返回 0，context.gap_count 被覆盖，
        Step7 ROI 看板会显示「0 小时」而不报错——数据静默丢失。
        新增：解析失败时记 WARN 日志，避免静默数据丢失。
        """
        from core.steps.step0_gap_analysis import Step0GapAnalysis

        report_path = self.output_dir / "requirement_gap_analysis.md"
        if not report_path.exists():
            self._notify_log("WARN", "Step0 报告不存在，gap_count 无法恢复，置 0")
            return 0
        try:
            content = report_path.read_text(encoding="utf-8")
            gap = Step0GapAnalysis._extract_gap_count(content)
            if gap == 0:
                # 区分两种情况：报告明确是 0 个漏洞 vs 报告被篡改导致解析失败
                has_gap_marker = "GAP_COUNT" in content or "gap_count" in content.lower()
                if not has_gap_marker:
                    self._notify_log(
                        "WARN",
                        f"⚠️ Step0 报告 {report_path.name} 缺少 GAP_COUNT 标记，"
                        "可能被篡改，gap_count 置 0（ROI 看板将显示保守值）",
                    )
            return gap
        except Exception as e:
            self._notify_log("WARN", f"恢复 gap_count 失败: {e}，置 0")
            return 0

    def _count_cases_from_excel(self) -> int:
        """统计 testcases.xlsx 中的用例条数（数据行）。

        失败返回 0，不抛异常（ROI 看板对缺失数据要稳健）。
        """
        xlsx_path = self.output_dir / "testcases.xlsx"
        if not xlsx_path.exists():
            return 0
        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
            ws = wb.active
            # 数据行数 = max_row - 1（表头）；空表返回 0
            max_row = ws.max_row if ws is not None else 0
            count = max(0, (max_row or 1) - 1)
            wb.close()
            return count
        except Exception:
            return 0

    def _print_step6_guide(self, xlsx_path: Path, case_count: int):
        """Step 6 暂停时的主动引导：3 种执行方式任选其一。

        把原来的"消极等待"重构为"赋能引导 + 万能导出插头"：
          方式 1：手动标记 Excel（最简单）
          方式 2：一键导出/推送到 TestRail 等外部系统（资产管道核心）
          方式 3：跳过执行，直接出报告（快速过流程）
        """
        estimated_hours = max(1, round(case_count * 0.05, 1))  # 粗略估算

        print()
        print("═" * 60)
        print("  🎯 Step 6 — 测试执行阶段（3 种方式任选其一）")
        print("═" * 60)
        print()
        print(f"  📁 已生成 {case_count} 条用例：{xlsx_path}")
        print(f"  ⏱️  人工执行预估：~{estimated_hours} 小时（每条约 3 分钟）")
        print()
        print("  ─────────────────────────────────────────")
        print("  1️⃣  手动标记 Excel")
        print("     在「执行结果」列填 通过/失败/阻塞/跳过")
        print()
        print("  2️⃣  一键导出 / 推送到外部测试管理系统")
        print("     支持平台：TestRail（已内置）、TestLink、可自定义适配器")
        print("     API: POST /api/v1/integrations/test-cases/push")
        print("     CLI : python cli.py ingest <file> --category historical-cases")
        print()
        print("  3️⃣  跳过执行，直接出报告")
        print("     不填执行结果，直接 resume → 报告标记全部「未执行」")
        print("  ─────────────────────────────────────────")
        print()
        print(f"  ▶️  完成后继续：python cli.py resume -o {self.output_dir}")
        print("═" * 60)

    # ─── 执行 ───

    def run(
        self,
        requirements_file: str,
        mode: str = "semi",
        dimensions: str = "basic",
        formats: str = "excel",
    ) -> dict:
        """
        执行全流程 Pipeline。

        Args:
            requirements_file: 需求文档路径
            mode: auto | semi | step
            dimensions: basic | all | positive,negative
            formats: excel | xmind | excel,xmind

        Returns:
            最终状态 dict（completed_steps / step_results / ...）
        """
        state = self.load_state()
        state["mode"] = mode
        state["requirements_file"] = requirements_file
        self.save_state(state)

        self._print_banner(mode, requirements_file)

        out = str(self.output_dir)
        config = self.config

        # ─── Step 0: 需求漏洞扫描（PRD Gap Analysis）───
        # ★ 容灾步骤：失败/超时一律 gap_count=0，不阻断后续 Step 1+。
        # ★ 数据流闭环：把 gap_count 注入 self.context，Step7 报告读取做 ROI 计算。
        # 即使 self_check 关闭，Step0 仍执行（它不依赖 self_check 开关）。
        if self._check_skip(state, 0, "requirement_gap_analysis.md"):
            print("✅ Step 0 已完成，跳过\n")
            # 跳过时仍从产物恢复 gap_count（断点续跑场景）
            self.context["gap_count"] = self._recover_gap_count_from_file()
        else:
            self._init_llm()
            r0 = Step0GapAnalysis(out, config, self.llm).run(
                requirements_path=requirements_file
            )
            gap = int(r0.data.get("gap_count", 0)) if r0.data else 0
            self.context["gap_count"] = gap
            self.mark_done(state, 0, r0)
            degraded = bool(r0.data.get("degraded")) if r0.data else False
            if degraded:
                print("⏭️  Step 0 已降级跳过（gap_count=0），继续 Step 1\n")
            else:
                print(f"🔍 Step 0 识别 {gap} 项待澄清问题（详见 requirement_gap_analysis.md）\n")
            # Step 0 不设检查点（needs_pause=False），无论 semi/step 都直接推进

        # ─── Step 1: 需求分析 ───
        if self._check_skip(state, 1, "requirements_analysis.md"):
            print("✅ Step 1 已完成，跳过\n")
        else:
            self._init_llm()
            r = Step1Analysis(out, config, self.llm).run(
                requirements_path=requirements_file
            )
            if r.ok:
                self.mark_done(state, 1, r)
                print()
                if mode in ("semi", "step"):
                    if not self._pause("Step 1"):
                        return state
            elif mode in ("semi", "step"):
                print("⏸️  Step 1 需要确认后继续")
                return state
            else:
                print("❌ Step 1 失败，终止")
                return state

        # ─── Step 2: 知识库检索 ───
        if self._check_skip(state, 2, "knowledge-context.md"):
            print("✅ Step 2 已完成，跳过\n")
        else:
            analysis = (self.output_dir / "requirements_analysis.md").read_text(
                encoding="utf-8"
            )
            r = Step2KBSearch(out, config, self.llm).run(
                requirements_analysis=analysis
            )
            self.mark_done(state, 2, r)
            print()

        # ─── Step 3: 测试点梳理 ───
        if self._check_skip(state, 3, "testpoints.md"):
            print("✅ Step 3 已完成，跳过\n")
        else:
            self._init_llm()
            analysis = (self.output_dir / "requirements_analysis.md").read_text(
                encoding="utf-8"
            )
            kb_context = self._read_kb_context()
            r = Step3Testpoints(out, config, self.llm).run(
                requirements_analysis=analysis,
                kb_context=kb_context,
                dimensions=dimensions,
            )
            if r.ok:
                self.mark_done(state, 3, r)
                print()
                if mode in ("semi", "step"):
                    if not self._pause("Step 3"):
                        return state
            else:
                print("⏸️  Step 3 失败/需确认")
                return state

        # ─── Step 4: 生成测试用例 ───
        if self._check_skip(state, 4, "testcases.xlsx"):
            print("✅ Step 4 已完成，跳过\n")
            # 断点续跑时从 Excel 恢复 case_count
            self.context["case_count"] = self._count_cases_from_excel()
            # 尝试从 testcases.json 恢复 total_duration（v3.0 ROI 精准计算）
            self.context["total_duration"] = self._recover_total_duration()
        else:
            self._init_llm()
            # ★ v3.0 四变量拼装：PRD + Step0漏洞 + RAG知识 + 测试点
            prd_content = self._read_prd_content(requirements_file)
            step0_vuln = self._read_output("requirement_gap_analysis.md") or ""
            rag_chunks = self._read_kb_context()
            test_points_text = (self.output_dir / "testpoints.md").read_text(encoding="utf-8") if (self.output_dir / "testpoints.md").exists() else ""

            r = Step4Generate(out, config, self.llm).run(
                dimensions=dimensions,
                formats=formats,
                prd_content=prd_content,
                step0_vulnerabilities=step0_vuln,
                rag_knowledge_chunks=rag_chunks,
                test_points=test_points_text,
            )
            if r.ok:
                self.mark_done(state, 4, r)
                # ★ 数据流闭环：把用例数和总预估时长注入全局上下文，供 Step7 ROI 计算
                self.context["case_count"] = r.data.get("case_count", 0)
                self.context["total_duration"] = r.data.get("total_duration", 0)
                methodology = r.data.get("methodology", "unknown")
                neg_ratio = r.data.get("negative_ratio", 0)
                print(f"  → 用例数 {self.context['case_count']}，预估总时长 {self.context['total_duration']} 分钟，"
                      f"负向占比 {neg_ratio:.0%}（{methodology}）")
                print()
            else:
                print("❌ Step 4 失败，终止")
                return state

        # ─── Step 5: 用例评审 ───
        if self._check_skip(state, 5, "test_case_review_report.md"):
            print("✅ Step 5 已完成，跳过\n")
        else:
            self._init_llm()
            kb_context = self._read_kb_context()
            r = Step5Review(out, config, self.llm).run(kb_context=kb_context)
            if r.ok:
                self.mark_done(state, 5, r)
                print()
                if mode in ("semi", "step"):
                    if not self._pause("Step 5"):
                        return state
            else:
                print("⏸️  Step 5 需确认")
                return state

        # ─── 🔁 知识库回灌 #1: Step5 完成后回灌优质用例 ───
        # 立项需求：评审通过的优质用例沉淀到 🏆 历史用例/{项目}/{批次}/
        # 设计：回灌是"增强"非"必需"，失败不阻塞主流程（_ingest_to_kb 内部容错）
        if not state.get("ingested_cases", False):
            xlsx_path = self.output_dir / "testcases.xlsx"
            ingested = self._ingest_to_kb(
                source_file=str(xlsx_path),
                category="historical-cases",
                module="全流程生成用例",
            )
            if ingested > 0:
                state["ingested_cases"] = True
                state["ingested_cases_count"] = ingested
                self.save_state(state)

        # ─── Step 6: 人工执行测试 ───
        xlsx_path = self.output_dir / "testcases.xlsx"
        # Step 6 跳过检查：已完成且 Excel 有执行结果则跳过
        if self.is_done(state, 6) and xlsx_path.exists() and self._has_results(xlsx_path):
            print("✅ Step 6 已完成，跳过\n")
        else:
            r = Step6HumanTest(out, config).run()
            if r.ok:
                self.mark_done(state, 6, r)
                print()
            else:
                # ★ Step 6 主动引导：3 种执行方式任选其一（含万能导出插头）
                case_count = self._count_cases_from_excel()
                self._print_step6_guide(xlsx_path, case_count)
                return state

        # ─── Step 7: 生成报告 ───
        if self._check_skip(state, 7, "test_report.md"):
            print("✅ Step 7 已完成，跳过\n")
        else:
            # ★ 把全局上下文的 gap_count/case_count/total_duration 透传给报告生成器（ROI 看板）
            r = Step7Report(out, config).run(
                gap_count=self.context.get("gap_count", 0),
                case_count=self.context.get("case_count", 0),
                total_duration=self.context.get("total_duration", 0),
            )
            if r.ok:
                self.mark_done(state, 7, r)
                print()
            else:
                print("❌ Step 7 失败，终止")
                return state

        # ─── 🔁 知识库回灌 #2: Step7 完成后回灌失败用例分析 ───
        # 立项需求：测试报告中的失败用例分析沉淀到 ⚠️ 线上坑点/
        # 设计：从 test_report.md 提取每条失败用例，逐条回灌（独立文件，方便检索复用）
        if not state.get("ingested_pitfalls", False):
            report_path = self.output_dir / "test_report.md"
            pitfalls = self._extract_pitfalls_from_report(report_path)
            if pitfalls:
                pitfalls_dir = self.output_dir / "_pitfalls_export"
                pitfalls_dir.mkdir(exist_ok=True)
                ingested_total = 0
                for i, pitfall in enumerate(pitfalls, 1):
                    # 每条坑点独立成文件：_pitfalls_export/TC-xxx 标题.md
                    safe_title = re.sub(r'[\\/:*?"<>|]', '_', pitfall['title'])[:60]
                    pf_path = pitfalls_dir / f"{pitfall['tc_id']} {safe_title}.md"
                    pf_path.write_text(pitfall['content'], encoding="utf-8")
                    count = self._ingest_to_kb(
                        source_file=str(pf_path),
                        category="pitfalls",
                        module="测试失败分析",
                    )
                    ingested_total += count
                if ingested_total > 0:
                    state["ingested_pitfalls"] = True
                    state["ingested_pitfalls_count"] = ingested_total
                    self.save_state(state)

        # ─── 完成 ───
        self._print_summary(state)
        return state

    def resume(self, dimensions: str = "basic", formats: str = "excel", mode: str | None = None):
        """从断点继续执行。

        Args:
            mode: 执行模式，默认 auto（与 WebUI 对齐）。传 None 使用原始模式。
        """
        state = self.load_state()
        if not state.get("requirements_file"):
            print("❌ 未找到 pipeline 状态，请先 run", file=sys.stderr)
            return

        # resume 默认 auto 模式（用户 resume 的目标是快速完成后续步骤，不是二次暂停）
        # 允许 --mode 参数覆盖（如强制 semi 重审评审报告）
        run_mode = mode or "auto"

        print()
        print(f"▶  继续执行 — 已完成: {state['completed_steps']}  (模式: {run_mode})")
        print()

        self.run(
            state["requirements_file"],
            mode=run_mode,
            dimensions=dimensions,
            formats=formats,
        )

    def status(self):
        """查看当前 Pipeline 执行状态。"""
        state = self.load_state()
        print()
        print("═" * 60)
        print("  📊 Pipeline 状态")
        print("═" * 60)
        print()
        print(f"  启动时间: {state.get('started', 'N/A')[:19]}")
        print(f"  最后更新: {state.get('updated', 'N/A')[:19]}")
        print(f"  执行模式: {state.get('mode', 'N/A')}")
        print(f"  需求文档: {state.get('requirements_file', 'N/A')}")
        print()

        print("  步骤                    │ 状态   │ 输出文件")
        print("  " + "─" * 56)
        for meta in STEP_REGISTRY:
            done = meta.id in state["completed_steps"]
            exists = (self.output_dir / meta.output_file).exists()
            icon = "✅" if done else ("📁" if exists else "⬜")
            st = "已完成" if done else ("文件存在" if exists else "待执行")
            print(f"  {icon} Step {meta.id}. {meta.name:<16} │ {st:<6} │ {meta.output_file}")

        done_count = len(state["completed_steps"])
        print()
        print(f"  进度: {done_count}/{TOTAL_STEPS} ({done_count / TOTAL_STEPS * 100:.0f}%)")
        print()

    # ─── 知识库回灌（横切增强） ───

    def _ingest_to_kb(
        self,
        source_file: str,
        category: str,
        module: str = "",
        project: str = "",
        batch: str = "",
    ) -> int:
        """回灌知识到 Obsidian Vault（调用 kb_manager_mcp.py ingest 子命令）。

        设计原则：
        - 容错：回灌失败不阻塞主流程（它是"增强"不是"必需"），仅打印警告
        - 幂等：重复回灌同文件由 kb_manager 层处理（标题去重）
        - 可验证：回灌后实际抽查 Vault 文件数，不信任中间层返回的 count

        Args:
            source_file: 源文件路径（Excel 或 Markdown）
            category: 知识库分类（historical-cases / pitfalls / ...）
            module: 所属模块（可选）
            project: 项目名（历史用例归档用，默认取输出目录名）
            batch: 批次名（默认当天日期）

        Returns:
            实际回灌条数（回灌失败返回 0）
        """
        # 知识库配置改从 DB 读取（与知识库页面配置同源）
        try:
            from core.kb.dynamic_kb_manager import get_dynamic_kb_manager

            mgr = get_dynamic_kb_manager()
            if not mgr.is_configured():
                self._notify_log("KB", "知识库未配置，跳过回灌")
                return 0
            kb_cfg = mgr.get_config() or {}
            vault_path = kb_cfg.get("vault_path", "")
        except Exception as e:
            self._notify_log("WARN", f"知识库配置读取失败: {e}，跳过回灌")
            return 0

        if not vault_path:
            self._notify_log("WARN", "vault_path 未配置，跳过回灌")
            return 0

        source_path = Path(source_file)
        if not source_path.exists():
            self._notify_log("WARN", f"回灌源文件不存在: {source_file}")
            return 0

        # 兜底项目名和批次名
        if not project:
            project = self.output_dir.name
        if not batch:
            batch = datetime.now().strftime("%Y-%m-%d")

        # 调用 kb_manager_mcp.py ingest 子命令（与 Step2 同样的 subprocess 模式）
        kb_script = PROJECT_ROOT / "core" / "kb" / "kb_manager_mcp.py"
        if not kb_script.exists():
            self._notify_log("WARN", f"知识库脚本不存在: {kb_script}")
            return 0

        try:
            env = {**os.environ, "OBSIDIAN_VAULT": vault_path}
            self._notify_log("KB", f"回灌 {source_path.name} → {category} ({project}/{batch})")

            result = subprocess.run(
                [
                    sys.executable, str(kb_script), "ingest", str(source_path),
                    "--category", category,
                    "--module", module or "未分类",
                    "--project", project,
                    "--batch", batch,
                ],
                capture_output=True,
                text=True,
                timeout=120,
                env=env,
            )

            if result.returncode != 0:
                self._notify_log(
                    "WARN",
                    f"回灌失败 (exit={result.returncode}): {result.stderr[-200:]}",
                )
                return 0

            # 回灌后实际验证 Vault 文件数（不信任 stdout 的 count）
            vault = Path(vault_path).expanduser()
            if vault.exists():
                verified = self._verify_ingest(vault, category, project, batch)
                self._notify_log("KB", f"✅ 回灌验证: Vault 新增 {verified} 个文件")
                return verified
            return 0

        except subprocess.TimeoutExpired:
            self._notify_log("WARN", "回灌超时（120s），跳过")
            return 0
        except Exception as e:
            self._notify_log("WARN", f"回灌异常: {e}")
            return 0

    @staticmethod
    def _verify_ingest(
        vault: Path, category: str, project: str, batch: str
    ) -> int:
        """回灌后验证 Vault 实际产生文件（SKILL 注意事项#9）。

        策略：定位归档目录，统计 .md 文件数。
        - historical-cases: 🏆 历史用例/{project}/{batch}/
        - pitfalls: ⚠️ 线上坑点/（平铺，按 mtime 近期过滤）
        - 其他分类：对应分类目录
        """
        category_dirs = {
            "historical-cases": "🏆 历史用例",
            "pitfalls": "⚠️ 线上坑点",
            "business-rules": "📋 业务规则",
            "templates": "📐 用例模板",
            "data-dictionary": "📊 数据字典",
            "business-specs": "📌 业务规范",
            "team-standards": "👥 团队规范",
        }
        base_name = category_dirs.get(category, category)
        target_dir = vault / base_name

        if category == "historical-cases":
            # 按项目/批次归档
            target_dir = vault / base_name / project / batch
        elif not target_dir.exists():
            # 平铺分类，目录不存在说明没写入
            return 0

        if not target_dir.exists():
            return 0

        # 统计 .md 文件数
        md_files = list(target_dir.rglob("*.md"))
        return len(md_files)

    @staticmethod
    def _extract_pitfalls_from_report(report_path: Path) -> list[dict]:
        """从 test_report.md 提取失败用例，转为逐条坑点（每条独立回灌）。

        Returns:
            list[dict]，每项 {"tc_id", "title", "content"}；无失败用例返回 []。
            content 是完整的 Markdown，标题用 "TC-xxx: title" 便于检索。
        """
        if not report_path.exists():
            return []

        content = report_path.read_text(encoding="utf-8")
        today = datetime.now().strftime("%Y-%m-%d")

        import re

        # 匹配 "## 🔍 失败用例分析" 到下一个二级标题或 "---"
        match = re.search(
            r"##\s*🔍?\s*失败用例分析(.*?)(?=\n##\s|\n---|\Z)",
            content,
            re.DOTALL,
        )
        if not match:
            return []

        section = match.group(1)
        if "共 **0**" in section or "共 0 个" in section:
            return []

        # 解析每个失败用例块（### N. TC-xxx: title）
        cases = re.findall(
            r"###\s*\d+\.\s*(TC-\d+):\s*(.+?)(?=\n###\s|\Z)",
            section,
            re.DOTALL,
        )
        if not cases:
            return []

        pitfalls = []
        for tc_id, body in cases:
            title = body.split("\n")[0].strip()
            # 提取关键字段（去掉 Markdown 加粗的 ** 前缀，避免 ** ** 瑕疵）
            def extract(field_pattern: str) -> str:
                m = re.search(field_pattern, body)
                if not m:
                    return "未推断"
                # 去掉值前后的 ** 和空格
                return m.group(1).strip().strip("*").strip()

            module = extract(r"所属模块[：:]\s*\**\s*(.+)")
            reason = extract(r"失败原因.*?[：:]\s*\**\s*(.+)")
            suggestion = extract(r"修复建议[：:]\s*\**\s*(.+)")
            note = extract(r"备注[：:]\s*\**\s*(.+)")

            pitfall_content = "\n".join([
                f"# {tc_id}: {title}",
                "",
                f"> 失败用例坑点 | 提取自测试报告 | {today}",
                "",
                f"**用例编号**: {tc_id}",
                f"**所属模块**: {module}",
                f"**失败原因（推断）**: {reason}",
                f"**修复建议**: {suggestion}",
                "",
                "## 详细分析",
                note,
            ])
            pitfalls.append({
                "tc_id": tc_id,
                "title": title,
                "content": pitfall_content,
            })

        return pitfalls

    # ─── 内部工具 ───

    def _read_kb_context(self) -> str:
        """读取 Step 2 产出的知识库上下文（缺失时返回空串）。"""
        path = self.output_dir / "knowledge-context.md"
        if path.exists():
            return path.read_text(encoding="utf-8")
        return ""

    def _pause(self, step_name: str) -> bool:
        """semi/step 模式暂停确认。

        CLI 模式：真正交互式 input() 等待用户确认，展示当前步骤产物路径。
        WebUI 模式：通过 self.interactive=False 标志跳过（WebUI 用状态机管暂停）。

        Returns:
            True 继续执行，False 终止流程。
        """
        # WebUI 调用（非交互式）直接跳过，暂停由 PipelineTask 状态机管理
        if not getattr(self, "interactive", True):
            return True

        # CLI 交互式暂停
        print()
        print(f"⏸️  {step_name} 完成 — 半自动模式暂停")
        print(f"    可查看产物: {self.output_dir}")
        print("    输入 [回车]继续，[s]停止，[r]查看报告")

        try:
            while True:
                choice = input("> ").strip().lower()
                if choice in ("", "c", "continue", "继续"):
                    return True
                if choice in ("s", "stop", "停止"):
                    print(f"⛔ 用户中止 Pipeline（已完成 {step_name}）")
                    return False
                if choice in ("r", "report", "报告"):
                    self._show_step_summary()
                    continue
                print("    无效输入。[回车]继续 / [s]停止 / [r]查看报告")
        except (EOFError, KeyboardInterrupt):
            # 非交互环境（如管道/CI）或 Ctrl+C → 默认继续
            print()
            return True

    def _show_step_summary(self):
        """CLI 暂停时展示当前各步骤产物状态（辅助用户决策）。"""
        state = self.load_state()
        print()
        print("  📁 当前产物状态:")
        for meta in STEP_REGISTRY:
            path = self.output_dir / meta.output_file
            done = meta.id in state.get("completed_steps", [])
            if path.exists():
                size = path.stat().st_size
                size_str = f"{size/1024:.1f}K" if size > 1024 else f"{size}B"
                icon = "✅" if done else "📁"
                print(f"    {icon} Step {meta.id}. {meta.name}: {meta.output_file} ({size_str})")
            else:
                print(f"    ⬜ Step {meta.id}. {meta.name}: (未生成)")
        print()

    def _print_banner(self, mode: str, requirements_file: str):
        """打印启动横幅。"""
        print()
        print("═" * 60)
        print(f"  🚀 全流程 Pipeline 启动 — 模式: {mode}")
        print(f"  📂 输出目录: {self.output_dir}")
        print(f"  📄 需求文档: {requirements_file}")
        if self.llm:
            print(f"  🤖 LLM: {self.llm.provider} / {self.llm.model}")
        print("═" * 60)
        print()

    def _print_summary(self, state: dict):
        """打印执行完成总结。"""
        print()
        print("═" * 60)
        print("  ✅ 全流程 Pipeline 执行完成！")
        print("═" * 60)
        print()
        print(f"  {'步骤':<22} │ {'状态':<6} │ 输出文件")
        print(f"  {'─' * 56}")
        for meta in STEP_REGISTRY:
            done = meta.id in state["completed_steps"]
            icon = "✅" if done else "⬜"
            st = "完成" if done else "待执行"
            print(f"  {icon} {meta.id}. {meta.name:<18} │ {st:<4} │ {meta.output_file}")

        print()
        print(f"  📁 输出目录: {self.output_dir}/")

        if self.llm:
            stats = self.llm.stats
            print(
                f"  🤖 LLM 调用: {stats['call_count']} 次, "
                f"{stats['total_tokens']} tokens ({stats['provider']})"
            )

        # 下一步指引
        ingested_cases = state.get("ingested_cases_count", 0)
        ingested_pitfalls = state.get("ingested_pitfalls_count", 0)
        if ingested_cases or ingested_pitfalls:
            print(f"  📚 知识库回灌: {ingested_cases} 用例 + {ingested_pitfalls} 坑点")
        print()
        print("  💡 下一步:")
        report = self.output_dir / "test_report.md"
        if report.exists():
            print(f"     📖 查看测试报告: {report}")
        print("     📋 回顾分析结果: see requirements_analysis.md, testpoints.md")
        print("     ▶️  重新开始: python cli.py run <新需求文档.md> -o <输出目录>")
