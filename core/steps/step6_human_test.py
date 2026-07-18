#!/usr/bin/env python3
"""
Step 6: 人工执行测试（人工步骤）

不自动执行，只检测 Excel 是否已填写执行结果。
未填写时自动生成测试执行指引（execution_guide.md），帮助用户高效执行。
"""

from pathlib import Path

from core.steps.base import BaseStep, StepResult


class Step6HumanTest(BaseStep):
    step_id = 6
    step_name = "执行测试（人工）"
    output_file = "testcases.xlsx"

    def run(self, **kwargs) -> StepResult:  # type: ignore[override]
        """检测 Excel 是否已填写执行结果，未填写时自动生成执行指引。"""
        self.log(f"Step {self.step_id}/7: {self.step_name}", "HUMAN")

        xlsx_path = self._out("testcases.xlsx")

        if self._check_has_results(str(xlsx_path)):
            self.log("检测到执行结果已填写，继续下一步", "OK")
            return StepResult(ok=True)

        # 未填写 → 自动生成执行指引，帮助用户高效执行
        guide_path = self._out("execution_guide.md")
        case_count = self._generate_execution_guide(str(xlsx_path), str(guide_path))

        self.log("⏸️  等待人工执行测试", "HUMAN")
        self.log(f"  📋 测试用例: {xlsx_path}（{case_count} 条）", "HUMAN")
        self.log(f"  📖 执行指引: {guide_path}（按模块分组，P0 优先）", "HUMAN")
        self.log("  ✏️  在「执行结果」列填写：通过/失败/阻塞/跳过", "HUMAN")
        self.log("  💡 建议先看 execution_guide.md 了解测什么和怎么测", "HUMAN")
        self.log(f"  ▶️  填完后运行: python cli.py resume -o {Path(xlsx_path).parent}", "HUMAN")
        return StepResult(ok=False, error="等待人工执行测试", human=True)

    @staticmethod
    def _generate_execution_guide(xlsx_path: str, guide_path: str) -> int:
        """生成测试执行指引（失败不阻塞，返回 0）。"""
        try:
            import sys
            from pathlib import Path
            script = Path(__file__).resolve().parents[2] / "scripts" / "generate_execution_guide.py"
            if script.exists():
                import subprocess
                r = subprocess.run(
                    [sys.executable, str(script), xlsx_path, "-o", guide_path],
                    capture_output=True, text=True, timeout=30,
                )
                if r.returncode == 0:
                    # 从输出提取用例数
                    import re
                    m = re.search(r"(\d+) 条用例", r.stdout)
                    return int(m.group(1)) if m else 0
        except Exception:
            pass
        return 0

    @staticmethod
    def _check_has_results(xlsx_path: str) -> bool:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            if not ws:
                wb.close()
                return False
            result_col = None
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col).value or "").strip()
                if header == "执行结果" or ("执行" in header and "结果" in header):
                    result_col = col
                    break
            if not result_col:
                wb.close()
                return False
            filled = 0
            for row in range(2, ws.max_row + 1):
                val = str(ws.cell(row=row, column=result_col).value or "").strip()
                if val:
                    filled += 1
            wb.close()
            return filled > 0
        except Exception:
            return False
