#!/usr/bin/env python3
"""
Step 5: 用例评审（AI 步骤）

读取测试用例 → 调用 LLM 四维质检 → 输出 test_case_review_report.md
"""

from core.llm_client import LLMError
from core.prompt_loader import build_kb_context, load_prompt, render
from core.steps.base import BaseStep, StepResult


class Step5Review(BaseStep):
    step_id = 5
    step_name = "用例评审"
    output_file = "test_case_review_report.md"

    def run(self, **kwargs) -> StepResult:  # type: ignore[override]
        """
        kwargs:
            test_cases: 测试用例文本（Markdown 表格或文本）
            kb_context: 知识库上下文（可选）
        """
        test_cases = kwargs.get("test_cases", "")
        kb_context = kwargs.get("kb_context", "")

        self.log(f"Step {self.step_id}/7: {self.step_name}", "STEP")

        if not test_cases:
            # 尝试从输出目录读取 Excel 并转换为文本
            test_cases = self._read_testcases_as_text()
        if not test_cases:
            return StepResult(ok=False, error="缺少测试用例数据")

        if not self.llm:
            return StepResult(ok=False, error="AI 步骤需要 LLM 客户端")

        # 1. 组装提示词
        template = load_prompt("case_review")
        prompt = render(
            template,
            test_cases=test_cases,
            kb_context=build_kb_context(kb_context),
        )

        # 2. 调用 LLM
        try:
            self.log("  调用 LLM 进行四维质检...", "INFO")
            response = self.llm.chat_with_retry(
                prompt,
                system="你是一位资深测试质量专家，擅长严格的用例评审。",
            )
        except LLMError as e:
            self.log(f"LLM 调用失败: {e}", "ERR")
            return StepResult(ok=False, error=str(e))

        # 3. 质量自检
        check_result = self.self_check(
            response,
            criteria=(
                "1. 评审是否覆盖了所有4个维度（完整性/清晰性/准确性/可执行性）？\n"
                "2. 完整性评审：是否识别了缺失用例、遗漏场景？\n"
                "3. 清晰性评审：是否指出了步骤含糊、预期结果不明确的用例？\n"
                "4. 准确性评审：是否发现预期结果与实际不符、逻辑错误的用例？\n"
                "5. 可执行性评审：是否评估了测试步骤的可操作性、数据依赖性？\n"
                "6. 整改清单中的建议是否具体可操作（非'需要补充'这种模糊建议）？"
            ),
        )
        score = check_result.get("score", 0)
        self.log(f"  自检评分: {score}/100", "INFO")

        # 不合格重跑
        if not check_result.get("passed", False) and score < 70:
            issues = check_result.get("issues", [])
            self.log(f"  自检未通过，带着改进意见重跑 (问题: {len(issues)} 个)", "WARN")
            improvement_hint = "\n".join(f"- {issue}" for issue in issues)
            retry_prompt = (
                prompt
                + f"\n\n## 上次输出的问题（请务必改进）\n{improvement_hint}\n"
                + "\n请重新生成改进后的版本。"
            )
            try:
                response = self.llm.chat(retry_prompt)
            except LLMError:
                self.log("  重跑失败，使用原始输出", "WARN")
            except Exception:
                self.log("  重跑异常，使用原始输出", "WARN")

        # 4. 写入文件
        self._write_output("test_case_review_report.md", response)

        # 5. 提取评分
        score = self._extract_score(response)
        if score:
            grade = self._score_to_grade(score)
            self.log(f"用例评审完成 — 评分 {score}/100 ({grade})", "OK")
        else:
            self.log("用例评审完成", "OK")

        return StepResult(
            ok=True,
            data={"score": score},
        )

    def _read_testcases_as_text(self) -> str:
        """读取 testcases.xlsx 并转换为文本摘要供 LLM 评审"""

        xlsx_path = self._out("testcases.xlsx")
        if not xlsx_path.exists():
            return ""

        try:
            from openpyxl import load_workbook

            wb = load_workbook(str(xlsx_path), data_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                return ""

            lines = []
            for row in ws.iter_rows(min_row=1, values_only=True):
                # 转为 | 分隔的表格行
                cells = [str(c or "").replace("\n", " ")[:80] for c in row]
                lines.append("| " + " | ".join(cells) + " |")

            wb.close()
            return "\n".join(lines[:200])  # 限制长度，避免超 token

        except ImportError:
            self.log("openpyxl 未安装，无法读取 Excel", "WARN")
            return ""

    @staticmethod
    def _extract_score(report_md: str) -> int:
        """从评审报告中提取总分"""
        import re

        # 匹配 | **总计** | **100** | **XX** | 或 总计 ... XX
        match = re.search(r"总计.*?(\d+)\s*\|?\s*$", report_md, re.MULTILINE)
        if match:
            return int(match.group(1))
        # 匹配评分: XX
        match = re.search(r"评分[:：]\s*(\d+)", report_md)
        if match:
            return int(match.group(1))
        return 0

    @staticmethod
    def _score_to_grade(score: int) -> str:
        if score >= 90:
            return "优秀"
        elif score >= 75:
            return "良好"
        elif score >= 60:
            return "中等"
        return "较差"
