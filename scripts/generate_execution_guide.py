#!/usr/bin/env python3
"""
测试执行指引生成器。

业务背景：Step6 人工执行测试时，用户面对 80+ 条用例的 Excel 不知从何下手。
本脚本从 testcases.xlsx 生成一份按模块分组的执行清单（Markdown），包含：
  - 每条用例的"测什么/怎么测/期望什么"精简摘要
  - 按模块和优先级分组，P0 用例高亮
  - 执行进度追踪表
  - 环境准备 checklist（从前置条件提取）

用法：
  python generate_execution_guide.py <testcases.xlsx> [-o execution_guide.md]
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path


def generate_guide(xlsx_path: str, output_path: str = "") -> str:
    """从 testcases.xlsx 生成测试执行指引 Markdown。

    Returns:
        生成的 Markdown 内容（同时写入 output_path）。
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # 定位列（首行表头）
    headers = [str(c.value or "").strip() for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    # 收集所有用例
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if not vals or not vals[0]:
            continue
        cases.append({
            "id": vals[col.get("用例编号", 0)] if "用例编号" in col else vals[0],
            "module": vals[col.get("所属模块", 1)] if "所属模块" in col else "",
            "feature": vals[col.get("功能点", 2)] if "功能点" in col else "",
            "dim": vals[col.get("测试维度", 3)] if "测试维度" in col else "",
            "title": vals[col.get("用例标题", 4)] if "用例标题" in col else "",
            "priority": vals[col.get("优先级", 5)] if "优先级" in col else "",
            "precond": vals[col.get("前置条件", 6)] if "前置条件" in col else "",
            "steps": vals[col.get("测试步骤", 7)] if "测试步骤" in col else "",
            "data": vals[col.get("测试数据", 8)] if "测试数据" in col else "",
            "expected": vals[col.get("预期结果", 9)] if "预期结果" in col else "",
            "result": vals[col.get("执行结果", 11)] if "执行结果" in col and len(vals) > 11 else "",
        })
    wb.close()

    if not cases:
        return "# 测试执行指引\n\n（未找到测试用例）\n"

    # 按模块分组
    by_module = defaultdict(list)
    for c in cases:
        by_module[c["module"] or "未分类"].append(c)

    # 统计
    total = len(cases)
    from collections import Counter
    pri_count = Counter(c["priority"] for c in cases if c["priority"])
    dim_count = Counter(c["dim"] for c in cases if c["dim"])
    filled = sum(1 for c in cases if c["result"])

    # 构建文档
    lines = []
    lines.append("# 📋 测试执行指引")
    lines.append("")
    lines.append(f"> 本文档由 testcases.xlsx 自动生成，帮助测试人员高效执行。")
    lines.append(f"> **总用例数**: {total} | **已执行**: {filled} | **待执行**: {total - filled}")
    lines.append("")

    # 概览
    lines.append("## 📊 用例概览")
    lines.append("")
    lines.append(f"- **优先级分布**: " + " / ".join(f"{k}:{v}" for k, v in sorted(pri_count.items())))
    lines.append(f"- **维度分布**: " + " / ".join(f"{k}:{v}" for k, v in dim_count.most_common()))
    lines.append(f"- **模块数**: {len(by_module)}")
    lines.append("")

    # 环境准备 checklist（提取所有前置条件的去重关键词）
    lines.append("## 🔧 环境准备 Checklist")
    lines.append("")
    preconds = set()
    for c in cases:
        if c["precond"]:
            # 取前置条件的第一句或关键短语
            pc = str(c["precond"]).split("。")[0].split("\n")[0].strip()
            if len(pc) > 5:
                preconds.add(pc)
    for pc in sorted(preconds)[:10]:  # 最多10条，避免冗长
        lines.append(f"- [ ] {pc}")
    lines.append("")

    # 执行进度追踪表
    lines.append("## 📈 执行进度追踪")
    lines.append("")
    lines.append("| 模块 | 总数 | P0 | P1 | P2 | 已执行 |")
    lines.append("|------|------|----|----|----|--------|")
    for mod, mod_cases in sorted(by_module.items()):
        mc = len(mod_cases)
        p0 = sum(1 for c in mod_cases if c["priority"] == "P0")
        p1 = sum(1 for c in mod_cases if c["priority"] == "P1")
        p2 = sum(1 for c in mod_cases if c["priority"] == "P2")
        done = sum(1 for c in mod_cases if c["result"])
        lines.append(f"| {mod} | {mc} | {p0} | {p1} | {p2} | {done} |")
    lines.append("")

    # 按模块详细执行清单
    lines.append("## 📝 执行清单（按模块）")
    lines.append("")
    lines.append("> 填写 testcases.xlsx 的「执行结果」列：通过/失败/阻塞/跳过")
    lines.append("> 失败时请在「备注」列记录现象和复现步骤")
    lines.append("")

    # 优先级符号
    pri_icon = {"P0": "🔴", "P1": "🟠", "P2": "🟡"}

    for mod, mod_cases in sorted(by_module.items()):
        lines.append(f"### {mod}（{len(mod_cases)} 条）")
        lines.append("")
        # 按优先级排序：P0 > P1 > P2
        mod_cases.sort(key=lambda c: ({"P0": 0, "P1": 1, "P2": 2}.get(c["priority"], 3), c["id"] or ""))
        for c in mod_cases:
            icon = pri_icon.get(c["priority"], "⚪")
            result_mark = {"通过": "✅", "失败": "❌", "阻塞": "⛔", "跳过": "⏭️"}.get(str(c["result"]).strip(), "⬜")
            lines.append(f"#### {result_mark} {icon} {c['id']} — {c['title']}")
            lines.append(f"- **维度**: {c['dim']} | **优先级**: {c['priority']} | **功能点**: {c['feature']}")
            if c["data"]:
                lines.append(f"- **测试数据**: {c['data']}")
            if c["expected"]:
                lines.append(f"- **预期**: {c['expected']}")
            lines.append("")
        lines.append("---")
        lines.append("")

    # 执行说明
    lines.append("## 💡 执行说明")
    lines.append("")
    lines.append("- **执行结果填写位置**: testcases.xlsx 的「执行结果」列")
    lines.append("- **可选值**: 通过 / 失败 / 阻塞 / 跳过")
    lines.append("  - 通过：实际结果符合预期")
    lines.append("  - 失败：实际结果与预期不符（请在备注列记录现象）")
    lines.append("  - 阻塞：因环境/依赖问题无法执行（请在备注列记录阻塞原因）")
    lines.append("  - 跳过：该用例不适用当前版本")
    lines.append("- **执行顺序建议**: 先 P0（必须通过）→ 再 P1（重要）→ 最后 P2（建议）")
    lines.append("- **填完后**: 运行 `python cli.py resume -o <输出目录>` 继续生成测试报告")
    lines.append("")

    content = "\n".join(lines)

    if output_path:
        Path(output_path).write_text(content, encoding="utf-8")

    return content


def main():
    parser = argparse.ArgumentParser(description="从 testcases.xlsx 生成测试执行指引")
    parser.add_argument("xlsx", help="testcases.xlsx 路径")
    parser.add_argument("-o", "--output", default="execution_guide.md", help="输出文件路径")
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        print(f"❌ 文件不存在: {args.xlsx}", file=sys.stderr)
        return 1

    content = generate_guide(args.xlsx, args.output)
    case_count = content.count("#### ")
    print(f"✅ 执行指引已生成: {args.output}（{case_count} 条用例）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
