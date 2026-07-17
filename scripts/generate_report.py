#!/usr/bin/env python3
"""
测试报告生成脚本
读取已执行完成的测试用例 Excel 文件，生成结构化 Markdown 测试质量报告。

用法:
    python generate_report.py <testcases.xlsx> [--output test_report.md] [--requirements requirements_analysis.md]

依赖:
    openpyxl — pip install openpyxl
"""

import argparse
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("错误: 缺少依赖库 openpyxl", file=sys.stderr)
    print("请安装:", file=sys.stderr)
    print("  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════
# 常量定义
# ═══════════════════════════════════════════════════════════════

# 执行结果归一化映射（中英文缩写 → 标准状态）
RESULT_NORMALIZE = {
    # 通过
    "通过": "pass", "pass": "pass", "passed": "pass", "✅": "pass", "pass(✅)": "pass",
    "成功": "pass", "success": "pass", "ok": "pass",
    # 失败
    "失败": "fail", "fail": "fail", "failed": "fail", "❌": "fail", "fail(❌)": "fail",
    # 阻塞
    "阻塞": "block", "block": "block", "blocked": "block", "⛔": "block", "block(⛔)": "block",
    # 跳过
    "跳过": "skip", "skip": "skip", "skipped": "skip", "⏭️": "skip", "skip(⏭️)": "skip",
    # 未执行
    "未执行": "not_run", "n/a": "not_run", "not run": "not_run", "na": "not_run",
    "pending": "not_run", "—": "not_run", "-": "not_run", "": "not_run",
}

# 标准状态中文显示 + 图标
STATUS_DISPLAY = {
    "pass":    ("通过", "✅"),
    "fail":    ("失败", "❌"),
    "block":   ("阻塞", "⛔"),
    "skip":    ("跳过", "⏭️"),
    "not_run": ("未执行", "⏳"),
}

# 质量评级标准
QUALITY_GRADES = [
    (0.95, "优秀 🏆", "质量优秀，可以发布"),
    (0.85, "良好 ✅", "质量较好，修复少量问题后可发布"),
    (0.70, "中等 ⚠️", "质量一般，需要修复较多问题"),
    (0.00, "较差 ❌", "质量不达标，不建议发布"),
]

# ═══════════════════════════════════════════════════════════════
# Excel 读取器
# ═══════════════════════════════════════════════════════════════

class ExcelReader:
    """读取已执行的测试用例 Excel 文件"""

    # 可能的列名（模糊匹配）
    COLUMN_ALIASES = {
        "id":          ["用例编号", "编号", "id", "case_id", "tc"],
        "module":      ["所属模块", "模块", "module"],
        "feature":     ["功能点", "feature", "功能"],
        "dimension":   ["测试维度", "维度", "类型", "type", "dimension", "category"],
        "title":       ["用例标题", "标题", "title", "用例名称", "名称", "描述"],
        "priority":    ["优先级", "priority", "级别"],
        "precondition": ["前置条件", "precondition"],
        "steps":       ["测试步骤", "步骤", "steps"],
        "test_data":   ["测试数据", "data", "test_data"],
        "expected":    ["预期结果", "预期", "expected", "期望"],
        "result":      ["执行结果", "结果", "result", "status", "执行状态", "测试结果"],
        "remark":      ["备注", "remark", "说明", "note", "comments"],
    }

    def __init__(self):
        self.header_map = {}

    def _normalize_header(self, text: str) -> str:
        """标准化表头文本：去除空白、标点、统一大小写"""
        if not text:
            return ""
        return re.sub(r"[\s\u3000（）()【】\[\]]", "", str(text)).strip().lower()

    def _match_column(self, header_text: str) -> str:
        """模糊匹配列名，返回标准字段名"""
        normalized = self._normalize_header(header_text)
        for field, aliases in self.COLUMN_ALIASES.items():
            for alias in aliases:
                if self._normalize_header(alias) == normalized:
                    return field
                # 部分匹配（如"执行结果"包含"结果"）
                if normalized and self._normalize_header(alias) in normalized:
                    return field
        return ""

    def _normalize_result(self, raw_value) -> str:
        """将执行结果归一化为标准状态

        查找顺序：精确匹配 RESULT_NORMALIZE → 模糊关键词匹配 → 默认 not_run。
        """
        if raw_value is None:
            return "not_run"
        text = str(raw_value).strip()
        if not text:
            return "not_run"

        normalized = text.lower()
        # 精确查表（用 .get 避免 O(n) 遍历）
        if text in RESULT_NORMALIZE:
            return RESULT_NORMALIZE[text]
        if normalized in RESULT_NORMALIZE:
            return RESULT_NORMALIZE[normalized]

        # 模糊匹配
        for keyword, status in [("pass", "pass"), ("fail", "fail"),
                                 ("block", "block"), ("skip", "skip")]:
            if keyword in normalized:
                return status
        return "not_run"

    def read(self, file_path: str) -> tuple:
        """
        读取 Excel 文件
        返回: (test_cases, metadata)
          test_cases: list[dict] 每行一个字典
          metadata: dict 文件信息
        """
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 解析表头（第1行）
        headers = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value:
                field = self._match_column(cell_value)
                if field:
                    headers[field] = col

        if not headers:
            print("⚠️  无法识别 Excel 表头，请检查列名", file=sys.stderr)
            return [], {}

        # 检查是否有执行结果列
        has_result_col = "result" in headers
        if not has_result_col:
            print("⚠️  Excel 中没有「执行结果」列", file=sys.stderr)

        # 读取数据行
        test_cases = []
        for row in range(2, ws.max_row + 1):
            case = {}
            for field, col in headers.items():
                cell_value = ws.cell(row=row, column=col).value
                case[field] = str(cell_value) if cell_value is not None else ""
            # 归一化执行结果
            if "result" in case:
                case["result"] = self._normalize_result(case.get("result", ""))
            else:
                case["result"] = "not_run"
            case["_row"] = row
            test_cases.append(case)

        wb.close()

        metadata = {
            "sheet_name": ws.title,
            "total_rows": ws.max_row - 1,
            "has_result_col": has_result_col,
            "columns_detected": list(headers.keys()),
        }

        return test_cases, metadata


# ═══════════════════════════════════════════════════════════════
# 报告分析器
# ═══════════════════════════════════════════════════════════════

class ReportAnalyzer:
    """分析测试结果数据"""

    def analyze(self, test_cases: list) -> dict:
        """执行完整分析，返回统计数据"""
        stats = {
            "total": len(test_cases),
            "pass": 0,
            "fail": 0,
            "block": 0,
            "skip": 0,
            "not_run": 0,
            "by_module": defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0, "block": 0, "skip": 0}),
            "by_priority": defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0, "block": 0, "skip": 0}),
            "by_dimension": defaultdict(lambda: {"total": 0, "pass": 0, "fail": 0, "block": 0, "skip": 0}),
            "failed_cases": [],
            "blocked_cases": [],
        }

        for tc in test_cases:
            status = tc.get("result", "not_run")
            module = tc.get("module", "未分类") or "未分类"
            priority = tc.get("priority", "P1") or "P1"
            dimension = tc.get("dimension", "未知") or "未知"

            # 总计
            if status in stats:
                stats[status] += 1

            # 按模块
            stats["by_module"][module]["total"] += 1
            if status in ("pass", "fail", "block", "skip"):
                stats["by_module"][module][status] += 1

            # 按优先级
            stats["by_priority"][priority]["total"] += 1
            if status in ("pass", "fail", "block", "skip"):
                stats["by_priority"][priority][status] += 1

            # 按维度
            stats["by_dimension"][dimension]["total"] += 1
            if status in ("pass", "fail", "block", "skip"):
                stats["by_dimension"][dimension][status] += 1

            # 失败用例
            if status == "fail":
                stats["failed_cases"].append(tc)

            # 阻塞用例
            if status == "block":
                stats["blocked_cases"].append(tc)

        # 计算通过率和执行率
        executed = stats["pass"] + stats["fail"] + stats["block"] + stats["skip"]
        stats["executed"] = executed
        stats["pass_rate"] = stats["pass"] / executed if executed > 0 else 0.0
        stats["execution_rate"] = executed / stats["total"] if stats["total"] > 0 else 0.0

        # 当有用例但无执行结果时，输出警告（避免误导性的 0% 通过率）
        if stats["total"] > 0 and executed == 0:
            print(
                "⚠️  所有用例均未执行（执行结果列为空），通过率 0% 无参考意义。",
                file=sys.stderr,
            )

        return stats

    def get_quality_grade(self, pass_rate: float) -> tuple:
        """根据通过率获取质量评级"""
        for threshold, grade, desc in QUALITY_GRADES:
            if pass_rate >= threshold:
                return grade, desc
        return "较差 ❌", "质量不达标"

    # 失败原因关键词映射（提升为类常量，避免每次调用重建字典）
    # ⚠️ 顺序很重要：infer_failure_cause 在同分时返回本字典中定义顺序最靠前的类别，
    #    因此更具体的类别必须放在前面：
    #    1)「代码缺陷」先于「功能未实现」：TC-006 文本同时含"未发现"+"不验证"，
    #       语义上"缺校验逻辑"属于代码缺陷而非整功能缺失，需让代码缺陷在平局时胜出；
    #       而纯"功能未实现"用例（TC-010/016）的文本不含代码缺陷关键词，不受影响。
    #    2) 两个新类别都先于"数据校验失败"等运行时类别：因为"未发现校验"比"校验"更具体，
    #       代码审查型失败（无运行时报错）若不优先匹配会被"接口/服务异常"/"权限/认证问题"
    #       等"功能测试型"类别误判。
    FAILURE_CAUSE_MAP = {
        # ── 代码审查型失败（优先匹配，无运行时报错）──
        # 关键词设计为"代码审查场景"的强信号短语，避免与运行时失败类别冲突：
        # 「校验」单独出现会命中"数据校验失败"，故此处用"未校验/无校验/不验证"
        # 等"否定+校验"组合，专指"代码缺校验逻辑"。
        "代码缺陷": ["未校验", "缺失", "无校验", "不验证", "漏掉", "缺少",
                    "未检查", "未做", "完整性校验"],
        "功能未实现": ["未发现", "不存在", "未实现", "未支持", "该功能",
                     "规划中", "未承诺", "不支持", "无法通过"],
        # ── 运行时失败（功能测试型）──
        "接口/服务异常": ["超时", "timeout", "502", "503", "500", "服务不可用", "连接",
                      "网络", "接口", "网关", "响应", "空指针", "null", "undefined"],
        "权限/认证问题": ["权限", "越权", "认证", "鉴权", "token", "未授权", "无权",
                       "登录", "会话", "过期", "身份", "401", "403"],
        "数据校验失败": ["校验", "验证", "唯一性", "重复", "冲突", "为空", "格式",
                      "长度", "范围", "枚举", "必填", "非法", "无效"],
        "状态流转错误": ["状态", "流转", "流程", "跳转", "返回", "回调", "待支付",
                      "已完成", "已取消", "已关闭"],
        "并发/竞态": ["并发", "同时", "竞态", "锁", "死锁", "超卖", "重复扣减"],
        "边界值问题": ["最小", "最大", "边界", "临界", "上限", "下限", "溢出",
                     "零", "0", "负数", "空"],
        "兼容性": ["兼容", "编码", "GBK", "UTF", "乱码", "版本", "浏览器"],
        "环境配置": ["环境", "配置", "部署", "Docker", "端口", "SSL", "证书"],
        "功能缺陷": ["金额", "计算", "显示", "保存", "提交", "跳转", "排序",
                    "筛选", "查询", "按钮", "页面", "渲染"],
        "数据问题": ["数据", "库存", "不存在", "脏数据", "同步", "一致性"],
        "用例缺陷": ["预期", "步骤", "用例", "描述"],
    }

    def infer_failure_cause(self, case: dict) -> str:
        """推断失败原因（增强版：综合标题、维度、优先级、备注多维信息）

        当多个原因类别得分相同时，返回 FAILURE_CAUSE_MAP 中定义顺序最靠前的类别，
        确保结果稳定可复现（不依赖字典遍历顺序）。

        代码审查型失败（功能未实现 / 代码缺陷）优先级高于运行时失败类别：
        当文本命中这两个类别时，运行时类别不再享受维度加权，避免"安全测试"
        维度的"缺校验逻辑"被加权成"权限/认证问题"，或"异常测试"维度的"功能
        未实现"被加权成"接口/服务异常"。
        """
        title = case.get("title", "")
        remark = case.get("remark", "")
        dimension = case.get("dimension", "")
        steps = case.get("steps", "")
        expected = case.get("expected", "")

        combined = f"{title} {remark} {dimension} {steps} {expected}"

        # 先算代码审查型类别的原始命中数；若 >0 则本用例属"代码审查型失败"
        review_causes = ("代码缺陷", "功能未实现")
        review_hit = sum(
            1 for c in review_causes
            for kw in self.FAILURE_CAUSE_MAP[c] if kw in combined
        )

        best_cause = "待确认"
        best_score = 0
        for cause, keywords in self.FAILURE_CAUSE_MAP.items():
            score = sum(1 for kw in keywords if kw in combined)

            # 代码审查型失败已确定：跳过运行时类别的维度加权，避免误判
            if review_hit > 0 and cause not in review_causes:
                pass
            else:
                # 维度加权：异常测试更可能是接口/环境问题
                if "异常" in dimension and cause in ("接口/服务异常", "环境配置"):
                    score += 1
                # 安全测试更可能是权限问题
                if "安全" in dimension and cause == "权限/认证问题":
                    score += 1

            # 代码审查型类别自身加权：安全测试缺校验→代码缺陷；异常测试缺功能→功能未实现
            if review_hit > 0:
                if "安全" in dimension and cause == "代码缺陷":
                    score += 1
                if "异常" in dimension and cause == "功能未实现":
                    score += 1

            # 严格大于：相同得分时保留先出现的类别（稳定排序）
            if score > best_score:
                best_score = score
                best_cause = cause

        return best_cause

    def get_fix_suggestion(self, cause: str) -> str:
        """根据失败原因生成修复建议"""
        suggestions = {
            "功能缺陷": "检查功能实现代码，确认逻辑是否正确，修复后回归测试。",
            "环境问题": "检查测试环境配置和网络连接，确认服务状态正常后重试。",
            "数据问题": "检查测试数据准备是否充分，清理脏数据后重新执行。",
            "用例缺陷": "核对预期结果是否正确，必要时更新用例描述和步骤。",
            "功能未实现": "确认该功能是否在需求范围内；若在则补实现，若不在则修正用例预期。",
            "代码缺陷": "补充对应的校验/检查逻辑，参考用例的预期结果。",
            "接口/服务异常": "检查后端接口返回状态码和响应体，确认服务可用性和超时配置。",
            "权限/认证问题": "检查用户权限配置和 Token/Session 有效性，确认鉴权中间件正确拦截。",
            "数据校验失败": "检查前后端校验规则是否一致，确认必填项、格式、范围限制实现完整。",
            "状态流转错误": "检查状态机配置和流程引擎，确认状态转移条件与需求一致。",
            "并发/竞态": "检查锁机制和事务隔离级别，确认并发控制（乐观锁/悲观锁）正确实现。",
            "边界值问题": "检查边界条件的代码处理，确认 off-by-one 和边界判断逻辑。",
            "兼容性": "检查不同编码/浏览器/版本下的兼容处理，确认字符集和 API 版本兼容。",
            "环境配置": "检查环境变量、端口、SSL 证书等配置项，确认与生产环境一致。",
            "待确认": "需要开发人员排查具体原因后确定修复方案。",
        }
        return suggestions.get(cause, suggestions["待确认"])

    def assess_risk(self, stats: dict) -> dict:
        """风险评估"""
        risks = {"high": [], "medium": [], "low": []}

        # P0 用例失败 → 高风险
        p0_fail = stats["by_priority"].get("P0", {}).get("fail", 0)
        if p0_fail > 0:
            risks["high"].append(f"P0 高优先级用例失败 {p0_fail} 个")

        # P1 用例失败 → 中风险
        p1_fail = stats["by_priority"].get("P1", {}).get("fail", 0)
        if p1_fail > 0:
            risks["medium"].append(f"P1 用例失败 {p1_fail} 个")

        # 模块通过率 — 单次遍历完成分级（避免重复遍历+字符串去重）
        for module, data in stats["by_module"].items():
            executed = data["pass"] + data["fail"] + data["block"] + data["skip"]
            if executed > 0:
                rate = data["pass"] / executed
                if rate < 0.70:
                    risks["high"].append(f"模块「{module}」通过率仅 {rate:.0%}")
                elif rate < 0.90:
                    risks["medium"].append(f"模块「{module}」通过率 {rate:.0%}，需关注")

        # 阻塞用例
        if stats["block"] > 0:
            risks["medium"].append(f"有 {stats['block']} 个用例被阻塞，可能影响覆盖率")

        # P2 失败 → 低风险
        p2_fail = stats["by_priority"].get("P2", {}).get("fail", 0)
        if p2_fail > 0:
            risks["low"].append(f"P2 用例失败 {p2_fail} 个")

        # 低通过率但不涉及核心模块
        if stats["pass_rate"] < 0.90 and stats["pass_rate"] >= 0.70:
            if not risks["high"]:
                risks["low"].append(f"整体通过率 {stats['pass_rate']:.1%}，有提升空间")

        # 综合风险等级
        if risks["high"]:
            level = "🔴 高风险"
        elif risks["medium"]:
            level = "🟡 中风险"
        else:
            level = "🟢 低风险"

        risks["level"] = level
        return risks

    def get_release_recommendation(self, stats: dict, risks: dict) -> str:
        """发布建议"""
        p0_fail = stats["by_priority"].get("P0", {}).get("fail", 0)
        p0_block = stats["by_priority"].get("P0", {}).get("block", 0)

        if p0_fail > 0 or p0_block > 0:
            return ("⛔ **不建议发布** — 存在 P0 级别用例失败或阻塞，"
                    "必须优先修复后再重新评估。")
        if risks["high"]:
            return ("⚠️ **暂缓发布** — 存在高风险问题，"
                    "修复后需回归测试。")
        if stats["pass_rate"] >= 0.95:
            return "✅ **建议发布** — 测试通过率优秀，质量达标。"
        if stats["pass_rate"] >= 0.85:
            return ("✅ **可以发布** — 通过率良好，建议修复已知问题后在下一版本迭代修复。")
        if stats["pass_rate"] >= 0.70:
            return ("⚠️ **谨慎发布** — 通过率一般，建议修复核心问题后发布。")
        return "❌ **不建议发布** — 通过率较低，需大量修复后重新测试。"


# ═══════════════════════════════════════════════════════════════
# 报告生成器
# ═══════════════════════════════════════════════════════════════

class ReportGenerator:
    """生成 Markdown 测试报告"""

    def __init__(self):
        self.analyzer = ReportAnalyzer()

    def _fmt_percent(self, value: float) -> str:
        """格式化百分比"""
        return f"{value:.1%}"

    def _pass_rate(self, data: dict) -> float:
        """计算通过率"""
        executed = data["pass"] + data["fail"] + data["block"] + data["skip"]
        return data["pass"] / executed if executed > 0 else 0.0

    def _rate_emoji(self, rate: float) -> str:
        """通过率对应的图标"""
        if rate >= 0.90:
            return "🟢"
        elif rate >= 0.75:
            return "🟡"
        elif rate >= 0.50:
            return "🟠"
        else:
            return "🔴"

    def generate(self, test_cases: list, file_path: str, source_file: str = "",
                 requirements_path: str = "") -> tuple:
        """生成完整 Markdown 报告

        Returns:
            (content, stats, risks): 报告文本、统计数据、风险信息
        """
        stats = self.analyzer.analyze(test_cases)
        grade, grade_desc = self.analyzer.get_quality_grade(stats["pass_rate"])
        risks = self.analyzer.assess_risk(stats)
        release = self.analyzer.get_release_recommendation(stats, risks)

        now = datetime.now().strftime("%Y-%m-%d %H:%M")

        lines = []
        lines.append("# 测试质量报告\n")
        lines.append(f"*由 AI 自动生成 | 生成时间：{now}*\n")
        if source_file:
            lines.append(f"*数据来源：`{source_file}`*\n")
        lines.append("")

        # ─── 1. 总体概览 ───
        lines.append("## 📊 总体概览\n")
        lines.append("| 指标 | 数值 | 说明 |")
        lines.append("|------|------|------|")
        lines.append(f"| 总用例数 | {stats['total']} 个 | — |")
        lines.append(f"| ✅ 通过 | {stats['pass']} 个 | {self._fmt_percent(stats['pass'] / stats['total']) if stats['total'] else '0%'} |")
        lines.append(f"| ❌ 失败 | {stats['fail']} 个 | {self._fmt_percent(stats['fail'] / stats['total']) if stats['total'] else '0%'} |")
        lines.append(f"| ⛔ 阻塞 | {stats['block']} 个 | {self._fmt_percent(stats['block'] / stats['total']) if stats['total'] else '0%'} |")
        lines.append(f"| ⏭️ 跳过 | {stats['skip']} 个 | {self._fmt_percent(stats['skip'] / stats['total']) if stats['total'] else '0%'} |")
        lines.append(f"| ⏳ 未执行 | {stats['not_run']} 个 | {self._fmt_percent(stats['not_run'] / stats['total']) if stats['total'] else '0%'} |")
        lines.append(f"| **通过率** | **{self._fmt_percent(stats['pass_rate'])}** | 通过/(通过+失败+阻塞+跳过) |")
        lines.append(f"| **执行率** | **{self._fmt_percent(stats['execution_rate'])}** | 已执行/总数 |")
        lines.append(f"| **质量评级** | **{grade}** | {grade_desc} |")
        lines.append("")

        # ─── 2. 模块通过率 ───
        lines.append("## 📈 模块通过率\n")
        lines.append("| 模块 | 总数 | 通过 | 失败 | 阻塞 | 跳过 | 通过率 | 状态 |")
        lines.append("|------|------|------|------|------|------|--------|------|")
        # 按通过率升序（最差的排前面）
        module_rates = []
        for module, data in stats["by_module"].items():
            rate = self._pass_rate(data)
            module_rates.append((module, data, rate))
        module_rates.sort(key=lambda x: x[2])  # 通过率低的排前面

        for module, data, rate in module_rates:
            emoji = self._rate_emoji(rate)
            lines.append(
                f"| {module} | {data['total']} | {data['pass']} | {data['fail']} | "
                f"{data['block']} | {data['skip']} | {self._fmt_percent(rate)} | {emoji} |"
            )
        lines.append("")

        # ─── 3. 优先级分析 ───
        lines.append("## 🔖 优先级分析\n")
        lines.append("| 优先级 | 总数 | 通过 | 失败 | 阻塞 | 跳过 | 通过率 | 状态 |")
        lines.append("|--------|------|------|------|------|------|--------|------|")
        for pri in ["P0", "P1", "P2"]:
            data = stats["by_priority"].get(pri)
            if data and data["total"] > 0:
                rate = self._pass_rate(data)
                emoji = self._rate_emoji(rate)
                lines.append(
                    f"| {pri} | {data['total']} | {data['pass']} | {data['fail']} | "
                    f"{data['block']} | {data['skip']} | {self._fmt_percent(rate)} | {emoji} |"
                )
        lines.append("")

        # ─── 4. 测试维度分析 ───
        if stats["by_dimension"]:
            lines.append("## 🧪 测试维度分析\n")
            lines.append("| 测试维度 | 总数 | 通过 | 失败 | 阻塞 | 跳过 | 通过率 |")
            lines.append("|---------|------|------|------|------|------|--------|")
            for dim, data in stats["by_dimension"].items():
                if data["total"] > 0:
                    rate = self._pass_rate(data)
                    lines.append(
                        f"| {dim} | {data['total']} | {data['pass']} | {data['fail']} | "
                        f"{data['block']} | {data['skip']} | {self._fmt_percent(rate)} |"
                    )
            lines.append("")

        # ─── 5. 失败用例分析 ───
        if stats["failed_cases"]:
            lines.append("## 🔍 失败用例分析\n")
            lines.append(f"共 **{len(stats['failed_cases'])}** 个失败用例。\n")

            for i, case in enumerate(stats["failed_cases"], 1):
                tc_id = case.get("id", "N/A")
                title = case.get("title", "未命名")
                module = case.get("module", "未分类")
                priority = case.get("priority", "P1")
                expected = case.get("expected", "—")
                remark = case.get("remark", "—")
                cause = self.analyzer.infer_failure_cause(case)

                lines.append(f"### {i}. {tc_id}: {title}\n")
                lines.append(f"- **所属模块：** {module}")
                lines.append(f"- **优先级：** {priority}")
                lines.append(f"- **预期结果：** {expected}")
                lines.append(f"- **失败原因（推断）：** {cause}")
                if remark and remark != "—":
                    lines.append(f"- **备注：** {remark}")
                lines.append(f"- **修复建议：** {self.analyzer.get_fix_suggestion(cause)}")
                lines.append("")
        else:
            lines.append("## 🔍 失败用例分析\n")
            lines.append("🎉 没有失败用例！\n")

        # ─── 6. 阻塞用例分析 ───
        if stats["blocked_cases"]:
            lines.append("## ⛔ 阻塞用例分析\n")
            lines.append(f"共 **{len(stats['blocked_cases'])}** 个阻塞用例。\n")
            lines.append("| 用例编号 | 标题 | 模块 | 优先级 | 阻塞原因 |")
            lines.append("|---------|------|------|--------|---------|")
            for case in stats["blocked_cases"]:
                tc_id = case.get("id", "N/A")
                title = case.get("title", "未命名")
                module = case.get("module", "未分类")
                priority = case.get("priority", "P1")
                remark = case.get("remark", "—") or "—"
                lines.append(f"| {tc_id} | {title} | {module} | {priority} | {remark} |")
            lines.append("")
        else:
            lines.append("## ⛔ 阻塞用例分析\n")
            lines.append("🎉 没有阻塞用例！\n")

        # ─── 7. 风险评估 ───
        lines.append("## ⚠️ 风险评估\n")
        lines.append(f"**综合风险等级：{risks['level']}**\n")

        if risks["high"]:
            lines.append("### 🔴 高风险")
            for r in risks["high"]:
                lines.append(f"- {r}")
            lines.append("")
        if risks["medium"]:
            lines.append("### 🟡 中风险")
            for r in risks["medium"]:
                lines.append(f"- {r}")
            lines.append("")
        if risks["low"]:
            lines.append("### 🟢 低风险")
            for r in risks["low"]:
                lines.append(f"- {r}")
            lines.append("")

        if not (risks["high"] or risks["medium"] or risks["low"]):
            lines.append("✅ 未发现显著风险。\n")

        # ─── 8. 测试结论与建议 ───
        lines.append("## 📝 测试结论与建议\n")
        lines.append("### 发布建议\n")
        lines.append(f"{release}\n")

        lines.append("### 后续行动项\n")
        action_items = []
        if stats["fail"] > 0:
            action_items.append(f"- [ ] 修复 {stats['fail']} 个失败用例对应的缺陷")
        if stats["block"] > 0:
            action_items.append(f"- [ ] 排查 {stats['block']} 个阻塞用例的阻塞原因")
        if stats["not_run"] > 0:
            action_items.append(f"- [ ] 补充执行 {stats['not_run']} 个未执行用例")
        # 模块通过率低的建议
        for module, data, rate in module_rates:
            if rate < 0.70:
                action_items.append(f"- [ ] 重点关注模块「{module}」（通过率 {rate:.0%}）")
        if not action_items:
            action_items.append("- [x] 所有测试通过，无需额外行动 🎉")

        for item in action_items:
            lines.append(item)
        lines.append("")

        # ─── 页脚 ───
        lines.append("---")
        lines.append(f"\n*报告由 generate-report Skill v1.0.0 自动生成 | {now}*\n")

        content = "\n".join(lines)

        # 写入文件
        output_path = Path(file_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(content, encoding="utf-8")

        return content, stats, risks


# ═══════════════════════════════════════════════════════════════
# 主函数
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="测试报告生成器 — 根据执行结果生成 Markdown 质量报告"
    )
    parser.add_argument("input", help="已执行完成的测试用例文件 (.xlsx)")
    parser.add_argument("-o", "--output", default="test_report.md",
                        help="输出报告路径 (默认: test_report.md)")
    parser.add_argument("-r", "--requirements", default="",
                        help="需求分析文件路径 (可选，用于需求覆盖率)")
    args = parser.parse_args()

    # 检查输入文件
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ 文件不存在: {args.input}", file=sys.stderr)
        return 1

    if input_path.suffix.lower() not in (".xlsx", ".xls"):
        print(f"⚠️  当前版本主要支持 .xlsx 格式，收到的文件: {input_path.suffix}", file=sys.stderr)

    # 读取 Excel
    print(f"📖 读取测试用例文件: {args.input}")
    reader = ExcelReader()
    test_cases, metadata = reader.read(args.input)

    if not test_cases:
        print("⚠️  未读取到任何测试用例数据", file=sys.stderr)
        return 1

    print(f"✅ 读取到 {len(test_cases)} 条测试用例")
    print(f"   识别的列: {', '.join(metadata.get('columns_detected', []))}")

    # 检查执行结果
    if not metadata.get("has_result_col"):
        print("⚠️  警告: 未检测到「执行结果」列，所有用例将标记为「未执行」", file=sys.stderr)

    executed = sum(1 for tc in test_cases if tc.get("result") in ("pass", "fail", "block", "skip"))
    if executed == 0:
        print("⚠️  警告: 所有用例均为「未执行」状态，请先执行测试后再生成报告", file=sys.stderr)

    # 生成报告
    print("\n🔨 生成测试报告...")
    generator = ReportGenerator()
    content, stats, risks = generator.generate(
        test_cases, args.output,
        source_file=args.input,
        requirements_path=args.requirements
    )
    grade, _ = generator.analyzer.get_quality_grade(stats["pass_rate"])

    report_size = Path(args.output).stat().st_size

    # 打印汇总
    print(f"\n{'='*50}")
    print("✅ 测试报告生成完成！")
    print(f"{'='*50}")
    print("📊 总体概览:")
    print(f"  - 总用例数: {stats['total']} 个")
    print(f"  - ✅ 通过: {stats['pass']} 个 ({stats['pass_rate']:.1%})")
    print(f"  - ❌ 失败: {stats['fail']} 个")
    print(f"  - ⛔ 阻塞: {stats['block']} 个")
    print(f"  - ⏭️  跳过: {stats['skip']} 个")
    print(f"  - ⏳ 未执行: {stats['not_run']} 个")
    print(f"  - 执行率: {stats['execution_rate']:.1%}")
    print(f"  - 质量评级: {grade}")
    print(f"\n📁 输出文件: {args.output} ({report_size / 1024:.1f} KB)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
