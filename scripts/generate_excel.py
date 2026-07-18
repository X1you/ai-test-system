#!/usr/bin/env python3
"""
生成 Excel 格式测试用例
读取测试点 Markdown 文件，生成结构化的 .xlsx 文件

用法:
    python generate_excel.py <testpoints.md> [--output testcases.xlsx] [--dimensions all|basic|positive,negative]
"""

import argparse
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少依赖库 openpyxl", file=sys.stderr)
    print("请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from scripts.common import (
    TestPointParser,
    assign_priority,
    filter_by_dimensions,
    filter_by_priorities,
)

# ═══════════════════════════════════════════════════════════════
# 常量定义
# ═════════════════════════════════════════════════════

class TestCaseGenerator:
    """根据测试点生成测试用例"""

    DIMENSION_MAP = {
        "正向测试": "positive",
        "负向测试": "negative",
        "边界测试": "boundary",
        "异常测试": "exception",
        "性能测试": "performance",
        "安全测试": "security",
    }

    # ── 动作关键词→步骤模板映射表 ──
    # 每条: (关键词列表, 正向步骤列表, 负向步骤列表)
    # 按优先级从上到下匹配，先匹配到的生效
    ACTION_TEMPLATES = [
        # ── 认证/登录类 ──
        (["登录", "登录成功", "登录失败"], [
            "2. 在登录页面输入正确的账号和密码（详见测试数据）",
            "3. 点击登录按钮提交",
            "4. 检查是否跳转到首页/正确页面，生成有效 Token",
        ], [
            "2. 在登录页面输入错误的/异常的账号或密码（详见测试数据）",
            "3. 点击登录按钮提交",
            "4. 检查系统是否拒绝登录并给出正确提示",
        ]),
        # ── 注册/开通类 ──
        (["注册", "开通", "激活"], [
            "2. 在注册页面输入合法的手机号，获取并输入验证码",
            "3. 设置符合强度要求的密码（详见测试数据）",
            "4. 点击注册/提交按钮",
            "5. 检查注册是否成功，数据库是否新增用户记录",
        ], [
            "2. 在注册页面输入无效/重复的数据（详见测试数据）",
            "3. 点击注册/提交按钮",
            "4. 检查系统是否拒绝注册并给出正确提示",
        ]),
        # ── 上传/导入类 ──
        (["上传", "导入"], [
            "2. 进入上传/导入页面",
            "3. 选择符合要求的文件（格式、大小、编码正确）",
            "4. 点击上传/导入按钮",
            "5. 确认上传成功，检查返回结果/解析结果",
        ], [
            "2. 准备不符合要求的文件（格式错误/超大/空文件/编码异常）",
            "3. 尝试上传/导入该文件",
            "4. 检查系统是否拒绝并给出正确的错误提示",
        ]),
        # ── 导出/下载类 ──
        (["导出", "下载"], [
            "2. 进入含可导出数据的页面/模块",
            "3. 选择导出范围/条件（如有）",
            "4. 点击导出/下载按钮",
            "5. 检查文件是否成功下载，内容是否完整",
        ], [
            "2. 在无数据/无权限状态下尝试导出",
            "3. 检查系统是否阻止并提示",
        ]),
        # ── 创建/新增/生成类 ──
        (["创建", "新增", "生成", "添加", "新建"], [
            "2. 进入创建/新增页面",
            "3. 填写/选择必要的信息（名称、描述等，详见测试数据）",
            "4. 点击提交/保存按钮",
            "5. 确认操作成功，检查生成的唯一标识和返回结果",
        ], [
            "2. 准备非法/重复/不完整的创建数据（详见测试数据）",
            "3. 尝试提交创建请求",
            "4. 检查系统是否拒绝并给出正确提示",
        ]),
        # ── 配置/设置类 ──
        (["配置", "设置", "勾选", "取消"], [
            "2. 进入配置/设置页面",
            "3. 按需求勾选/填写配置项（详见测试数据）",
            "4. 点击保存/应用按钮",
            "5. 确认配置已保存生效，触发后续操作验证效果",
        ], [
            "2. 准备非法/无效的配置（如取消所有必选项、不存在路径）",
            "3. 尝试保存配置",
            "4. 检查系统是否拒绝并给出正确提示",
        ]),
        # ── 编辑/修改/更新类 ──
        (["编辑", "修改", "更新", "变更", "重置"], [
            "2. 进入编辑模式，定位目标记录",
            "3. 修改目标字段为新值（详见测试数据）",
            "4. 点击保存/确认按钮",
            "5. 确认修改已生效，检查相关数据是否同步更新",
        ], [
            "2. 尝试用非法/无效/冲突的数据修改目标字段（详见测试数据）",
            "3. 提交修改",
            "4. 检查系统是否拒绝并给出正确提示",
        ]),
        # ── 删除/取消/作废类 ──
        (["删除", "取消", "作废", "撤销"], [
            "2. 定位待删除/取消的目标数据",
            "3. 执行删除/取消操作",
            "4. 确认删除/取消结果，检查关联数据状态",
        ], [
            "2. 在无权限/数据被引用状态下尝试删除",
            "3. 检查系统是否阻止并给出正确提示",
        ]),
        # ── 查询/搜索/检索/列表类 ──
        (["查询", "搜索", "检索", "列表", "筛选"], [
            "2. 进入查询/检索页面",
            "3. 输入有效的查询关键词/条件（详见测试数据）",
            "4. 执行查询操作",
            "5. 检查返回结果是否正确匹配，排序是否合理",
        ], [
            "2. 输入无效/空/超长查询条件（详见测试数据）",
            "3. 执行查询操作",
            "4. 检查系统是否正确处理（空结果或错误提示）",
        ]),
        # ── 分析/解析/处理类 ──
        (["分析", "解析", "处理", "读取", "识别"], [
            "2. 准备待分析/处理的输入文件或数据（详见测试数据）",
            "3. 执行分析/处理操作",
            "4. 检查输出结果是否符合预期，字段识别是否准确",
        ], [
            "2. 准备无效/空/损坏的输入数据（详见测试数据）",
            "3. 执行分析/处理操作",
            "4. 检查系统是否给出正确提示或容错处理",
        ]),
        # ── 评审/质检/检查类 ──
        (["评审", "质检", "检查", "审查"], [
            "2. 导入待评审/检查的数据文件",
            "3. 启动评审/检查流程",
            "4. 检查评审结果是否准确，评分是否合理",
        ], [
            "2. 导入异常/空/格式错误的文件进行评审",
            "3. 检查系统是否正确处理并提示",
        ]),
        # ── 回灌/写入/同步类 ──
        (["回灌", "写入", "同步", "导入知识"], [
            "2. 准备待回灌/同步的数据文件",
            "3. 执行回灌/同步操作",
            "4. 确认数据已正确写入目标存储，字段完整",
        ], [
            "2. 准备格式不正确/空的文件尝试回灌",
            "3. 检查系统是否拒绝并给出正确提示",
        ]),
        # ── 报告/统计类 ──
        (["报告", "统计", "生成报告"], [
            "2. 选择已执行完成的测试用例文件",
            "3. 执行报告生成操作",
            "4. 检查报告各部分是否完整，数据是否准确",
        ], [
            "2. 选择未执行/空的用例文件生成报告",
            "3. 检查系统是否提示先执行测试",
        ]),
        # ── 集成/串联/全流程类 ──
        (["集成", "串联", "全流程", "恢复", "续跑", "执行"], [
            "2. 准备完整的输入文件和前置输出（详见测试数据）",
            "3. 启动全流程/串联执行",
            "4. 逐步检查各环节输出是否正确，数据流是否打通",
        ], [
            "2. 准备缺失/损坏的输入文件",
            "3. 启动流程执行",
            "4. 检查系统是否在缺失环节报错并提示",
        ]),
        # ── 校验/验证类 ──
        (["校验", "验证", "完整性", "一致性", "连续性"], [
            "2. 准备待校验的数据/文件",
            "3. 执行校验操作",
            "4. 比对实际结果与预期是否一致",
        ], [
            "2. 准备不符合校验规则的数据",
            "3. 执行校验操作",
            "4. 检查系统是否正确识别并提示问题",
        ]),
    ]

    def __match_action(self, title: str, dimension: str):
        """匹配动作模板，返回 (positive_steps, negative_steps) 或 None"""
        for keywords, pos_steps, neg_steps in self.ACTION_TEMPLATES:
            if any(k in title for k in keywords):
                return (pos_steps, neg_steps)
        return None

    def __generate_steps(self, tp: dict) -> str:
        """根据测试点生成具体的测试步骤"""
        title = tp["title"]
        dimension = tp["dimension"]
        module = tp["module"]
        feature = tp["feature"]
        test_data = tp.get("test_data", "")

        # ── 正向测试 ──
        if "正向" in dimension:
            steps = ["1. 准备测试环境，确认前置条件满足"]
            matched = self.__match_action(title, dimension)
            if matched:
                steps.extend(matched[0])
            else:
                # 智能兜底：利用测试点标题和测试数据生成半具体化步骤
                steps.append(f"2. 进入「{module}」模块的「{feature}」功能")
                steps.append(f"3. 按需求执行: {title}")
                if test_data:
                    steps.append(f"   - 测试数据: {test_data}")
                steps.append("4. 检查操作结果与预期是否一致")
            return "\n".join(steps)

        # ── 负向测试 ──
        if "负向" in dimension:
            steps = ["1. 准备测试环境，确保系统处于可测试状态"]
            matched = self.__match_action(title, dimension)
            if matched:
                steps.extend(matched[1])
            else:
                # 智能兜底
                steps.append(f"2. 准备非法/无效条件（详见测试数据: {test_data or '非法数据'}）")
                steps.append(f"3. 执行操作: {title}")
                steps.append("4. 验证系统是否正确拒绝并给出提示")
            return "\n".join(steps)

        # ── 边界测试 ──
        if "边界" in dimension:
            steps = ["1. 确定目标字段的边界值"]
            if any(k in title for k in ["最小值", "最小", "下限", "刚好", "恰好"]):
                steps.append("2. 设置测试数据为边界最小值")
            elif any(k in title for k in ["最大值", "最大", "上限", "超出", "超过"]):
                steps.append("2. 设置测试数据为边界最大值（或略超）")
            elif "超长" in title or "长度" in title:
                steps.append("2. 构造长度在边界值附近（±1）的测试数据")
            elif "数量" in title:
                steps.append("2. 设置数量为边界值（如刚好 N 个）")
            elif "金额" in title:
                steps.append("2. 设置金额为边界值（如 0.01 / 999999.99）")
            elif any(k in title for k in ["时间", "有效期", "过期", "频率", "间隔"]):
                steps.append("2. 设置时间/有效期刚好在边界值（如刚好 N 分钟/N 秒）")
            elif any(k in title for k in ["次数", "第", "锁定", "限制"]):
                steps.append("2. 重复操作至边界次数（如第 N-1 次和第 N 次）")
            elif any(k in title for k in ["空值", "null", "为空", "留空", "不填"]):
                steps.append("2. 将目标字段设置为空值/null/不填写")
                steps.append("3. 提交请求，观察系统是否正确处理空值边界")
            elif any(k in title for k in ["特殊字符", "SQL注入", "XSS", "emoji", "表情"]):
                steps.append("2. 在目标字段中输入特殊字符（如 ' \" < > & | \\ emoji 等）")
                steps.append("3. 提交请求，观察系统是否正确转义/过滤")
            else:
                # 智能兜底：从标题中提取关键信息
                steps.append(f"2. 针对「{feature}」设置边界值测试数据（详见测试数据字段）")
            if test_data:
                steps.append(f"   - 测试数据: {test_data}")
            steps.append("3. 执行操作")
            steps.append("4. 检查系统是否正确处理边界情况")
            return "\n".join(steps)

        # ── 异常测试 ──
        if "异常" in dimension:
            steps = ["1. 准备正常的测试环境"]
            if any(k in title for k in ["超时", "timeout"]):
                steps.append("2. 模拟网络超时/服务无响应场景")
                steps.append("3. 执行操作")
                steps.append("4. 观察系统是否给出超时提示")
            elif any(k in title for k in ["网络", "断网", "中断", "断开"]):
                steps.append("2. 执行操作过程中断开网络连接")
                steps.append("3. 检查系统是否有重试/恢复机制")
            elif any(k in title for k in ["并发", "同时", "并行"]):
                steps.append("2. 使用多线程/多进程模拟并发请求")
                steps.append("3. 检查并发处理结果的一致性")
            elif any(k in title for k in ["编码", "字符集", "乱码", "GBK"]):
                steps.append(f"2. 使用异常编码/字符集的数据: {test_data or '非标编码数据'}")
                steps.append("3. 执行操作")
                steps.append("4. 检查是否正确处理编码问题")
            elif any(k in title for k in ["不可读", "不存在", "损坏", "异常", "失败"]):
                steps.append(f"2. 模拟异常条件: {test_data or '异常条件'}")
                steps.append("3. 执行操作")
                steps.append("4. 验证系统的异常处理机制")
            else:
                # 智能兜底
                steps.append(f"2. 模拟异常场景（详见测试数据: {test_data or '异常场景'}）")
                steps.append(f"3. 执行操作: {title}")
                steps.append("4. 观察系统响应是否合理")
            return "\n".join(steps)

        # ── 性能测试 ──
        if "性能" in dimension:
            steps = ["1. 准备性能测试工具和监控环境"]
            if "并发" in title or "高并发" in title or "大量" in title:
                steps.append("2. 使用性能测试工具（如 JMeter/Locust）模拟并发用户")
                steps.append("3. 按测试数据设置并发数量和脚本参数")
                steps.append("4. 执行测试，记录响应时间、TPS、错误率等指标")
                steps.append("5. 与性能基线对比，判断是否达标")
            elif "响应时间" in title or "延迟" in title:
                steps.append("2. 单次执行目标操作")
                steps.append("3. 记录从请求发起到收到响应的时间")
                steps.append("4. 重复执行 N 次（N>=10），计算平均响应时间")
                steps.append("5. 判断是否满足性能要求")
            elif "效率" in title or "生成" in title or "耗时" in title:
                steps.append(f"2. 准备较大规模输入数据（详见测试数据: {test_data or '大规模数据'}）")
                steps.append("3. 执行目标操作并计时")
                steps.append("4. 记录完成耗时，判断是否满足时间要求")
            else:
                steps.append(f"2. 针对「{feature}」准备性能测试场景")
                steps.append(f"3. 执行操作: {title}")
                steps.append("4. 记录性能指标数据，判断是否达标")
            return "\n".join(steps)

        # ── 安全测试 ──
        if "安全" in dimension:
            steps = ["1. 安全测试准备"]
            if any(k in title for k in ["越权", "未授权", "权限"]):
                steps.append("2. 使用低权限用户 A 的凭证登录")
                steps.append("3. 尝试访问/操作高权限用户 B 的数据或功能")
                steps.append("4. 验证系统是否阻止越权访问并给出提示")
            elif any(k in title for k in ["注入", "SQL", "XSS", "脚本", "xss"]):
                steps.append(f"2. 在输入框/参数中注入恶意载荷: {test_data or 'SQL/XSS 注入 payload'}")
                steps.append("3. 提交请求")
                steps.append("4. 检查是否被拦截或过滤，不泄露原始错误信息")
            elif any(k in title for k in ["篡改", "伪造"]):
                steps.append("2. 使用抓包工具（如 Burp Suite/Charles）拦截请求")
                steps.append("3. 篡改关键参数（如金额、用户ID、token）")
                steps.append("4. 发送篡改后的请求")
                steps.append("5. 验证服务端是否校验了数据的合法性")
            elif any(k in title for k in ["敏感", "泄露", "加密"]):
                steps.append("2. 检查 API 响应和页面渲染是否包含敏感信息")
                steps.append("3. 检查敏感字段在传输和存储时是否加密")
                steps.append("4. 验证日志中是否过滤了敏感数据")
            elif any(k in title for k in ["认证", "身份", "Token", "token", "暴力", "破解", "封禁", "频率"]):
                steps.append(f"2. 准备安全测试攻击载荷/场景（详见测试数据: {test_data or '攻击载荷'}）")
                steps.append(f"3. 执行攻击操作: {title}")
                steps.append("4. 验证系统的安全防护机制是否触发")
            else:
                steps.append(f"2. 针对「{feature}」执行安全测试")
                steps.append(f"3. 操作: {title}")
                steps.append("4. 验证系统的安全防护机制")
            return "\n".join(steps)

        # 兜底
        return f"1. 准备测试环境\n2. 执行操作: {title}\n3. 验证结果与预期一致"

    def __generate_precondition(self, tp: dict) -> str:
        """生成前置条件"""
        title = tp["title"]
        dimension = tp["dimension"]

        # 根据模块和维度生成前置条件
        if any(k in title for k in ["登录", "注册", "认证"]):
            return "用户处于登录/注册页面"

        if any(k in title for k in ["查询", "列表", "搜索", "筛选", "搜索"]):
            return "用户已登录，系统中有可查询的数据"

        if any(k in title for k in ["创建", "新增", "添加", "提交", "上传"]):
            return "用户已登录，具备创建/新增权限"

        if any(k in title for k in ["编辑", "修改", "更新", "变更"]):
            return "用户已登录，有可编辑的数据记录"

        if any(k in title for k in ["删除", "取消", "作废"]):
            return "用户已登录，有待删除/取消的数据记录"

        if any(k in title for k in ["导出", "下载"]):
            return "用户已登录，有可导出的数据"

        if any(k in title for k in ["校验", "验证", "检查", "对比"]):
            return "基准数据和待校验数据已准备就绪"

        if any(k in title for k in ["集成", "串联"]):
            return "前置环节已完成，相关服务运行正常"

        if any(k in title for k in ["性能", "并发", "效率"]):
            return "性能测试环境已就绪，监控工具已配置"

        # 异常/负向/安全
        if "异常" in dimension or "安全" in dimension or "负向" in dimension:
            return "测试环境已就绪，准备好测试工具和数据"

        if "边界" in dimension:
            return "测试环境已就绪，确定好边界值参数"

        # 默认
        return "用户已登录系统，测试环境正常可用"

    def generate(self, test_points: list) -> list:
        """生成测试用例"""
        test_cases = []
        for i, tp in enumerate(test_points, 1):
            test_cases.append({
                "id": f"TC-{i:03d}",
                "module": tp["module"],
                "feature": tp["feature"],
                "dimension": tp["dimension"],
                "title": tp["title"],
                "priority": assign_priority(tp),
                "precondition": self.__generate_precondition(tp),
                "steps": self.__generate_steps(tp),
                "test_data": tp["test_data"],
                "expected": tp["expected"],
            })
        return test_cases

# ═══════════════════════════════════════════════════════════════
# Excel 写入器
# ═════════════════════════════════

class ExcelWriter:
    """写入 Excel 文件，带格式化"""

    # 表头定义
    HEADERS = [
        ("用例编号", 12),
        ("所属模块", 16),
        ("功能点", 20),
        ("测试维度", 12),
        ("用例标题", 30),
        ("优先级", 10),
        ("前置条件", 25),
        ("测试步骤", 45),
        ("测试数据", 25),
        ("预期结果", 40),
        ("备注", 15),
        ("执行结果", 12),
    ]

    # 优先级颜色
    PRIORITY_FILLS = {
        "P0": PatternFill(start_color="FFCDD2", fill_type="solid"),  # 浅红
        "P1": PatternFill(start_color="FFE0B2", fill_type="solid"),  # 浅橙
        "P2": PatternFill(start_color="FFF9C4", fill_type="solid"),  # 浅黄
    }

    def write(self, test_cases: list, output_path: str):
        """写入 Excel 文件，带格式化"""
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # ── 表头 ──
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F4F4F", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, (header, width) in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 30

        # ── 数据行 ──
        data_font = Font(name="微软雅黑", size=10)
        for row_idx, tc in enumerate(test_cases, 2):
            row_data = [
                tc["id"], tc["module"], tc["feature"], tc["dimension"],
                tc["title"], tc["priority"], tc["precondition"],
                tc["steps"], tc["test_data"], tc["expected"], "", ""
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

            # 优先级着色
            if tc["priority"] in self.PRIORITY_FILLS:
                ws.cell(row=row_idx, column=6).fill = self.PRIORITY_FILLS[tc["priority"]]

            # 行高根据步骤数调整
            step_count = tc["steps"].count("\n") + 1
            ws.row_dimensions[row_idx].height = max(30, step_count * 18)

        # ── 冻结首行 ──
        ws.freeze_panes = "A2"

        # ── 自动筛选 ──
        last_col = get_column_letter(len(self.HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{len(test_cases)+1}"

        wb.save(output_path)


# ═════════════════════════════════════════ 计数助手 ═══════════════════════════

def count_stats(test_cases: list) -> dict:
    """统计用例分布"""
    stats = {
        "total": len(test_cases),
        "modules": set(),
        "features": set(),
        "dimensions": {},
        "priorities": {},
    }
    for tc in test_cases:
        stats["modules"].add(tc["module"])
        stats["features"].add(tc["feature"])
        stats["dimensions"][tc["dimension"]] = stats["dimensions"].get(tc["dimension"], 0) + 1
        stats["priorities"][tc["priority"]] = stats["priorities"].get(tc["priority"], 0) + 1
    # 转为可序列化格式（set → sorted list）
    stats["modules"] = sorted(stats["modules"])
    stats["features"] = sorted(stats["features"])
    return stats


# ═══════════════ SkillScript ┐ ═══════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="生成 Excel 格式测试用例")
    parser.add_argument("input", help="测试点 Markdown 文件路径")
    parser.add_argument("-o", "--output", default="testcases.xlsx", help="输出文件路径")
    parser.add_argument(
        "-d", "--dimensions",
        default="all",
        help="测试维度过滤: all|basic|positive,negative,boundary,exception,performance,security"
    )
    parser.add_argument(
        "-p", "--priority", default="all",
        help="优先级过滤: all|P0|P0,P1|P1,P2 等"
    )
    args = parser.parse_args()

    # 1. 解析测试点
    print(f"📖 读取测试点文件: {args.input}")
    test_point_parser = TestPointParser()
    test_points = test_point_parser.parse_file(args.input)

    if not test_points:
        print("⚠️  未解析到任何测试点，请检查文件格式", file=sys.stderr)
        return 1

    print(f"✅ 解析到 {len(test_points)} 个测试点")

    # 2. 过滤
    if args.dimensions != "all":
        test_points = filter_by_dimensions(test_points, args.dimensions)
        print(f"🔍 维度过滤后剩余 {len(test_points)} 个测试点")

    # 3. 生成用例
    print("🔨 生成测试用例...")
    generator = TestCaseGenerator()
    test_cases = generator.generate(test_points)

    if args.priority != "all":
        test_cases = filter_by_priorities(test_cases, args.priority)
        print(f"🔍 优先级过滤后剩余 {len(test_cases)} 个用例")

    # 4. 统计
    stats = count_stats(test_cases)

    print("\n✅ 测试用例生成完成！")
    print("📊 统计信息：")
    print(f"  - 测试模块：{len(stats['modules'])} 个")
    print(f"  - 功能点：{len(stats['features'])} 个")
    print(f"  - 用例总数：{stats['total']} 个")
    print("\n📊 测试维度分布：")
    for dim, count in stats["dimensions"].items():
        print(f"  - {dim}: {count} 个")
    print("\n📊 优先级分布：")
    for pri in ["P0", "P1", "P2"]:
        count = stats["priorities"].get(pri, 0)
        print(f"  - {pri}: {count} 个")

    # 5. 写入 Excel
    print(f"\n📁 写入文件: {args.output}")
    excel_writer = ExcelWriter()
    excel_writer.write(test_cases, args.output)

    print(f"\n✅ 文件已保存: {args.output}")
    return 0

if __name__ == "__main__":
    sys.exit(main())
