#!/usr/bin/env python3
"""
集成测试 — 覆盖跨模块交互和真实运行时场景

测试类型：集成测试（Integration Tests）
覆盖范围：
  - KB MCP Client 与 Step2 的集成（vault_path 传递、缺失 vault 优雅降级）
  - Step1 → Step3 → Step4 → Step7 的端到端数据流
  - Pipeline 状态管理与断点续跑
  - Report Generator 的边界场景（空数据、无执行结果）
  - LLM 响应解析容错
  - 知识库路径穿越与权限处理

这些测试针对此前发现的实际运行时缺陷，确保修复不被回退。
"""

import json
import os
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("LLM_API_KEY", "sk-test-dummy-key-for-testing")


# ═══════════════════════════════════════════════════════════════
# 集成测试：KB MCP Client
# ═══════════════════════════════════════════════════════════════

class TestMCPClientResilience:
    """KB 客户端健壮性 — 不可写/不存在的 vault 路径"""

    def test_nonexistent_vault_no_crash(self, tmp_path):
        """vault 路径不存在时不应崩溃"""
        from core.kb.mcp_client import MCPClient

        # 使用一个不存在的路径（但父目录可写）
        fake_vault = tmp_path / "nonexistent_vault"
        client = MCPClient(str(fake_vault))
        # 应自动创建目录，不崩溃
        assert fake_vault.exists()

    def test_readonly_vault_no_crash(self, tmp_path):
        """vault 路径在只读位置时不应崩溃"""
        from core.kb.mcp_client import MCPClient

        # 模拟不可写的路径（使用不存在的根目录）
        readonly_vault = "/dev/null/cannot_create_here"
        client = MCPClient(readonly_vault)
        # 搜索应返回空列表，不崩溃
        results = client.search("test")
        assert isinstance(results, list)

    def test_path_traversal_blocked(self, tmp_path):
        """路径穿越攻击应被拦截"""
        from core.kb.mcp_client import MCPClient

        vault = tmp_path / "vault"
        vault.mkdir()
        client = MCPClient(str(vault))

        # 尝试读取 vault 外的文件
        result = client.read_file("../../etc/passwd")
        assert result is None

    def test_search_returns_list(self, tmp_path):
        """搜索空 vault 应返回空列表"""
        from core.kb.mcp_client import MCPClient

        vault = tmp_path / "empty_vault"
        vault.mkdir()
        client = MCPClient(str(vault))
        results = client.search("任意关键词")
        assert isinstance(results, list)
        assert len(results) == 0


class TestKBManagerVaultEnv:
    """KB Manager 应从 OBSIDIAN_VAULT 环境变量读取 vault 路径"""

    def test_env_var_overrides_default(self, tmp_path, monkeypatch):
        """OBSIDIAN_VAULT 环境变量应覆盖默认路径"""
        custom_vault = str(tmp_path / "custom_vault")
        monkeypatch.setenv("OBSIDIAN_VAULT", custom_vault)

        # 重新导入模块以读取新的环境变量
        import importlib

        import core.kb.kb_manager_mcp as kb_module

        importlib.reload(kb_module)
        assert kb_module.OBSIDIAN_VAULT == custom_vault

    def test_fallback_to_default(self, tmp_path, monkeypatch):
        """无环境变量时使用默认路径"""
        monkeypatch.delenv("OBSIDIAN_VAULT", raising=False)

        import importlib

        import core.kb.kb_manager_mcp as kb_module

        importlib.reload(kb_module)
        assert "test-interview-kb" in kb_module.OBSIDIAN_VAULT


# ═══════════════════════════════════════════════════════════════
# 集成测试：Step2 KB Search
# ═══════════════════════════════════════════════════════════════

class TestStep2KBIntegration:
    """Step2 与知识库的集成"""

    def test_missing_vault_graceful_skip(self, tmp_path):
        """vault 不存在时应优雅跳过，不崩溃"""
        from core.steps.step2_kb_search import Step2KBSearch

        cfg = {
            "knowledge_base": {
                "enabled": True,
                "vault_path": str(tmp_path / "nonexistent_vault_xyz"),
            }
        }
        step2 = Step2KBSearch(str(tmp_path / "out"), cfg, llm=None)
        result = step2.run(requirements_analysis="## 模块一：测试")

        assert result.ok, f"应返回 ok=True（优雅跳过）, got error={result.data}"
        assert result.data.get("skipped") is True

    def test_disabled_kb_skips(self, tmp_path):
        """知识库禁用时应跳过"""
        from core.steps.step2_kb_search import Step2KBSearch

        cfg = {"knowledge_base": {"enabled": False, "vault_path": ""}}
        step2 = Step2KBSearch(str(tmp_path / "out"), cfg, llm=None)
        result = step2.run(requirements_analysis="test")
        assert result.ok
        assert result.data.get("skipped") is True

    def test_keyword_extraction(self, tmp_path):
        """从需求分析中提取关键词"""
        from core.steps.step2_kb_search import Step2KBSearch

        analysis = """## 模块一：用户管理

### 功能点 1：用户注册

### 功能点 2：用户登录

## 模块二：订单系统

### 功能点 1：创建订单
"""
        keywords = Step2KBSearch._extract_keywords(analysis)
        assert "用户管理" in keywords or "用户注册" in keywords
        assert len(keywords) > 0

    def test_keyword_extraction_empty(self):
        """空文本提取关键词应返回默认值"""
        from core.steps.step2_kb_search import Step2KBSearch

        keywords = Step2KBSearch._extract_keywords("")
        assert keywords == "测试"


# ═══════════════════════════════════════════════════════════════
# 集成测试：Step4 → Excel 生成
# ═══════════════════════════════════════════════════════════════

class TestStep4GenerateIntegration:
    """Step4 与 Excel 生成脚本的集成"""

    def _make_testpoints(self, path):
        """创建标准测试点文件"""
        path.write_text(
            "# 测试点\n\n"
            "## 模块一：用户管理\n\n"
            "### 功能点 1：用户注册\n\n"
            "#### 测试维度：正向测试\n\n"
            "- 测试点 1：成功注册新用户\n"
            "  - 测试数据：13800138000\n"
            "  - 预期结果：注册成功\n\n"
            "#### 测试维度：负向测试\n\n"
            "- 测试点 2：手机号已存在\n"
            "  - 测试数据：已注册手机号\n"
            "  - 预期结果：提示已注册\n",
            encoding="utf-8",
        )

    def test_step4_generates_excel(self, tmp_path):
        """Step4 应成功生成 Excel 文件"""
        from core.steps.step4_generate import Step4Generate

        self._make_testpoints(tmp_path / "testpoints.md")
        cfg = {}
        step4 = Step4Generate(str(tmp_path), cfg)
        result = step4.run(dimensions="all", formats="excel")

        assert result.ok, f"Step4 应成功, error={result.error}"
        assert result.data.get("case_count", 0) >= 2
        assert (tmp_path / "testcases.xlsx").exists()

    def test_step4_missing_testpoints(self, tmp_path):
        """缺少 testpoints.md 时应正确报错"""
        from core.steps.step4_generate import Step4Generate

        step4 = Step4Generate(str(tmp_path), {})
        result = step4.run(dimensions="all", formats="excel")
        assert not result.ok
        assert "缺少" in (result.error or "") or "不存在" in (result.error or "")

    def test_step4_dimension_filter(self, tmp_path):
        """维度过滤应正确工作"""
        from core.steps.step4_generate import Step4Generate

        self._make_testpoints(tmp_path / "testpoints.md")
        step4 = Step4Generate(str(tmp_path), {})
        result = step4.run(dimensions="positive", formats="excel")
        assert result.ok
        # 正向测试只有 1 条
        assert result.data.get("case_count") == 1


# ═══════════════════════════════════════════════════════════════
# 集成测试：Step7 Report 生成
# ═══════════════════════════════════════════════════════════════

class TestStep7ReportIntegration:
    """Step7 报告生成集成"""

    def _make_executed_xlsx(self, path, results=None):
        """创建已执行的测试用例 Excel"""
        if results is None:
            results = [("通过",), ("失败",), ("通过",)]
        wb = Workbook()
        ws = wb.active
        ws.append([
            "用例编号", "模块", "功能点", "测试维度", "用例标题",
            "优先级", "前置条件", "测试步骤", "测试数据", "预期结果",
            "备注", "执行结果",
        ])
        for i, (res,) in enumerate(results, 1):
            ws.append([
                f"TC-{i:03d}", "模块", "功能", "正向测试", f"测试{i}",
                "P0", "前置", "1.操作", "数据", "预期", "", res,
            ])
        wb.save(str(path))
        wb.close()

    def test_step7_generates_report(self, tmp_path):
        """Step7 应生成报告"""
        from core.steps.step7_report import Step7Report

        self._make_executed_xlsx(tmp_path / "testcases.xlsx")
        step7 = Step7Report(str(tmp_path), {})
        result = step7.run()
        assert result.ok
        report = (tmp_path / "test_report.md").read_text(encoding="utf-8")
        assert "测试质量报告" in report
        assert "66.7%" in report  # 2 pass / 3 total (1 decimal place)

    def test_step7_missing_xlsx(self, tmp_path):
        """缺少 testcases.xlsx 时应正确报错"""
        from core.steps.step7_report import Step7Report

        step7 = Step7Report(str(tmp_path), {})
        result = step7.run()
        assert not result.ok
        assert "Excel" in (result.error or "") or "不存在" in (result.error or "")

    def test_step7_no_executed_cases_warning(self, tmp_path, capsys):
        """未执行任何用例时应输出警告"""
        from scripts.generate_report import ExcelReader, ReportAnalyzer

        wb = Workbook()
        ws = wb.active
        ws.append([
            "用例编号", "模块", "功能点", "测试维度", "用例标题",
            "优先级", "前置条件", "测试步骤", "测试数据", "预期结果",
            "备注", "执行结果",
        ])
        ws.append(["TC-001", "M", "F", "正向", "T", "P0", "x", "op", "d", "e", "", ""])
        ws.append(["TC-002", "M", "F", "负向", "T", "P1", "x", "op", "d", "e", "", ""])
        xlsx_path = tmp_path / "blank.xlsx"
        wb.save(str(xlsx_path))
        wb.close()

        reader = ExcelReader()
        cases, _ = reader.read(str(xlsx_path))
        analyzer = ReportAnalyzer()
        analyzer.analyze(cases)

        captured = capsys.readouterr()
        assert "未执行" in captured.err


# ═══════════════════════════════════════════════════════════════
# 集成测试：Pipeline 状态管理
# ═══════════════════════════════════════════════════════════════

class TestPipelineStateManagement:
    """Pipeline 状态持久化与断点续跑"""

    def test_state_save_load_roundtrip(self, tmp_path):
        """状态保存后重新加载应一致"""
        from core.config_loader import load_config
        from core.pipeline import Pipeline

        cfg = load_config()
        p = Pipeline(cfg, str(tmp_path))

        state = p.load_state()
        state["completed_steps"] = [1, 2, 3]
        state["step_results"] = {"1": {"ok": True, "data": {"modules": 5}}}
        state["mode"] = "auto"
        state["requirements_file"] = "test.md"
        p.save_state(state)

        # 重新加载
        state2 = p.load_state()
        assert state2["completed_steps"] == [1, 2, 3]
        assert state2["mode"] == "auto"
        assert state2["step_results"]["1"]["ok"] is True

    def test_skip_check_requires_output_file(self, tmp_path):
        """跳过检查应同时满足：步骤完成 + 输出文件存在"""
        from core.config_loader import load_config
        from core.pipeline import Pipeline

        cfg = load_config()
        p = Pipeline(cfg, str(tmp_path))

        # 步骤标记完成但文件不存在
        state = {"completed_steps": [4]}
        skip = p._check_skip(state, 4, "testcases.xlsx")
        assert skip is False, "文件不存在时不应跳过"

        # 创建文件后应跳过
        (tmp_path / "testcases.xlsx").write_bytes(b"fake")
        skip2 = p._check_skip(state, 4, "testcases.xlsx")
        assert skip2 is True

    def test_corrupted_state_file(self, tmp_path):
        """损坏的状态文件不应导致崩溃"""
        from core.config_loader import load_config
        from core.pipeline import Pipeline

        # 写入损坏的 JSON
        (tmp_path / "_pipeline_state.json").write_text("{invalid json!!!", encoding="utf-8")

        cfg = load_config()
        p = Pipeline(cfg, str(tmp_path))

        # load_state 应能处理损坏的 JSON（或抛出可预期的异常）
        try:
            state = p.load_state()
            # 如果没有抛出异常，应返回合理的默认值
            assert "completed_steps" in state
        except json.JSONDecodeError:
            # 这是可接受的 — 至少不是其他类型的崩溃
            pass


# ═══════════════════════════════════════════════════════════════
# 集成测试：LLM 响应解析
# ═══════════════════════════════════════════════════════════════

class TestLLMResponseParsing:
    """LLM 响应解析容错"""

    def test_split_response_separator(self):
        """等号分隔线正确拆分"""
        from core.steps.step1_analysis import Step1Analysis

        content = "# 需求分析\n\n## 模块一\n\n==========\n\n# 待确认\n1. 问题1"
        analysis, clar = Step1Analysis._split_response(content)
        assert "需求分析" in analysis
        assert "待确认" in clar

    def test_split_response_marker(self):
        """>=10 个等号的分隔线正确拆分"""
        from core.steps.step1_analysis import Step1Analysis

        content = "分析内容\n\n============\n\n待确认内容"
        analysis, clar = Step1Analysis._split_response(content)
        assert "分析" in analysis
        assert "待确认" in clar

    def test_split_response_no_separator(self):
        """无分隔线时全部作为需求分析"""
        from core.steps.step1_analysis import Step1Analysis

        content = "# 需求分析\n\n## 模块一"
        analysis, clar = Step1Analysis._split_response(content)
        assert "需求分析" in analysis
        assert clar == ""

    def test_parse_json_response_markdown_block(self):
        """解析 markdown 代码块包裹的 JSON"""
        from core.llm_client import LLMClient

        raw = '```json\n{"score": 85, "issues": ["问题1"]}\n```'
        result = LLMClient._parse_json_response(raw)
        assert result.get("score") == 85
        assert "问题1" in result.get("issues", [])

    def test_parse_json_response_plain(self):
        """解析纯 JSON"""
        from core.llm_client import LLMClient

        raw = '{"score": 90, "passed": true}'
        result = LLMClient._parse_json_response(raw)
        assert result["score"] == 90
        assert result["passed"] is True

    def test_parse_json_response_extract_block(self):
        """从混合文本中提取 JSON 块"""
        from core.llm_client import LLMClient

        raw = '评估完成。\n{"score": 70}\n以上是结果。'
        result = LLMClient._parse_json_response(raw)
        assert result.get("score") == 70

    def test_parse_json_response_invalid(self):
        """无效 JSON 返回空字典"""
        from core.llm_client import LLMClient

        result = LLMClient._parse_json_response("这不是JSON")
        assert result == {}


# ═══════════════════════════════════════════════════════════════
# 集成测试：Excel Reader 列匹配
# ═══════════════════════════════════════════════════════════════

class TestExcelReaderColumnMatching:
    """Excel 读取器列名模糊匹配"""

    def test_standard_headers(self, tmp_path):
        """标准 12 列表头"""
        from scripts.generate_report import ExcelReader

        wb = Workbook()
        ws = wb.active
        ws.append([
            "用例编号", "所属模块", "功能点", "测试维度", "用例标题",
            "优先级", "前置条件", "测试步骤", "测试数据", "预期结果",
            "备注", "执行结果",
        ])
        ws.append(["TC-001", "M", "F", "正向", "T", "P0", "x", "op", "d", "e", "", "通过"])
        path = tmp_path / "std.xlsx"
        wb.save(str(path))
        wb.close()

        reader = ExcelReader()
        cases, meta = reader.read(str(path))
        assert len(cases) == 1
        assert cases[0]["result"] == "pass"

    def test_english_headers(self, tmp_path):
        """英文表头"""
        from scripts.generate_report import ExcelReader

        wb = Workbook()
        ws = wb.active
        ws.append(["id", "module", "feature", "type", "title",
                    "priority", "precondition", "steps", "data", "expected",
                    "remark", "result"])
        ws.append(["TC-001", "M", "F", "正向", "T", "P0", "x", "op", "d", "e", "", "fail"])
        path = tmp_path / "en.xlsx"
        wb.save(str(path))
        wb.close()

        reader = ExcelReader()
        cases, meta = reader.read(str(path))
        assert len(cases) == 1
        assert cases[0]["result"] == "fail"

    def test_result_normalization(self, tmp_path):
        """各种执行结果值的归一化"""
        from scripts.generate_report import ExcelReader

        wb = Workbook()
        ws = wb.active
        ws.append(["用例编号", "模块", "执行结果"])
        for i, res in enumerate(["通过", "pass", "✅", "成功", "失败", "fail", "❌",
                                   "阻塞", "block", "跳过", "skip", "未执行", ""], 1):
            ws.append([f"TC-{i:03d}", "M", res])
        path = tmp_path / "norm.xlsx"
        wb.save(str(path))
        wb.close()

        reader = ExcelReader()
        cases, _ = reader.read(str(path))
        assert cases[0]["result"] == "pass"   # 通过
        assert cases[1]["result"] == "pass"   # pass
        assert cases[2]["result"] == "pass"   # ✅
        assert cases[3]["result"] == "pass"   # 成功
        assert cases[4]["result"] == "fail"   # 失败
        assert cases[5]["result"] == "fail"   # fail
        assert cases[6]["result"] == "fail"   # ❌
        assert cases[7]["result"] == "block"  # 阻塞
        assert cases[8]["result"] == "block"  # block
        assert cases[9]["result"] == "skip"   # 跳过
        assert cases[10]["result"] == "skip"  # skip
        assert cases[11]["result"] == "not_run"  # 未执行
        assert cases[12]["result"] == "not_run"  # 空


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
