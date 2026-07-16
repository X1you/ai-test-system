#!/usr/bin/env python3
"""
测试数据生成器 — 生成多样化测试输入数据。

生成内容：
  - 需求文档（Markdown）：多种业务场景、格式变体
  - 测试点文件（Markdown）：标准/非标准/边缘格式
  - Excel 测试用例：已执行/未执行/混合结果/特殊字符
  - 配置文件变体：正常/缺失字段/无效值
"""

import random
import string
from pathlib import Path

from openpyxl import Workbook

# ═══════════════════════════════════════════════════════════════
# 需求文档模板（8 种业务场景）
# ═══════════════════════════════════════════════════════════════

REQUIREMENT_TEMPLATES = {

    "user_management": """# 用户管理系统需求文档

## 1. 用户注册模块

### 1.1 注册方式
用户可以通过手机号注册，需要验证短信验证码。

### 1.2 注册流程
1. 用户输入手机号
2. 点击"获取验证码"
3. 输入短信验证码
4. 设置密码（8-20位，需包含字母和数字）
5. 完成注册

### 1.3 密码要求
密码长度 8-20 位，必须包含大小写字母和数字。

## 2. 用户登录模块

### 2.1 癆录方式
- 账号密码登录
- 手机验证码登录
- 第三方登录（微信、支付宝）

### 2.2 登录限制
连续输错密码 5 次锁定账号 30 分钟。

## 3. 权限管理

### 3.1 角色体系
系统分为超级管理员、普通管理员、普通用户三级。

### 3.2 权限分配
管理员可以分配和回收用户权限。
""",

    "order_system": """# 订单管理系统需求

## 1. 订单创建

### 1.1 创建流程
用户选择商品 → 填写收货地址 → 选择支付方式 → 提交订单。

### 1.2 订单号规则
订单号格式：yyyyMMdd + 6位流水号，如 20260716000001。

### 1.3 库存校验
下单时实时校验库存，库存不足时提示并阻止下单。

## 2. 订单状态

### 2.1 状态流转
待支付 → 已支付 → 已发货 → 已签收 → 已完成
待支付 → 已取消
已支付 → 退款中 → 已退款

### 2.2 超时取消
订单创建后 30 分钟未支付自动取消。

## 3. 退款管理

### 3.1 退款条件
已支付订单可以申请退款，需管理员审核。

### 3.2 退款金额
按实际支付金额原路退回。
""",

    "payment": """# 支付系统需求

## 1. 支付方式

### 1.1 支持渠道
微信支付、支付宝支付、银行卡支付、余额支付。

### 1.2 支付限额
单笔最低 0.01 元，最高 50000 元。
日累计限额 100000 元。

## 2. 支付安全

### 2.1 交易加密
所有支付请求使用 HTTPS + AES-256 加密。

### 2.2 防重放
每笔交易携带唯一 nonce，5 分钟内不可重复。

### 2.3 风控规则
异常交易（高频、大额）触发人工审核。

## 3. 对账

### 3.1 自动对账
每日凌晨自动对账，差异订单生成报告。

### 3.2 差异处理
T+1 完成差异订单处理。
""",

    "content_management": """# 内容管理系统需求

## 1. 文章管理

### 1.1 发布流程
创建草稿 → 编辑内容 → 预览 → 发布 → 下线。

### 1.2 内容审核
发布前需通过敏感词过滤和人工审核。

### 1.3 富文本编辑
支持 Markdown 和富文本两种编辑模式。

## 2. 评论系统

### 2.1 评论规则
用户登录后可评论，每条评论最多 500 字。

### 2.2 评论审核
首次评论需审核，后续评论自动通过（信誉积分制）。

## 3. 分类与标签

### 3.1 分类管理
支持三级分类树，每篇文章归属一个主分类。

### 3.2 标签系统
每篇文章最多 5 个标签，标签自动聚合。
""",

    "inventory": """# 库存管理系统需求

## 1. 入库管理

### 1.1 采购入库
创建采购单 → 到货验收 → 入库登记 → 库存更新。

### 1.2 入库单号
RK + yyyyMMdd + 4位流水号。

## 2. 出库管理

### 2.1 销售出库
订单驱动出库，拣货 → 复核 → 打包 → 发货。

### 2.2 出库校验
出库时校验库存可用量，不足时阻断。

## 3. 库存盘点

### 3.1 定期盘点
每月全盘，每日抽盘（高频SKU）。

### 3.2 盘点差异
差异 > 1% 触发复盘流程，差异 > 5% 升级处理。

## 4. 库存预警

### 4.1 低库存预警
库存低于安全水位时推送预警通知。

### 4.2 滞销预警
90 天无动销的 SKU 标记为滞销。
""",

    "notification": """# 消息通知系统需求

## 1. 通知渠道

### 1.1 支持渠道
站内信、短信、邮件、APP推送、企业微信。

### 1.2 渠道优先级
用户可配置渠道优先级，系统按序尝试发送。

## 2. 模板管理

### 2.1 模板变量
支持 {{username}}、{{order_id}} 等变量占位符。

### 2.2 模板审核
短信和邮件模板需服务商审核通过后使用。

## 3. 发送策略

### 3.1 重试机制
发送失败自动重试 3 次，间隔 1min/5min/30min。

### 3.2 频率限制
同一用户同一模板，1 小时内最多 5 条。
""",

    "data_export": """# 数据导出系统需求

## 1. 导出功能

### 1.1 导出格式
支持 Excel、CSV、PDF 三种格式。

### 1.2 异步导出
数据量 > 10000 行时使用异步任务，完成后邮件通知下载。

## 2. 数据范围

### 2.1 时间筛选
支持按创建时间、更新时间筛选。

### 2.2 字段选择
用户可自定义导出字段列。

## 3. 权限控制

### 3.1 数据隔离
普通用户只能导出本人数据，管理员可导出全部。

### 3.2 敏感字段
手机号、身份证号默认脱敏导出（显示前3后4）。
""",

    "authentication": """# 认证授权系统需求

## 1. 登录认证

### 1.1 认证方式
用户名密码、手机验证码、LDAP、SSO单点登录。

### 1.2 密码策略
密码长度 12-32 位，需包含大小写字母、数字、特殊字符。
密码 90 天强制更换，不可与最近 5 次重复。

## 2. 会话管理

### 2.1 Token 机制
JWT Token 有效期 2 小时，Refresh Token 有效期 7 天。

### 2.2 多设备登录
最多 5 个设备同时在线，超出自动踢出最早的。

## 3. 权限控制

### 3.1 RBAC 模型
基于角色的访问控制，用户 → 角色 → 权限三级映射。

### 3.2 数据权限
行级权限（部门数据隔离）+ 列级权限（敏感字段隐藏）。

## 4. 安全审计

### 4.1 登录日志
记录所有登录尝试（成功/失败），保留 180 天。

### 4.2 异常检测
异地登录、新设备登录触发二次验证。
""",
}


# ═══════════════════════════════════════════════════════════════
# 测试点 Markdown 生成器
# ═══════════════════════════════════════════════════════════════

def generate_testpoints_standard(modules: list, points_per_module: int = 5) -> str:
    """生成标准格式测试点文件"""
    lines = ["# 测试点清单\n"]
    dim_cn = ["正向测试", "负向测试", "边界测试", "异常测试", "安全测试"]

    for i, module in enumerate(modules, 1):
        cn_num = "一二三四五六七八九十"[i - 1] if i <= 10 else str(i)
        lines.append(f"\n## 模块{cn_num}：{module}\n")
        lines.append(f"### 功能点 1：{module}核心功能\n")

        for j in range(points_per_module):
            dim = dim_cn[j % len(dim_cn)]
            lines.append(f"\n#### 测试维度：{dim}\n")
            lines.append(f"- 测试点 {j + 1}：{module}的{dim}场景测试")
            lines.append(f"  - 测试数据：{module}测试数据{j + 1}")
            lines.append(f"  - 预期结果：系统正确处理{dim}场景")
    return "\n".join(lines)


def generate_testpoints_variant_no_prefix(modules: list) -> str:
    """生成无 '测试维度：' 前缀的测试点变体"""
    lines = ["# 测试点\n"]
    for module in modules[:3]:
        lines.append(f"\n## 模块一：{module}\n")
        lines.append("### 功能点 1：功能\n")
        lines.append("\n#### 正向测试\n")
        lines.append(f"- 测试点 1：测试{module}")
        lines.append("  - 测试数据：数据")
        lines.append("  - 预期结果：结果")
    return "\n".join(lines)


def generate_testpoints_mixed_colons(modules: list) -> str:
    """生成半角/全角冒号混合的测试点"""
    lines = ["# 测试点\n"]
    for i, module in enumerate(modules[:3], 1):
        colon = ":" if i % 2 == 0 else "："
        lines.append(f"\n## 模块{i}{colon} {module}\n")
        lines.append(f"### 功能点 1{colon} 功能\n")
        lines.append(f"\n#### 测试维度{colon} 正向测试\n")
        lines.append(f"- 测试点 1{colon} 测试{module}")
    return "\n".join(lines)


def generate_testpoints_with_special_chars() -> str:
    """包含特殊字符的测试点"""
    return """# 特殊字符测试点

## 模块一：🎉特殊字符模块

### 功能点 1：emoji和unicode测试

#### 测试维度：正向测试

- 测试点 1：处理 emoji 🚀⚡🔥
  - 测试数据：café résumé naïve 日本語 한국어
  - 预期结果：正确处理多语言和特殊字符

#### 测试维度：安全测试

- 测试点 2：SQL注入防护
  - 测试数据：' OR '1'='1'; DROP TABLE users; --
  - 预期结果：注入被拦截

- 测试点 3：XSS防护
  - 测试数据：<script>alert('xss')</script>
  - 预期结果：脚本被过滤
"""


# ═══════════════════════════════════════════════════════════════
# Excel 测试用例生成器
# ═══════════════════════════════════════════════════════════════

def generate_executed_excel(
    path: str,
    case_count: int = 30,
    pass_rate: float = 0.8,
    include_special: bool = False,
):
    """生成已执行的测试用例 Excel

    Args:
        path: 输出路径
        case_count: 用例数量
        pass_rate: 通过率 (0-1)
        include_special: 是否包含特殊字符用例
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    headers = [
        "用例编号", "所属模块", "功能点", "测试维度", "用例标题",
        "优先级", "前置条件", "测试步骤", "测试数据", "预期结果",
        "备注", "执行结果",
    ]
    ws.append(headers)

    modules = ["用户管理", "订单系统", "支付系统", "内容管理"]
    dims = ["正向测试", "负向测试", "边界测试", "异常测试", "安全测试"]
    priorities = ["P0", "P1", "P2"]

    results = []
    for i in range(case_count):
        if i < int(case_count * pass_rate):
            results.append(random.choice(["通过", "pass", "✅"]))
        elif i < int(case_count * (pass_rate + 0.1)):
            results.append(random.choice(["失败", "fail", "❌"]))
        elif i < int(case_count * (pass_rate + 0.15)):
            results.append("阻塞")
        else:
            results.append("")

    random.shuffle(results)

    for i in range(1, case_count + 1):
        module = modules[i % len(modules)]
        dim = dims[i % len(dims)]
        pri = priorities[i % len(priorities)]

        title = f"{module}_{dim}_用例_{i}"
        if include_special and i == case_count:
            title = "特殊字符: <script>alert(1)</script> 'OR'1'='1"

        ws.append([
            f"TC-{i:03d}",
            module,
            f"功能{i % 5 + 1}",
            dim,
            title,
            pri,
            f"前置条件{i}",
            f"1.打开{module}\n2.执行操作\n3.检查结果",
            f"测试数据{i}",
            f"预期结果{i}",
            f"备注{i}" if i % 3 == 0 else "",
            results[i - 1],
        ])

    wb.save(path)
    wb.close()


def generate_empty_result_excel(path: str, case_count: int = 20):
    """生成执行结果列为空的 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.append([
        "用例编号", "所属模块", "功能点", "测试维度", "用例标题",
        "优先级", "前置条件", "测试步骤", "测试数据", "预期结果",
        "备注", "执行结果",
    ])
    for i in range(1, case_count + 1):
        ws.append([
            f"TC-{i:03d}", "模块", "功能", "正向测试", f"用例{i}",
            "P1", "前置", "步骤", "数据", "预期", "", "",
        ])
    wb.save(path)
    wb.close()


def generate_no_result_column_excel(path: str, case_count: int = 15):
    """生成没有执行结果列的 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.append([
        "用例编号", "模块", "功能点", "测试维度", "用例标题", "优先级",
    ])
    for i in range(1, case_count + 1):
        ws.append([f"TC-{i:03d}", "模块", "功能", "正向", f"标题{i}", "P1"])
    wb.save(path)
    wb.close()


def generate_english_header_excel(path: str, case_count: int = 15):
    """生成英文表头的 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.append([
        "id", "module", "feature", "type", "title",
        "priority", "precondition", "steps", "data", "expected",
        "remark", "result",
    ])
    for i in range(1, case_count + 1):
        res = "pass" if i % 3 != 0 else "fail"
        ws.append([
            f"TC-{i:03d}", "Module", "Feature", "positive", f"Test Case {i}",
            "P1", "precondition", "1. step", "data", "expected", "", res,
        ])
    wb.save(path)
    wb.close()


def generate_large_testpoints(path: str, count: int = 500):
    """生成大规模测试点文件（用于性能测试）"""
    lines = ["# 大规模测试点\n"]
    for i in range(1, count + 1):
        lines.append(f"\n## 模块{i}：模块{i}\n")
        lines.append(f"### 功能点 1：功能{i}\n")
        lines.append("#### 测试维度：正向测试\n")
        lines.append(f"- 测试点 {i}：测试{i}")
        lines.append(f"  - 测试数据：数据{i}")
        lines.append(f"  - 预期结果：结果{i}\n")
    Path(path).write_text("\n".join(lines), encoding="utf-8")


def generate_malformed_markdown(path: str):
    """生成格式错误的 Markdown"""
    Path(path).write_text(
        "# 标题\n\n"
        "这段没有正确的模块结构\n\n"
        "#### 测试维度：正向测试\n"
        "- 这不是正确的测试点格式\n\n"
        "一些随机的文本内容",
        encoding="utf-8",
    )


def generate_random_string(length: int) -> str:
    """生成指定长度的随机字符串"""
    chars = string.ascii_letters + string.digits + "测试用例数据"
    return "".join(random.choice(chars) for _ in range(length))
