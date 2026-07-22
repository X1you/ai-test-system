#!/usr/bin/env python3
"""
Step5Review.run() 路径单元测试

现有 test_step5_review_extra.py 已覆盖 _read_testcases_as_text I/O 和 run() 错误路径。
本文件覆盖 run() 的成功 + 自检 + 重跑主路径，以及 _read_testcases_as_text 的异常分支：
  - run() 成功（self_check 关闭 → 短路满分，不重跑）
  - run() 成功（self_check 通过 → 不重跑）
  - run() 自检未过 → 重跑成功
  - run() 自检未过 → 重跑 LLMError → 用原始输出
  - run() 自检未过 → 重跑其他异常 → 用原始输出（覆盖 89-91 行）
  - run() 提取评分（带/不带评分行）
  - _read_testcases_as_text：空 sheet（ws=None）/ ImportError 分支
"""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from core.llm_client import LLMError
from core.steps.step5_review import Step5Review


def _make_step(tmp_path, llm=None, config=None):
    return Step5Review(str(tmp_path), config=config or {}, llm=llm)


# ============================================================================
# run() — 成功路径
# ============================================================================


class TestRunSuccess:
    """测试用例评审成功流程"""

    def test_run_success_no_self_check(self, tmp_path):
        # config 无 self_check → 短路满分，不重跑
        llm = MagicMock()
        llm.chat_with_retry.return_value = (
            "# 评审报告\n\n| 维度 | 满分 | 得分 |\n| **总计** | **100** | **88** |\n"
        )
        step = _make_step(tmp_path, llm=llm)
        result = step.run(test_cases="| 编号 | 标题 |\n| TC-1 | 登录 |", kb_context="")
        assert result.ok is True
        assert result.data["score"] == 88
        assert step._read_output("test_case_review_report.md") is not None
        llm.chat.assert_not_called()  # 未触发重跑

    def test_run_success_self_check_pass(self, tmp_path):
        # 开启 self_check 且通过 → 不重跑
        llm = MagicMock()
        llm.chat_with_retry.return_value = "# 评审报告\n综合评分：92\n"
        llm.evaluate.return_value = {"score": 95, "passed": True, "issues": [], "suggestions": []}
        cfg = {"pipeline": {"self_check": True}}
        step = _make_step(tmp_path, llm=llm, config=cfg)
        result = step.run(test_cases="用例数据")
        assert result.ok is True
        assert result.data["score"] == 92
        llm.chat.assert_not_called()

    def test_run_no_score_in_report(self, tmp_path):
        # 报告中无评分 → score=0（覆盖 101-102 行 else 分支）
        llm = MagicMock()
        llm.chat_with_retry.return_value = "# 评审报告\n无评分信息\n"
        step = _make_step(tmp_path, llm=llm)
        result = step.run(test_cases="用例")
        assert result.ok is True
        assert result.data["score"] == 0


# ============================================================================
# run() — 自检未过 + 重跑
# ============================================================================


class TestRunSelfCheckRetry:
    """测试自检未过触发重跑的各路径"""

    def test_self_check_fail_retry_success(self, tmp_path):
        # 自检未过 → 重跑成功，用重跑结果（覆盖 76-88 行）
        llm = MagicMock()
        llm.chat_with_retry.return_value = "# 初版评审\n综合评分：60\n"
        llm.chat.return_value = "# 改进评审\n综合评分：85\n"
        llm.evaluate.return_value = {
            "score": 50, "passed": False, "issues": ["不完整"], "suggestions": [],
        }
        cfg = {"pipeline": {"self_check": True}}
        step = _make_step(tmp_path, llm=llm, config=cfg)
        result = step.run(test_cases="用例")
        assert result.ok is True
        llm.chat.assert_called_once()
        assert "改进评审" in step._read_output("test_case_review_report.md")

    def test_self_check_fail_retry_llm_error(self, tmp_path):
        # 重跑抛 LLMError → 用原始输出（覆盖 87-88 行）
        llm = MagicMock()
        llm.chat_with_retry.return_value = "# 评审\n综合评分：55\n"
        llm.chat.side_effect = LLMError("retry fail")
        llm.evaluate.return_value = {
            "score": 40, "passed": False, "issues": ["x"], "suggestions": [],
        }
        cfg = {"pipeline": {"self_check": True}}
        step = _make_step(tmp_path, llm=llm, config=cfg)
        result = step.run(test_cases="用例")
        assert result.ok is True
        assert "评审" in step._read_output("test_case_review_report.md")

    def test_self_check_fail_retry_unexpected_error(self, tmp_path):
        # 重跑抛非 LLMError 异常 → 用原始输出（覆盖 89-91 行）
        llm = MagicMock()
        llm.chat_with_retry.return_value = "# 评审\n综合评分：50\n"
        llm.chat.side_effect = RuntimeError("unexpected")
        llm.evaluate.return_value = {
            "score": 30, "passed": False, "issues": ["y"], "suggestions": [],
        }
        cfg = {"pipeline": {"self_check": True}}
        step = _make_step(tmp_path, llm=llm, config=cfg)
        result = step.run(test_cases="用例")
        assert result.ok is True
        assert "评审" in step._read_output("test_case_review_report.md")

    def test_run_llm_with_retry_raises(self, tmp_path):
        # 主调用 chat_with_retry 抛 LLMError → ok=False（覆盖 56-58 行）
        llm = MagicMock()
        llm.chat_with_retry.side_effect = LLMError("api down")
        step = _make_step(tmp_path, llm=llm)
        result = step.run(test_cases="用例")
        assert result.ok is False
        assert "api down" in result.error


# ============================================================================
# _read_testcases_as_text — 异常分支
# ============================================================================


class TestReadTestcasesEdgeCases:
    """测试 _read_testcases_as_text 的边界分支"""

    def test_no_sheet_returns_empty(self, tmp_path):
        # load_workbook 返回的 wb.active 为 None → 返回空（覆盖 121-123 行）
        xlsx = tmp_path / "testcases.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        wb.save(str(xlsx))
        wb.close()
        step = _make_step(tmp_path)
        # patch openpyxl.load_workbook 返回 active=None 的假 wb
        fake_wb = MagicMock()
        fake_wb.active = None
        with patch("openpyxl.load_workbook", return_value=fake_wb):
            text = step._read_testcases_as_text()
        assert text == ""
        fake_wb.close.assert_called_once()

    def test_import_error_returns_empty(self, tmp_path):
        # openpyxl 导入失败 → 返回空（覆盖 134-136 行）
        xlsx = tmp_path / "testcases.xlsx"
        xlsx.write_bytes(b"fake")
        step = _make_step(tmp_path)
        with patch.dict(sys.modules, {"openpyxl": None}):
            text = step._read_testcases_as_text()
        assert text == ""
