#!/usr/bin/env python3
"""测试 core/kb/kb_manager_mcp.py — KnowledgeBaseManager 管理器。

覆盖目标：
  - __init__ / search / add / status / add_single 委托给 MCPClient
  - ingest：Excel 标准12列 / 通用3列 / Markdown / 文件不存在 / 异常路径
  - export：按分类分组、Markdown 生成、标签/严重级别字段
  - main() CLI 入口（search / add / ingest / export / status / tags）
  - 文件不存在、不支持格式、openpyxl 缺失等边界

设计：所有 MCPClient 交互用 MagicMock 替换，不触碰真实文件系统。
"""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch, mock_open

import pytest

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


# ═══════════════════════════════════════════════════════════════
# fixtures
# ═══════════════════════════════════════════════════════════════


@pytest.fixture
def mock_mcp_client():
    """构造一个 mock MCPClient，供 KnowledgeBaseManager 注入。"""
    client = MagicMock()
    client.search.return_value = []
    client.create_file.return_value = True
    client.status.return_value = {"source": "mcp-obsidian", "total": 0}
    client.list_files.return_value = []
    client.read_file.return_value = None
    return client


@pytest.fixture
def kb_manager(mock_mcp_client):
    """构造一个 KnowledgeBaseManager，内部 MCPClient 被 mock 替换。"""
    from core.kb import kb_manager_mcp

    with patch.object(kb_manager_mcp, "MCPClient", return_value=mock_mcp_client):
        mgr = kb_manager_mcp.KnowledgeBaseManager(vault_path="/tmp/fake_vault")
    return mgr


# ═══════════════════════════════════════════════════════════════
# 委托方法
# ═══════════════════════════════════════════════════════════════


class TestDelegateMethods:
    """测试 KnowledgeBaseManager 委托给 mcp_client 的方法。"""

    def test_init_creates_mcp_client(self, mock_mcp_client):
        # 测试 __init__ 正确创建内部 MCPClient
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "MCPClient", return_value=mock_mcp_client) as mc:
            mgr = kb_manager_mcp.KnowledgeBaseManager(
                vault_path="/tmp/v", obsidian_api_base="http://x", obsidian_api_key="k"
            )
        mc.assert_called_once()
        assert mgr.mcp_client is mock_mcp_client

    def test_search_delegates(self, kb_manager, mock_mcp_client):
        # 测试 search 委托给 mcp_client.search
        mock_mcp_client.search.return_value = [{"title": "x"}]
        result = kb_manager.search("kw", category="pitfalls", limit=5)
        mock_mcp_client.search.assert_called_once_with("kw", category="pitfalls", limit=5)
        assert result == [{"title": "x"}]

    def test_add_delegates(self, kb_manager, mock_mcp_client):
        # 测试 add 委托给 mcp_client.create_file
        from core.kb.mcp_client import KnowledgeItem

        item = KnowledgeItem(id="", title="t", content="c", category="pitfalls", module="m", tags=[])
        mock_mcp_client.create_file.return_value = True
        assert kb_manager.add(item) is True
        mock_mcp_client.create_file.assert_called_once_with(item)

    def test_status_delegates(self, kb_manager, mock_mcp_client):
        # 测试 status 委托给 mcp_client.status
        mock_mcp_client.status.return_value = {"total": 3}
        result = kb_manager.status()
        assert result == {"total": 3}

    def test_add_single_builds_item_and_adds(self, kb_manager, mock_mcp_client):
        # 测试 add_single 组装 KnowledgeItem 后调用 add
        mock_mcp_client.create_file.return_value = True
        result = kb_manager.add_single(
            "标题", "内容", "pitfalls", module="mod", tags=["a"], severity="high"
        )
        assert result is True
        called_item = mock_mcp_client.create_file.call_args[0][0]
        assert called_item.title == "标题"
        assert called_item.severity == "high"


# ═══════════════════════════════════════════════════════════════
# ingest — Excel 标准12列格式
# ═══════════════════════════════════════════════════════════════


class TestIngestExcelStandard:
    """测试 ingest 处理 testcases.xlsx 标准12列格式。"""

    def test_ingest_excel_testcase_format(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试标准12列 Excel 格式（用例编号|模块|功能点|...）回灌成功
        import openpyxl

        f = tmp_path / "testcases.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["用例编号", "模块", "功能点", "维度", "用例标题", "优先级",
                   "前置条件", "步骤", "测试数据", "预期结果", "备注", "结果"])
        ws.append(["TC-001", "登录", "登录功能", "功能", "正常登录", "P0",
                   "用户存在", "1.输入账号", "acc", "登录成功", "", "通过"])
        wb.save(f)

        mock_mcp_client.create_file.return_value = True
        count = kb_manager.ingest(str(f), "historical-cases", project="p", batch="2024-01")
        assert count == 1
        item = mock_mcp_client.create_file.call_args[0][0]
        assert "TC-001" in item.title
        assert "正常登录" in item.title
        # historical-cases 时 module 编码为 project/batch/module
        assert item.module == "p/2024-01/登录"

    def test_ingest_excel_testcase_with_result_column(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试标准格式包含执行结果列（第12列）时内容拼接 **执行结果**
        import openpyxl

        f = tmp_path / "tc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["用例编号", "模块", "功能点", "维度", "用例标题", "优先级",
                   "前置条件", "步骤", "测试数据", "预期结果", "备注", "结果"])
        ws.append(["TC-002", "m", "f", "d", "标题行", "P1",
                   "pre", "step", "data", "exp", "", "失败"])
        wb.save(f)

        kb_manager.ingest(str(f), "business-rules")
        item = mock_mcp_client.create_file.call_args[0][0]
        assert "**执行结果**: 失败" in item.content

    def test_ingest_excel_testcase_no_title_skipped(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试标准格式中标题为空的行被跳过
        import openpyxl

        f = tmp_path / "tc.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["用例编号", "模块", "功能点", "维度", "用例标题", "优先级",
                   "前置条件", "步骤", "测试数据", "预期结果", "备注", "结果"])
        ws.append(["TC-003", "m", "f", "d", "", "P1", "p", "s", "d", "e", "", ""])
        wb.save(f)

        count = kb_manager.ingest(str(f), "business-rules")
        assert count == 0
        mock_mcp_client.create_file.assert_not_called()

    def test_ingest_excel_testcase_defaults_project_batch(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试未传 project/batch 时使用文件名和当前日期兜底
        import openpyxl

        f = tmp_path / "mycases.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["用例编号", "模块", "功能点", "维度", "用例标题", "优先级",
                   "前置条件", "步骤", "测试数据", "预期结果", "备注", "结果"])
        ws.append(["TC-004", "modX", "f", "d", "有标题", "P0", "p", "s", "d", "e", "", ""])
        wb.save(f)

        kb_manager.ingest(str(f), "historical-cases")
        item = mock_mcp_client.create_file.call_args[0][0]
        # project 兜底为文件名 mycases
        assert item.module.startswith("mycases/")


# ═══════════════════════════════════════════════════════════════
# ingest — Excel 通用3列格式
# ═══════════════════════════════════════════════════════════════


class TestIngestExcelGeneric:
    """测试 ingest 处理通用3列格式（标题|内容|标签）。"""

    def test_ingest_excel_generic_3col(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试通用3列格式（标题|内容|标签）回灌成功
        import openpyxl

        f = tmp_path / "g.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["标题", "内容", "标签"])
        ws.append(["规则A", "内容A", "tag1,tag2"])
        wb.save(f)

        mock_mcp_client.create_file.return_value = True
        count = kb_manager.ingest(str(f), "business-rules", module="mod")
        assert count == 1
        item = mock_mcp_client.create_file.call_args[0][0]
        assert item.title == "规则A"
        assert item.tags == ["tag1", "tag2"]

    def test_ingest_excel_generic_empty_title_skipped(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试通用3列格式标题或内容为空时跳过
        import openpyxl

        f = tmp_path / "g.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["标题", "内容", "标签"])
        ws.append(["", "内容A", "t"])
        ws.append(["标题B", "", "t"])
        wb.save(f)

        count = kb_manager.ingest(str(f), "business-rules")
        assert count == 0

    def test_ingest_excel_row_too_short_skipped(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试行数据不足3列时被跳过
        import openpyxl

        f = tmp_path / "g.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["标题", "内容", "标签"])
        ws.append(["only-one-col"])
        wb.save(f)

        count = kb_manager.ingest(str(f), "business-rules")
        assert count == 0

    def test_ingest_excel_none_row_skipped(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试数据行不足3列时被跳过（覆盖 line 107: len(row) < 3 分支）
        # openpyxl 会把行填充到表头列数，所以用2列表头使数据行 len=2 < 3
        import openpyxl

        f = tmp_path / "g.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["标题", "内容"])  # 只有2列表头
        ws.append(["val1", "val2"])  # 数据行 len=2 < 3 → 被跳过
        wb.save(f)

        count = kb_manager.ingest(str(f), "business-rules")
        assert count == 0


# ═══════════════════════════════════════════════════════════════
# ingest — Markdown / 不存在 / 异常
# ═══════════════════════════════════════════════════════════════


class TestIngestMarkdownAndErrors:
    """测试 Markdown 回灌及错误路径。"""

    def test_ingest_markdown(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试 Markdown 文件回灌成功
        f = tmp_path / "note.md"
        f.write_text("# 标题\n内容正文", encoding="utf-8")

        mock_mcp_client.create_file.return_value = True
        count = kb_manager.ingest(str(f), "business-rules", module="mod")
        assert count == 1
        item = mock_mcp_client.create_file.call_args[0][0]
        assert item.title == "note"
        assert "内容正文" in item.content

    def test_ingest_file_not_exists(self, kb_manager):
        # 测试源文件不存在时返回 0
        assert kb_manager.ingest("/nonexistent/file.xlsx", "business-rules") == 0

    def test_ingest_unsupported_format(self, kb_manager, tmp_path):
        # 测试不支持的文件格式返回 0
        f = tmp_path / "data.csv"
        f.write_text("a,b,c", encoding="utf-8")
        assert kb_manager.ingest(str(f), "business-rules") == 0

    def test_ingest_excel_read_exception(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试 openpyxl 读取异常时捕获并返回 0
        f = tmp_path / "bad.xlsx"
        f.write_text("not-an-excel", encoding="utf-8")
        count = kb_manager.ingest(str(f), "business-rules")
        assert count == 0


# ═══════════════════════════════════════════════════════════════
# export
# ═══════════════════════════════════════════════════════════════


class TestExport:
    """测试 export 导出增强上下文。"""

    def test_export_basic(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试 export 生成 Markdown 并写入文件
        mock_mcp_client.search.return_value = [
            {
                "category": "business-rules",
                "filepath": "path/to/file.md",
                "module": "mod",
                "tags": ["t1"],
                "severity": None,
                "content": "短内容",
            }
        ]
        out = tmp_path / "out" / "ctx.md"
        result = kb_manager.export("kw", output_file=str(out))
        assert result == str(out)
        text = out.read_text(encoding="utf-8")
        assert "知识库增强上下文" in text
        assert "业务规则" in text
        assert "[[path/to/file.md]]" in text

    def test_export_with_severity_and_long_content(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试 severity 字段和超长内容截断（>500字）
        long_content = "x" * 600
        mock_mcp_client.search.return_value = [
            {
                "category": "pitfalls",
                "filepath": "p.md",
                "module": "m",
                "tags": [],
                "severity": "high",
                "content": long_content,
            }
        ]
        out = tmp_path / "ctx2.md"
        result = kb_manager.export("kw", output_file=str(out))
        text = out.read_text(encoding="utf-8")
        assert "严重级别" in text
        assert "..." in text

    def test_export_empty_results(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试无搜索结果时仍生成基本框架
        mock_mcp_client.search.return_value = []
        out = tmp_path / "empty.md"
        result = kb_manager.export("kw", output_file=str(out))
        text = out.read_text(encoding="utf-8")
        assert "命中 0 条" in text

    def test_export_content_with_frontmatter_split(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试内容含 --- 分隔符时截取 frontmatter 之后部分
        mock_mcp_client.search.return_value = [
            {
                "category": "business-rules",
                "filepath": "f.md",
                "module": "",
                "tags": [],
                "severity": None,
                "content": "---\nid: 1\n---\n真实内容部分",
            }
        ]
        out = tmp_path / "fm.md"
        kb_manager.export("kw", output_file=str(out))
        text = out.read_text(encoding="utf-8")
        assert "真实内容部分" in text


# ═══════════════════════════════════════════════════════════════
# CLI main()
# ═══════════════════════════════════════════════════════════════


class TestMainCLI:
    """测试 main() CLI 入口的各子命令。"""

    def test_main_search(self, capsys):
        # 测试 CLI search 子命令输出 JSON
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.search.return_value = [{"title": "r"}]
            with patch.object(sys, "argv", ["kb", "search", "keyword"]):
                kb_manager_mcp.main()
            inst.search.assert_called_once()
            captured = capsys.readouterr()
            assert '"title"' in captured.out

    def test_main_add(self):
        # 测试 CLI add 子命令调用 add_single
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            with patch.object(sys, "argv", [
                "kb", "add", "--title", "T", "--content", "C",
                "--category", "pitfalls", "--module", "m", "--tags", "a,b"
            ]):
                kb_manager_mcp.main()
            inst.add_single.assert_called_once()
            args = inst.add_single.call_args
            assert args[0][0] == "T"

    def test_main_add_with_severity(self):
        # 测试 CLI add 带 severity 参数
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            with patch.object(sys, "argv", [
                "kb", "add", "--title", "T", "--content", "C",
                "--category", "pitfalls", "--severity", "high"
            ]):
                kb_manager_mcp.main()
            inst.add_single.assert_called_once()

    def test_main_ingest(self):
        # 测试 CLI ingest 子命令
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.ingest.return_value = 2
            with patch.object(sys, "argv", [
                "kb", "ingest", "file.xlsx", "--category", "business-rules"
            ]):
                kb_manager_mcp.main()
            inst.ingest.assert_called_once()

    def test_main_export(self):
        # 测试 CLI export 子命令
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.export.return_value = "/tmp/out.md"
            with patch.object(sys, "argv", ["kb", "export", "kw", "--output", "o.md"]):
                kb_manager_mcp.main()
            inst.export.assert_called_once_with("kw", "o.md")

    def test_main_status(self, capsys):
        # 测试 CLI status 子命令输出 JSON
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.status.return_value = {"total": 5}
            with patch.object(sys, "argv", ["kb", "status"]):
                kb_manager_mcp.main()
            captured = capsys.readouterr()
            assert "5" in captured.out

    def test_main_tags_with_data(self, capsys):
        # 测试 CLI tags 子命令收集并输出标签统计
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.mcp_client.list_files.return_value = ["f1.md", "f2.md"]
            inst.mcp_client.read_file.return_value = {"content": "---\ntags: [alpha, beta]\n---\nbody"}
            inst.mcp_client._parse_yaml_frontmatter.return_value = {"tags": ["alpha", "beta"]}
            with patch.object(sys, "argv", ["kb", "tags"]):
                kb_manager_mcp.main()
            captured = capsys.readouterr()
            assert "alpha" in captured.out

    def test_main_tags_empty(self, capsys):
        # 测试 CLI tags 知识库无标签时输出提示
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.mcp_client.list_files.return_value = []
            with patch.object(sys, "argv", ["kb", "tags"]):
                kb_manager_mcp.main()
            captured = capsys.readouterr()
            assert "暂无标签" in captured.out

    def test_main_tags_string_tags(self, capsys):
        # 测试 CLI tags 中 frontmatter tags 为字符串（逗号分隔）时的拆分
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.mcp_client.list_files.return_value = ["f1.md"]
            inst.mcp_client.read_file.return_value = {"content": "x"}
            inst.mcp_client._parse_yaml_frontmatter.return_value = {"tags": "x, y, z"}
            with patch.object(sys, "argv", ["kb", "tags"]):
                kb_manager_mcp.main()
            captured = capsys.readouterr()
            assert "x" in captured.out

    def test_main_tags_read_file_none_skipped(self, capsys):
        # 测试 CLI tags 中 read_file 返回 None 时跳过该文件（覆盖 line 360）
        from core.kb import kb_manager_mcp

        with patch.object(kb_manager_mcp, "KnowledgeBaseManager") as MockMgr:
            inst = MockMgr.return_value
            inst.mcp_client.list_files.return_value = ["f1.md", "f2.md"]
            # f1.md 返回 None（跳过），f2.md 有数据
            inst.mcp_client.read_file.side_effect = [None, {"content": "x"}]
            inst.mcp_client._parse_yaml_frontmatter.return_value = {"tags": ["ok"]}
            with patch.object(sys, "argv", ["kb", "tags"]):
                kb_manager_mcp.main()
            captured = capsys.readouterr()
            assert "ok" in captured.out


# ═══════════════════════════════════════════════════════════════
# ingest — openpyxl 缺失
# ═══════════════════════════════════════════════════════════════


class TestIngestOpenpyxlMissing:
    """测试 openpyxl 不可用时的降级。"""

    def test_ingest_openpyxl_import_error(self, kb_manager, mock_mcp_client, tmp_path):
        # 测试 openpyxl import 失败时 ingest 返回 0（不崩溃）
        f = tmp_path / "tc.xlsx"
        f.write_bytes(b"fake")

        import builtins

        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("no openpyxl")
            return real_import(name, *args, **kwargs)

        with patch.object(builtins, "__import__", side_effect=fake_import):
            count = kb_manager.ingest(str(f), "business-rules")
        assert count == 0
