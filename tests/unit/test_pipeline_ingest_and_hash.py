#!/usr/bin/env python3
"""测试 Pipeline 核心功能：产物 hash 校验 + 知识库回灌。

覆盖本次新增（commit 5baf95d）的逻辑：
  - TC-006/018 修复：_check_skip 的 SHA256 hash 篡改检测
  - P0 回灌：_ingest_to_kb / _verify_ingest / _extract_pitfalls_from_report
  - mark_done 自动记录 output_hashes

设计原则：不真实调用 LLM/子进程/Vault，全部用 tmp_path + mock 隔离。
"""

import hashlib
import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


@pytest.fixture
def pipeline_factory(tmp_path):
    """构造一个 Pipeline 实例（用 tmp_path 隔离，不触真实 LLM/Vault）。"""
    from core.pipeline import Pipeline

    config = {
        "llm": {"provider": "deepseek", "api_key": "sk-test", "model": "m"},
        "knowledge_base": {"enabled": True, "vault_path": str(tmp_path / "vault")},
    }

    def _make():
        return Pipeline(config, str(tmp_path / "output"))

    return _make


# ═══════════════════════════════════════════════════════════════
# TC-006/018：产物 hash 校验（篡改检测）
# ═══════════════════════════════════════════════════════════════


class TestProductHashCheck:
    """测试 _check_skip 的篡改检测逻辑。"""

    def test_skip_when_not_tampered(self, pipeline_factory):
        """场景A：文件未篡改 → _check_skip 返回 True（正常跳过）。"""
        pipe = pipeline_factory()
        f = pipe.output_dir / "requirements_analysis.md"
        f.write_text("原始内容", encoding="utf-8")
        original_hash = hashlib.sha256(f.read_bytes()).hexdigest()
        state = {
            "completed_steps": [1],
            "output_hashes": {"1": original_hash},
        }
        assert pipe._check_skip(state, 1, "requirements_analysis.md") is True

    def test_force_rerun_when_tampered(self, pipeline_factory):
        """场景B：文件被篡改 → _check_skip 返回 False + 移除完成标记。"""
        pipe = pipeline_factory()
        f = pipe.output_dir / "requirements_analysis.md"
        f.write_text("原始内容", encoding="utf-8")
        # 记录原始 hash
        original_hash = hashlib.sha256(f.read_bytes()).hexdigest()
        state = {
            "completed_steps": [1],
            "output_hashes": {"1": original_hash},
        }
        pipe.save_state(state)

        # 篡改文件
        f.write_text("被篡改的内容", encoding="utf-8")

        # 重新加载 state（模拟下一次 run 的 load_state）
        state = pipe.load_state()
        result = pipe._check_skip(state, 1, "requirements_analysis.md")

        # 应返回 False（不跳过，强制重跑）
        assert result is False
        # state 中应已移除该步骤的完成标记
        state_after = pipe.load_state()
        assert 1 not in state_after["completed_steps"]
        assert "1" not in state_after.get("output_hashes", {})

    def test_skip_when_no_hash_recorded(self, pipeline_factory):
        """场景C：旧 state 无 output_hashes（向后兼容）→ 退化为只检查文件存在。"""
        pipe = pipeline_factory()
        f = pipe.output_dir / "requirements_analysis.md"
        f.write_text("内容", encoding="utf-8")
        state = {"completed_steps": [1]}  # 无 output_hashes 键
        assert pipe._check_skip(state, 1, "requirements_analysis.md") is True

    def test_not_skip_when_file_missing(self, pipeline_factory):
        """场景D：文件不存在 → 返回 False（产物丢失需重跑）。"""
        pipe = pipeline_factory()
        state = {"completed_steps": [1], "output_hashes": {}}
        assert pipe._check_skip(state, 1, "nonexistent.md") is False

    def test_not_skip_when_step_not_done(self, pipeline_factory):
        """场景E：步骤未在 completed_steps → 返回 False。"""
        pipe = pipeline_factory()
        f = pipe.output_dir / "requirements_analysis.md"
        f.write_text("内容", encoding="utf-8")
        state = {"completed_steps": [], "output_hashes": {}}
        assert pipe._check_skip(state, 1, "requirements_analysis.md") is False

    def test_xlsx_with_results_not_treated_as_tampered(self, pipeline_factory):
        """场景F：testcases.xlsx 被填了执行结果（hash 变化）不视为篡改。

        Step4/Step6 共用 testcases.xlsx，用户填写执行结果后 hash 自然改变。
        只要检测到「执行结果」列有值，就应豁免篡改检测，避免丢失用户数据。
        豁免覆盖任意以 testcases.xlsx 为产物的步骤（防御性编程）。
        """
        from openpyxl import Workbook

        pipe = pipeline_factory()
        xlsx = pipe.output_dir / "testcases.xlsx"

        # 构造"原始"Excel（Step4 生成，无执行结果）
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["用例编号", "标题", "执行结果"])
        ws.append(["TC-1", "登录", ""])  # 空执行结果
        wb.save(str(xlsx))
        wb.close()

        original_hash = hashlib.sha256(xlsx.read_bytes()).hexdigest()
        state = {"completed_steps": [4], "output_hashes": {"4": original_hash}}

        # 用户填写执行结果（Excel hash 变化）
        wb2 = Workbook()
        ws2 = wb2.active
        assert ws2 is not None
        ws2.append(["用例编号", "标题", "执行结果"])
        ws2.append(["TC-1", "登录", "通过"])  # 填了执行结果
        wb2.save(str(xlsx))
        wb2.close()

        # hash 已变化，但因有执行结果 → 豁免篡改检测，返回 True（跳过）
        result = pipe._check_skip(state, 4, "testcases.xlsx")
        assert result is True, "填了执行结果的 testcases.xlsx 不应被视为篡改"

    def test_xlsx_without_results_treated_as_tampered(self, pipeline_factory):
        """场景G：testcases.xlsx hash 变化但无执行结果 → 视为篡改。

        对照测试：如果 Excel 被修改但没有执行结果，仍应触发篡改检测。
        """
        from openpyxl import Workbook

        pipe = pipeline_factory()
        xlsx = pipe.output_dir / "testcases.xlsx"

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["用例编号", "标题", "执行结果"])
        ws.append(["TC-1", "登录", ""])
        wb.save(str(xlsx))
        wb.close()

        original_hash = hashlib.sha256(xlsx.read_bytes()).hexdigest()
        state = {"completed_steps": [4], "output_hashes": {"4": original_hash}}
        pipe.save_state(state)

        # 篡改但保持无执行结果
        wb2 = Workbook()
        ws2 = wb2.active
        assert ws2 is not None
        ws2.append(["用例编号", "标题", "执行结果"])
        ws2.append(["TC-1", "被篡改的标题", ""])  # 改了内容，仍无执行结果
        wb2.save(str(xlsx))
        wb2.close()

        result = pipe._check_skip(state, 4, "testcases.xlsx")
        assert result is False, "无执行结果的 hash 变化应视为篡改"


class TestMarkDoneRecordsHash:
    """测试 mark_done 自动记录 output_hashes。"""

    def test_mark_done_records_hash(self, pipeline_factory):
        """mark_done 后 state 应包含对应步骤的 output_hashes。"""
        from core.steps.base import StepResult

        pipe = pipeline_factory()
        # 预置输出文件（模拟 Step1 产生的 requirements_analysis.md）
        f = pipe.output_dir / "requirements_analysis.md"
        f.write_text("Step1 产物内容", encoding="utf-8")
        expected_hash = hashlib.sha256(f.read_bytes()).hexdigest()

        state = pipe.load_state()
        result = StepResult(ok=True, data={"count": 5})
        pipe.mark_done(state, 1, result)

        state_after = pipe.load_state()
        assert "output_hashes" in state_after
        assert state_after["output_hashes"]["1"] == expected_hash
        assert 1 in state_after["completed_steps"]

    def test_compute_file_hash_stable(self, pipeline_factory):
        """相同内容 hash 一致，不同内容 hash 不同。"""
        pipe = pipeline_factory()
        f1 = pipe.output_dir / "a.txt"
        f1.write_text("hello", encoding="utf-8")
        f2 = pipe.output_dir / "b.txt"
        f2.write_text("hello", encoding="utf-8")
        f3 = pipe.output_dir / "c.txt"
        f3.write_text("world", encoding="utf-8")

        h1 = pipe._compute_file_hash(f1)
        h2 = pipe._compute_file_hash(f2)
        h3 = pipe._compute_file_hash(f3)

        assert h1 == h2  # 相同内容
        assert h1 != h3  # 不同内容
        assert len(h1) == 64  # SHA256 hex 长度

    def test_compute_file_hash_missing(self, pipeline_factory):
        """文件不存在时返回 None（不抛异常）。"""
        pipe = pipeline_factory()
        assert pipe._compute_file_hash(pipe.output_dir / "nope.md") is None

    def test_step_output_file_mapping(self, pipeline_factory):
        """_step_output_file 正确映射 STEP_REGISTRY。"""
        pipe = pipeline_factory()
        assert pipe._step_output_file(1) == "requirements_analysis.md"
        assert pipe._step_output_file(3) == "testpoints.md"
        assert pipe._step_output_file(7) == "test_report.md"
        assert pipe._step_output_file(999) == ""  # 不存在的步骤


# ═══════════════════════════════════════════════════════════════
# P0 回灌：_ingest_to_kb 容错逻辑
# ═══════════════════════════════════════════════════════════════


class TestIngestToKBFaultTolerance:
    """测试 _ingest_to_kb 的容错（失败不阻塞主流程）。"""

    def test_skip_when_kb_disabled(self, pipeline_factory, tmp_path):
        """知识库未启用 → 返回 0，不报错。"""
        from core.pipeline import Pipeline

        config = {"llm": {}, "knowledge_base": {"enabled": False}}
        pipe = Pipeline(config, str(tmp_path / "out"))
        src = tmp_path / "src.md"
        src.write_text("x", encoding="utf-8")

        count = pipe._ingest_to_kb(str(src), "historical-cases")
        assert count == 0

    def test_skip_when_no_vault_path(self, pipeline_factory, tmp_path):
        """vault_path 未配置 → 返回 0。"""
        from core.pipeline import Pipeline

        config = {
            "llm": {},
            "knowledge_base": {"enabled": True},  # 无 vault_path
        }
        pipe = Pipeline(config, str(tmp_path / "out"))
        src = tmp_path / "src.md"
        src.write_text("x", encoding="utf-8")

        count = pipe._ingest_to_kb(str(src), "historical-cases")
        assert count == 0

    def test_skip_when_source_missing(self, pipeline_factory, tmp_path):
        """源文件不存在 → 返回 0。"""
        from core.pipeline import Pipeline

        config = {
            "llm": {},
            "knowledge_base": {"enabled": True, "vault_path": str(tmp_path / "v")},
        }
        pipe = Pipeline(config, str(tmp_path / "out"))
        count = pipe._ingest_to_kb(str(tmp_path / "nonexistent.md"), "historical-cases")
        assert count == 0

    def test_skip_when_kb_script_missing(self, pipeline_factory, tmp_path, monkeypatch):
        """kb_manager_mcp.py 脚本不存在 → 返回 0（容错）。"""
        from core.pipeline import Pipeline

        config = {
            "llm": {},
            "knowledge_base": {"enabled": True, "vault_path": str(tmp_path / "v")},
        }
        pipe = Pipeline(config, str(tmp_path / "out"))
        src = tmp_path / "src.md"
        src.write_text("x", encoding="utf-8")

        # 模拟 PROJECT_ROOT 下找不到 kb_manager_mcp.py
        monkeypatch.setattr(
            "core.pipeline.PROJECT_ROOT", tmp_path / "fake_root"
        )

        count = pipe._ingest_to_kb(str(src), "historical-cases")
        assert count == 0

    def test_subprocess_failure_returns_zero(self, pipeline_factory, tmp_path):
        """子进程执行失败（非0退出）→ 返回 0，不抛异常。"""
        from core.pipeline import Pipeline

        config = {
            "llm": {},
            "knowledge_base": {"enabled": True, "vault_path": str(tmp_path / "v")},
        }
        pipe = Pipeline(config, str(tmp_path / "out"))
        src = tmp_path / "src.md"
        src.write_text("x", encoding="utf-8")

        # mock subprocess.run 返回失败
        fake_result = MagicMock()
        fake_result.returncode = 1
        fake_result.stderr = "模拟失败"
        with patch("core.pipeline.subprocess.run", return_value=fake_result):
            count = pipe._ingest_to_kb(str(src), "historical-cases")
        assert count == 0


# ═══════════════════════════════════════════════════════════════
# P0 回灌：_verify_ingest 验证逻辑
# ═══════════════════════════════════════════════════════════════


class TestVerifyIngest:
    """测试 _verify_ingest 的 Vault 文件计数。"""

    def test_count_historical_cases(self, tmp_path):
        """historical-cases 按 project/batch 归档，统计 .md 数。"""
        from core.pipeline import Pipeline

        vault = tmp_path / "vault"
        archive = vault / "🏆 历史用例" / "myproj" / "2026-07-18"
        archive.mkdir(parents=True)
        (archive / "TC-001.md").write_text("a", encoding="utf-8")
        (archive / "TC-002.md").write_text("b", encoding="utf-8")
        (archive / "TC-003.md").write_text("c", encoding="utf-8")

        count = Pipeline._verify_ingest(vault, "historical-cases", "myproj", "2026-07-18")
        assert count == 3

    def test_count_pitfalls_flat(self, tmp_path):
        """pitfalls 平铺分类，统计目录下 .md 数。"""
        from core.pipeline import Pipeline

        vault = tmp_path / "vault"
        pitfalls = vault / "⚠️ 线上坑点"
        pitfalls.mkdir(parents=True)
        (pitfalls / "坑点1.md").write_text("a", encoding="utf-8")
        (pitfalls / "坑点2.md").write_text("b", encoding="utf-8")

        count = Pipeline._verify_ingest(vault, "pitfalls", "", "")
        assert count == 2

    def test_zero_when_dir_missing(self, tmp_path):
        """归档目录不存在 → 返回 0。"""
        from core.pipeline import Pipeline

        vault = tmp_path / "vault"
        vault.mkdir()
        count = Pipeline._verify_ingest(vault, "historical-cases", "x", "y")
        assert count == 0


# ═══════════════════════════════════════════════════════════════
# P0 回灌：_extract_pitfalls_from_report 解析逻辑
# ═══════════════════════════════════════════════════════════════


class TestExtractPitfalls:
    """测试 _extract_pitfalls_from_report 的报告解析。"""

    def test_extract_multiple_failures(self, tmp_path):
        """正常场景：解析多个失败用例，返回 list[dict]。"""
        from core.pipeline import Pipeline

        report = tmp_path / "test_report.md"
        report.write_text(
            """# 测试报告

## 📊 总体概览

| 通过 | 失败 |
|------|------|
| 38 | 2 |

## 🔍 失败用例分析

共 **2** 个失败用例。

### 1. TC-006: 步骤产物被篡改

- **所属模块：** Pipeline
- **优先级：** P0
- **失败原因（推断）：** 数据校验失败
- **修复建议：** 补充 hash 校验
- **备注：** 代码中未发现产物完整性校验逻辑

### 2. TC-010: 动态修改模式

- **所属模块：** Pipeline
- **优先级：** P1
- **失败原因（推断）：** 接口异常
- **备注：** 未发现运行中动态修改模式的接口

## ⛔ 阻塞用例分析
""",
            encoding="utf-8",
        )

        pitfalls = Pipeline._extract_pitfalls_from_report(report)

        assert len(pitfalls) == 2
        assert pitfalls[0]["tc_id"] == "TC-006"
        assert "篡改" in pitfalls[0]["title"]
        assert "Pipeline" in pitfalls[0]["content"]
        assert "数据校验失败" in pitfalls[0]["content"]
        assert pitfalls[1]["tc_id"] == "TC-010"
        # 格式瑕疵修复：不应出现 ** ** 残留
        assert "** **" not in pitfalls[0]["content"]

    def test_empty_when_no_failures(self, tmp_path):
        """报告无失败用例（"共 **0** 个"）→ 返回空列表。"""
        from core.pipeline import Pipeline

        report = tmp_path / "test_report.md"
        report.write_text(
            "## 🔍 失败用例分析\n\n共 **0** 个失败用例。\n",
            encoding="utf-8",
        )
        assert Pipeline._extract_pitfalls_from_report(report) == []

    def test_empty_when_report_missing(self, tmp_path):
        """报告文件不存在 → 返回空列表（不抛异常）。"""
        from core.pipeline import Pipeline

        report = tmp_path / "nonexistent.md"
        assert Pipeline._extract_pitfalls_from_report(report) == []

    def test_empty_when_no_failure_section(self, tmp_path):
        """报告无"失败用例分析"章节 → 返回空列表。"""
        from core.pipeline import Pipeline

        report = tmp_path / "test_report.md"
        report.write_text("# 报告\n\n无失败\n", encoding="utf-8")
        assert Pipeline._extract_pitfalls_from_report(report) == []
