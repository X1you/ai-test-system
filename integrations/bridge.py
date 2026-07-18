#!/usr/bin/env python3
"""
Integration Bridge — Core Engine 和 Adapter Layer 之间的唯一连接点

将 Core Engine 的产物（testcases.xlsx）转换为 Canonical Model。
"""

from integrations.models import TestCase, TestResult

# ─── Excel 列索引常量（12 列标准格式）───
# 用例编号 | 模块 | 功能点 | 测试维度 | 用例标题 | 优先级 |
# 前置条件 | 步骤 | 测试数据 | 预期结果 | 备注 | 执行结果
_COL_ID = 0
_COL_MODULE = 1
_COL_FEATURE = 2
_COL_DIMENSION = 3
_COL_TITLE = 4
_COL_PRIORITY = 5
_COL_PRECONDITION = 6
_COL_STEPS = 7
_COL_TEST_DATA = 8
_COL_EXPECTED = 9
_COL_COMMENT = 10
_COL_STATUS = 11
_MIN_COLUMNS = 10  # excel_to_testcases 的最小有效列数
_RESULTS_MIN_COLUMNS = 12  # excel_to_results 需要完整 12 列

_EXCEL_HEADERS = [
    "用例编号", "所属模块", "功能点", "测试维度", "用例标题",
    "优先级", "前置条件", "测试步骤", "测试数据", "预期结果",
    "备注", "执行结果",
]

# 状态归一化映射（中文 → 内部标准）
_STATUS_MAP: dict[str, str] = {
    "通过": "passed",
    "失败": "failed",
    "阻塞": "blocked",
    "跳过": "skipped",
    "未执行": "",
    "": "",
}


def _normalize_status(val: str | None) -> str:
    """标准化执行结果状态

    将中文状态（通过/失败/阻塞/跳过）映射为内部标准（passed/failed/...）。
    未知状态保持原样返回。
    """
    if not val:
        return ""
    return _STATUS_MAP.get(str(val).strip(), str(val).strip())


def _require_openpyxl():
    """延迟导入 openpyxl，缺失时抛出带提示的 ImportError"""
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError:
        raise ImportError("需要 openpyxl 来读写 Excel 文件（pip install openpyxl）")


def _safe_str(val, default: str = "") -> str:
    """安全转字符串：None → default，其他 → str()"""
    return str(val) if val is not None else default


class IntegrationBridge:
    """桥接层 — 将 Core Engine 的产物转换为 Canonical Model"""

    @staticmethod
    def excel_to_testcases(xlsx_path: str) -> list[TestCase]:
        """读取 testcases.xlsx → List[TestCase]

        12 列格式：用例编号 | 模块 | 功能点 | 测试维度 | 用例标题 | 优先级 |
                  前置条件 | 步骤 | 测试数据 | 预期结果 | 备注 | 执行结果
        """
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return []

            cases: list[TestCase] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < _MIN_COLUMNS:
                    continue

                steps_raw = row[_COL_STEPS]
                cases.append(TestCase(
                    id=_safe_str(row[_COL_ID]),
                    title=_safe_str(row[_COL_TITLE]),
                    module=_safe_str(row[_COL_MODULE]),
                    feature=_safe_str(row[_COL_FEATURE]),
                    dimension=_safe_str(row[_COL_DIMENSION]),
                    priority=_safe_str(row[_COL_PRIORITY]),
                    precondition=_safe_str(row[_COL_PRECONDITION]),
                    steps=_safe_str(steps_raw).split("\n") if steps_raw else [],
                    test_data=_safe_str(row[_COL_TEST_DATA]),
                    expected_result=_safe_str(row[_COL_EXPECTED]),
                    status=_normalize_status(
                        str(row[_COL_STATUS]) if len(row) > _COL_STATUS else None
                    ),
                ))

            return cases
        finally:
            wb.close()

    @staticmethod
    def excel_to_results(xlsx_path: str) -> list[TestResult]:
        """读取 testcases.xlsx 的执行结果列 → List[TestResult]

        仅提取有执行结果的行（执行结果列非空）。
        """
        openpyxl = _require_openpyxl()
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        try:
            ws = wb.active
            if ws is None:
                return []

            results: list[TestResult] = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < _RESULTS_MIN_COLUMNS:
                    continue

                tc_id = _safe_str(row[_COL_ID])
                status_raw = _safe_str(row[_COL_STATUS]).strip()
                if not status_raw:
                    continue

                results.append(TestResult(
                    test_case_id=tc_id,
                    status=_normalize_status(status_raw),
                    comment=_safe_str(row[_COL_COMMENT]),
                ))

            return results
        finally:
            wb.close()

    @staticmethod
    def testcases_to_excel(cases: list[TestCase], output_path: str) -> str:
        """反向：Canonical → Excel（用于 pull 后写回）"""
        openpyxl = _require_openpyxl()
        wb = openpyxl.Workbook()
        try:
            ws = wb.active
            assert ws is not None  # Workbook() 总会创建一个 active sheet
            ws.title = "测试用例"
            ws.append(_EXCEL_HEADERS)

            for tc in cases:
                ws.append([
                    tc.id, tc.module, tc.feature, tc.dimension, tc.title,
                    tc.priority, tc.precondition,
                    "\n".join(tc.steps) if tc.steps else "",
                    tc.test_data, tc.expected_result, "", tc.status,
                ])

            wb.save(output_path)
            return output_path
        finally:
            wb.close()
