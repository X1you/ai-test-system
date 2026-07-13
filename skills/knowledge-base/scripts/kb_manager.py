#!/usr/bin/env python3
"""
测试知识库管理脚本

本地优先的测试知识库系统，支持：
  init    — 初始化知识库目录结构
  search  — TF-IDF + BM25 关键词检索
  add     — 添加单条知识
  ingest  — 从文件批量回灌知识（Excel/Markdown）
  export  — 导出增强上下文（Markdown 格式）
  status  — 知识库概况统计
  list    — 列出知识条目

用法:
    python kb_manager.py init   [--kb-dir DIR]
    python kb_manager.py search "关键词" [--kb-dir DIR] [--category CAT] [--limit N]
    python kb_manager.py add    --category CAT --title TITLE --content TEXT [--module M] [--severity S] [--kb-dir DIR]
    python kb_manager.py ingest <file> --category CAT [--module M] [--kb-dir DIR]
    python kb_manager.py export  "关键词" [--output FILE] [--kb-dir DIR]
    python kb_manager.py status  [--kb-dir DIR]
    python kb_manager.py list    [--category CAT] [--kb-dir DIR]

依赖:
    openpyxl (可选，仅回灌 Excel 时需要，在 Hermes venv 中)
"""

import argparse
import json
import math
import re
import sys
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path


# ═══════════════════════════════════════════════════════════════
# 常量
# ═══════════════════════════════════════════════════════════════

CATEGORIES = ["business-rules", "historical-cases", "pitfalls", "templates"]

CATEGORY_LABELS = {
    "business-rules": "业务规则",
    "historical-cases": "历史优质用例",
    "pitfalls": "线上坑点",
    "templates": "用例模板",
}

CATEGORY_ICONS = {
    "business-rules": "📋",
    "historical-cases": "🏆",
    "pitfalls": "⚠️",
    "templates": "📐",
}


# ═══════════════════════════════════════════════════════════════
# 中文分词器（简单但有效）
# ═══════════════════════════════════════════════════════════════

# 停用词
STOP_WORDS = frozenset([
    "的", "了", "在", "是", "和", "与", "或", "对", "为", "从", "到", "由",
    "这", "那", "它", "他", "她", "我", "你", "们", "个", "中", "上", "下",
    "不", "也", "都", "就", "还", "又", "被", "把", "给", "向", "已", "一",
    "the", "a", "an", "is", "are", "was", "were", "in", "on", "at", "to",
    "for", "of", "with", "by", "from", "and", "or", "not", "no", "yes",
])


def tokenize(text: str) -> list:
    """
    混合分词：英文按单词、中文按 2-gram + 单字
    对中文检索效果较好
    """
    if not text:
        return []

    text = text.lower().strip()
    tokens = []

    # 英文单词
    for m in re.finditer(r"[a-z][a-z0-9_\-]{1,}", text):
        word = m.group()
        if word not in STOP_WORDS and len(word) > 1:
            tokens.append(word)

    # 中文字符序列
    for seg in re.finditer(r"[\u4e00-\u9fff]+", text):
        chars = seg.group()
        # 单字
        for ch in chars:
            if ch not in STOP_WORDS:
                tokens.append(ch)
        # 2-gram（bigram）
        for i in range(len(chars) - 1):
            tokens.append(chars[i:i + 2])

    return tokens


# ═══════════════════════════════════════════════════════════════
# BM25 检索引擎
# ═══════════════════════════════════════════════════════════════

class BM25Engine:
    """纯标准库实现的 BM25 检索引擎"""

    def __init__(self, k1=1.5, b=0.75):
        self.k1 = k1
        self.b = b
        self.documents = []       # [{id, category, source, title, content, tokens, token_freq, metadata}]
        self.df = defaultdict(int)  # document frequency per token
        self.avg_dl = 0.0          # average document length
        self._built = False

    def add_document(self, doc_id: str, category: str, source: str,
                     title: str, content: str, metadata: dict = None):
        """添加文档到索引"""
        full_text = f"{title} {content}"
        tokens = tokenize(full_text)
        token_freq = Counter(tokens)

        self.documents.append({
            "id": doc_id,
            "category": category,
            "source": source,
            "title": title,
            "content": content,
            "tokens": tokens,
            "token_freq": token_freq,
            "dl": len(tokens),
            "metadata": metadata or {},
        })

        # 更新 DF
        for token in token_freq:
            self.df[token] += 1

        self._built = False

    def _build(self):
        """构建检索索引（计算 IDF 等）"""
        n = len(self.documents)
        self.idf = {}
        for token, df in self.df.items():
            # BM25 IDF
            self.idf[token] = math.log((n - df + 0.5) / (df + 0.5) + 1)

        self.avg_dl = (
            sum(d["dl"] for d in self.documents) / n if n > 0 else 0
        )
        self._built = True

    def search(self, query: str, category: str = None, limit: int = 20) -> list:
        """BM25 检索"""
        if not self.documents:
            return []
        if not self._built:
            self._build()

        query_tokens = tokenize(query)
        if not query_tokens:
            return []

        results = []
        for doc in self.documents:
            if category and doc["category"] != category:
                continue

            score = 0.0
            dl = doc["dl"]
            for qt in query_tokens:
                if qt not in self.idf:
                    continue
                tf = doc["token_freq"].get(qt, 0)
                if tf == 0:
                    continue

                # BM25 score
                idf = self.idf[qt]
                numerator = tf * (self.k1 + 1)
                denominator = tf + self.k1 * (1 - self.b + self.b * dl / max(self.avg_dl, 1))
                score += idf * numerator / denominator

            if score > 0:
                # 提取匹配片段
                snippet = self._extract_snippet(doc, query_tokens)
                results.append({
                    "id": doc["id"],
                    "category": doc["category"],
                    "source": doc["source"],
                    "title": doc["title"],
                    "score": round(score, 4),
                    "snippet": snippet,
                    "metadata": doc["metadata"],
                })

        results.sort(key=lambda x: x["score"], reverse=True)
        return results[:limit]

    def _extract_snippet(self, doc: dict, query_tokens: list, max_len: int = 200) -> str:
        """提取匹配上下文片段"""
        content = doc["content"]
        if len(content) <= max_len:
            return content

        # 找最佳匹配位置
        best_pos = 0
        best_score = 0
        for qt in query_tokens:
            pos = content.find(qt)
            if pos >= 0:
                # 简化：取第一个匹配位置
                best_pos = pos
                break

        start = max(0, best_pos - max_len // 3)
        end = min(len(content), start + max_len)
        snippet = content[start:end]
        if start > 0:
            snippet = "..." + snippet
        if end < len(content):
            snippet = snippet + "..."
        return snippet


# ═══════════════════════════════════════════════════════════════
# 知识库管理器
# ═══════════════════════════════════════════════════════════════

class KnowledgeBase:
    """知识库管理"""

    def __init__(self, kb_dir: str):
        self.kb_dir = Path(kb_dir)
        self.index_path = self.kb_dir / "index.json"
        self.engine = BM25Engine()

    def _ensure_dir(self):
        """确保知识库目录存在"""
        self.kb_dir.mkdir(parents=True, exist_ok=True)
        for cat in CATEGORIES:
            (self.kb_dir / cat).mkdir(exist_ok=True)

    def init(self):
        """初始化知识库目录结构"""
        self._ensure_dir()
        if not self.index_path.exists():
            self._save_index({"version": "1.0", "created": datetime.now().isoformat(), "entries": []})
        print(f"✅ 知识库已初始化: {self.kb_dir}")
        for cat in CATEGORIES:
            cat_dir = self.kb_dir / cat
            file_count = len(list(cat_dir.glob("*.md")))
            print(f"   {CATEGORY_ICONS[cat]} {cat}/ — {file_count} 条")

    def _load_index(self) -> dict:
        """加载索引"""
        if self.index_path.exists():
            return json.loads(self.index_path.read_text(encoding="utf-8"))
        return {"version": "1.0", "created": datetime.now().isoformat(), "entries": []}

    def _save_index(self, index: dict):
        """保存索引"""
        self._ensure_dir()
        index["updated"] = datetime.now().isoformat()
        self.index_path.write_text(
            json.dumps(index, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )

    def _load_all_documents(self):
        """加载所有知识文档到检索引擎"""
        self.engine = BM25Engine()
        index = self._load_index()

        for entry in index["entries"]:
            self.engine.add_document(
                doc_id=entry["id"],
                category=entry["category"],
                source=entry["source"],
                title=entry["title"],
                content=entry["content"],
                metadata=entry.get("metadata", {}),
            )

    def _parse_markdown(self, file_path: Path) -> list:
        """
        解析 Markdown 知识文件，提取结构化知识条目
        格式：每个 ### 标题下是一段知识内容
        """
        content = file_path.read_text(encoding="utf-8")
        entries = []
        current_title = None
        current_lines = []

        for line in content.split("\n"):
            # ### 标题 = 知识条目标题
            m = re.match(r"^###\s+(.+)", line)
            if m:
                if current_title and current_lines:
                    entries.append({
                        "title": current_title,
                        "content": "\n".join(current_lines).strip(),
                    })
                current_title = m.group(1).strip()
                current_lines = []
            elif current_title:
                current_lines.append(line)

        # 最后一条
        if current_title and current_lines:
            entries.append({
                "title": current_title,
                "content": "\n".join(current_lines).strip(),
            })

        return entries

    def search(self, query: str, category: str = None, limit: int = 20) -> list:
        """检索知识库"""
        self._load_all_documents()
        return self.engine.search(query, category=category, limit=limit)

    def add(self, category: str, title: str, content: str,
            module: str = "", severity: str = ""):
        """添加单条知识到 Markdown 文件"""
        self._ensure_dir()

        # 生成文件名
        safe_title = re.sub(r"[^\w\u4e00-\u9fff]", "_", title)[:30]
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{safe_title}_{timestamp}.md"
        file_path = self.kb_dir / category / filename

        # 构建内容
        lines = [f"# {CATEGORY_LABELS.get(category, category)}\n"]
        lines.append(f"### {title}\n")
        if module:
            lines.append(f"**模块：** {module}\n")
        if severity:
            lines.append(f"**严重程度：** {severity}\n")
        lines.append(f"\n{content}\n")
        lines.append(f"\n*记录时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}*\n")

        file_path.write_text("\n".join(lines), encoding="utf-8")

        # 更新索引
        index = self._load_index()
        entry_id = f"{category}_{timestamp}"
        index["entries"].append({
            "id": entry_id,
            "category": category,
            "source": str(file_path.relative_to(self.kb_dir)),
            "title": title,
            "content": content,
            "metadata": {"module": module, "severity": severity},
            "created": datetime.now().isoformat(),
        })
        self._save_index(index)

        print(f"✅ 已添加到知识库")
        print(f"   分类：{CATEGORY_ICONS.get(category, '')} {category}")
        print(f"   标题：{title}")
        print(f"   文件：{file_path}")

    def ingest_excel(self, file_path: str, category: str, module: str = ""):
        """从 Excel 文件回灌知识"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("❌ 回灌 Excel 需要 openpyxl，请使用 Hermes venv Python", file=sys.stderr)
            return 0

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        # 解析表头
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=1, column=col).value or "").strip()
            if val:
                headers[val] = col

        # 找关键列
        def find_col(*names):
            for name in names:
                for h, c in headers.items():
                    if name in h:
                        return c
            return None

        col_id = find_col("用例编号", "编号", "id")
        col_title = find_col("用例标题", "标题", "名称", "title")
        col_module = find_col("所属模块", "模块", "module")
        col_steps = find_col("测试步骤", "步骤", "steps")
        col_expected = find_col("预期结果", "预期", "expected")
        col_precondition = find_col("前置条件", "precondition")
        col_priority = find_col("优先级", "priority")
        col_data = find_col("测试数据", "data")

        if not col_title:
            print("⚠️  无法识别用例标题列", file=sys.stderr)
            return 0

        # 提取用例
        cases = []
        for row in range(2, ws.max_row + 1):
            def get(col):
                if col and col <= ws.max_column:
                    v = ws.cell(row=row, column=col).value
                    return str(v).strip() if v else ""
                return ""

            title = get(col_title)
            if not title:
                continue

            case_module = get(col_module) or module or "未分类"
            content_parts = []
            if col_precondition:
                pre = get(col_precondition)
                if pre:
                    content_parts.append(f"前置条件：{pre}")
            if col_steps:
                steps = get(col_steps)
                if steps:
                    content_parts.append(f"测试步骤：\n{steps}")
            if col_data:
                data = get(col_data)
                if data:
                    content_parts.append(f"测试数据：{data}")
            if col_expected:
                exp = get(col_expected)
                if exp:
                    content_parts.append(f"预期结果：{exp}")

            cases.append({
                "id": get(col_id),
                "title": title,
                "module": case_module,
                "content": "\n".join(content_parts),
                "priority": get(col_priority),
            })

        wb.close()

        # 按模块分组写入 Markdown 文件
        self._ensure_dir()
        index = self._load_index()
        count = 0

        modules = defaultdict(list)
        for case in cases:
            modules[case["module"]].append(case)

        for mod_name, mod_cases in modules.items():
            safe_mod = re.sub(r"[^\w\u4e00-\u9fff]", "_", mod_name)[:20]
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            filename = f"{safe_mod}_{timestamp}.md"
            file_path = self.kb_dir / category / filename

            lines = [f"# {mod_name} — 优质用例集\n"]
            lines.append(f"*回灌时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}*\n")

            for case in mod_cases:
                lines.append(f"### {case['id']}: {case['title']}\n")
                if case["priority"]:
                    lines.append(f"**优先级：** {case['priority']}\n")
                lines.append(f"{case['content']}\n")

            file_path.write_text("\n".join(lines), encoding="utf-8")

            # 更新索引
            for case in mod_cases:
                entry_id = f"{category}_{mod_name}_{case['id']}_{timestamp}"
                index["entries"].append({
                    "id": entry_id,
                    "category": category,
                    "source": str(file_path.relative_to(self.kb_dir)),
                    "title": f"{case['id']}: {case['title']}",
                    "content": case["content"],
                    "metadata": {
                        "module": case["module"],
                        "priority": case["priority"],
                    },
                    "created": datetime.now().isoformat(),
                })
                count += 1

        self._save_index(index)
        print(f"✅ 已回灌 {count} 条知识到知识库")
        print(f"   分类：{CATEGORY_ICONS.get(category, '')} {category}")
        print(f"   来源：{file_path}")
        print(f"   模块：{', '.join(modules.keys())}")
        return count

    def ingest_markdown(self, file_path: str, category: str):
        """从 Markdown 文件回灌知识"""
        path = Path(file_path)
        content = path.read_text(encoding="utf-8")

        # 尝试按 ### 拆分
        entries = self._parse_markdown(path)

        if not entries:
            # 整体作为一条
            entries = [{"title": path.stem, "content": content}]

        self._ensure_dir()
        index = self._load_index()
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{path.stem}_{timestamp}.md"
        dest_path = self.kb_dir / category / filename

        # 复制文件
        dest_path.write_text(content, encoding="utf-8")

        count = 0
        for entry in entries:
            entry_id = f"{category}_{timestamp}_{count}"
            index["entries"].append({
                "id": entry_id,
                "category": category,
                "source": str(dest_path.relative_to(self.kb_dir)),
                "title": entry["title"],
                "content": entry["content"][:2000],  # 限制长度
                "metadata": {},
                "created": datetime.now().isoformat(),
            })
            count += 1

        self._save_index(index)
        print(f"✅ 已回灌 {count} 条知识到知识库")
        print(f"   分类：{CATEGORY_ICONS.get(category, '')} {category}")
        print(f"   来源：{file_path}")
        return count

    def ingest(self, file_path: str, category: str, module: str = ""):
        """自动检测文件类型并回灌"""
        path = Path(file_path)
        if not path.exists():
            print(f"❌ 文件不存在: {file_path}", file=sys.stderr)
            return 0

        if path.suffix.lower() in (".xlsx", ".xls"):
            return self.ingest_excel(file_path, category, module)
        elif path.suffix.lower() in (".md", ".markdown", ".txt"):
            return self.ingest_markdown(file_path, category)
        else:
            print(f"⚠️  不支持的文件格式: {path.suffix}", file=sys.stderr)
            return 0

    def export(self, query: str, output: str = None, category: str = None, limit: int = 20) -> str:
        """导出增强上下文为 Markdown"""
        results = self.search(query, category=category, limit=limit)

        if not results:
            content = f"# 知识库增强上下文\n\n> 检索关键词：{query} | 未命中相关知识\n"
        else:
            lines = [f"# 知识库增强上下文\n"]
            lines.append(f"> 检索关键词：{query} | 命中 {len(results)} 条相关知识\n")

            # 按分类分组
            by_cat = defaultdict(list)
            for r in results:
                by_cat[r["category"]].append(r)

            for cat in CATEGORIES:
                if cat not in by_cat:
                    continue
                items = by_cat[cat]
                icon = CATEGORY_ICONS.get(cat, "")
                label = CATEGORY_LABELS.get(cat, cat)
                lines.append(f"## {icon} {label}（{len(items)} 条）\n")

                for i, r in enumerate(items, 1):
                    lines.append(f"### {i}. [{r['source']}] {r['title']}\n")
                    # 添加元数据
                    meta = r.get("metadata", {})
                    if meta.get("module"):
                        lines.append(f"**模块：** {meta['module']}")
                    if meta.get("priority"):
                        lines.append(f"**优先级：** {meta['priority']}")
                    if meta.get("severity"):
                        lines.append(f"**严重程度：** {meta['severity']}")
                    if meta:
                        lines.append("")
                    lines.append(f"{r['snippet']}\n")

            content = "\n".join(lines)

        if output:
            Path(output).write_text(content, encoding="utf-8")
            print(f"✅ 知识上下文已导出: {output}")

        return content

    def status(self):
        """知识库概况"""
        self._ensure_dir()
        index = self._load_index()
        entries = index.get("entries", [])

        print(f"📊 知识库概况\n")
        print(f"   路径：{self.kb_dir}")
        print(f"   创建时间：{index.get('created', 'N/A')[:19]}")
        print(f"   最后更新：{index.get('updated', 'N/A')[:19]}")
        print(f"   总条目：{len(entries)}\n")

        # 按分类统计
        cat_counts = defaultdict(int)
        for e in entries:
            cat_counts[e["category"]] += 1

        print(f"   {'分类':<20} {'条目数':>6}")
        print(f"   {'─' * 30}")
        for cat in CATEGORIES:
            count = cat_counts.get(cat, 0)
            icon = CATEGORY_ICONS.get(cat, "")
            label = CATEGORY_LABELS.get(cat, cat)
            print(f"   {icon} {label:<18} {count:>6}")

        # 文件统计
        print(f"\n   {'目录':<20} {'文件数':>6}")
        print(f"   {'─' * 30}")
        for cat in CATEGORIES:
            cat_dir = self.kb_dir / cat
            file_count = len(list(cat_dir.glob("*.md"))) if cat_dir.exists() else 0
            print(f"   {cat:<20} {file_count:>6}")

    def list_entries(self, category: str = None, limit: int = 50):
        """列出知识条目"""
        index = self._load_index()
        entries = index.get("entries", [])

        if category:
            entries = [e for e in entries if e["category"] == category]

        if not entries:
            print("（空）")
            return

        print(f"📚 知识条目（{len(entries)} 条）\n")
        print(f"   {'ID':<30} {'分类':<18} {'标题':<40}")
        print(f"   {'─' * 90}")

        for i, e in enumerate(entries[:limit]):
            cat_label = CATEGORY_LABELS.get(e["category"], e["category"])
            title = e["title"][:38]
            eid = e["id"][:28]
            print(f"   {eid:<30} {cat_label:<18} {title}")

        if len(entries) > limit:
            print(f"\n   ... 还有 {len(entries) - limit} 条")


# ═══════════════════════════════════════════════════════════════
# 主函数
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="测试知识库管理器 — 本地优先的 RAG 知识库"
    )
    subparsers = parser.add_subparsers(dest="command", help="操作命令")

    # 默认知识库路径
    default_kb = str(Path.home() / "Documents" / "ai-test-system" / "knowledge-base")

    # init
    p_init = subparsers.add_parser("init", help="初始化知识库")
    p_init.add_argument("--kb-dir", default=default_kb, help="知识库目录")

    # search
    p_search = subparsers.add_parser("search", help="检索知识库")
    p_search.add_argument("query", help="检索关键词")
    p_search.add_argument("--kb-dir", default=default_kb, help="知识库目录")
    p_search.add_argument("--category", choices=CATEGORIES, help="限定分类")
    p_search.add_argument("--limit", type=int, default=20, help="返回条数")

    # add
    p_add = subparsers.add_parser("add", help="添加单条知识")
    p_add.add_argument("--category", choices=CATEGORIES, required=True, help="知识分类")
    p_add.add_argument("--title", required=True, help="标题")
    p_add.add_argument("--content", required=True, help="内容")
    p_add.add_argument("--module", default="", help="所属模块")
    p_add.add_argument("--severity", default="", choices=["", "high", "medium", "low", "critical"], help="严重程度")
    p_add.add_argument("--kb-dir", default=default_kb, help="知识库目录")

    # ingest
    p_ingest = subparsers.add_parser("ingest", help="从文件回灌知识")
    p_ingest.add_argument("file", help="源文件路径")
    p_ingest.add_argument("--category", choices=CATEGORIES, required=True, help="目标分类")
    p_ingest.add_argument("--module", default="", help="所属模块（用于 Excel）")
    p_ingest.add_argument("--kb-dir", default=default_kb, help="知识库目录")

    # export
    p_export = subparsers.add_parser("export", help="导出增强上下文")
    p_export.add_argument("query", help="检索关键词")
    p_export.add_argument("--output", default="", help="输出文件路径")
    p_export.add_argument("--kb-dir", default=default_kb, help="知识库目录")
    p_export.add_argument("--category", choices=CATEGORIES, help="限定分类")
    p_export.add_argument("--limit", type=int, default=20, help="返回条数")

    # status
    p_status = subparsers.add_parser("status", help="知识库概况")
    p_status.add_argument("--kb-dir", default=default_kb, help="知识库目录")

    # list
    p_list = subparsers.add_parser("list", help="列出知识条目")
    p_list.add_argument("--kb-dir", default=default_kb, help="知识库目录")
    p_list.add_argument("--category", choices=CATEGORIES, help="限定分类")
    p_list.add_argument("--limit", type=int, default=50, help="返回条数")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 1

    kb = KnowledgeBase(args.kb_dir)

    if args.command == "init":
        kb.init()

    elif args.command == "search":
        results = kb.search(args.query, category=args.category, limit=args.limit)
        if not results:
            print(f"🔍 检索 \"{args.query}\" — 未命中")
            return 0

        print(f"🔍 检索 \"{args.query}\" — 命中 {len(results)} 条\n")
        for i, r in enumerate(results, 1):
            icon = CATEGORY_ICONS.get(r["category"], "")
            label = CATEGORY_LABELS.get(r["category"], r["category"])
            print(f"{i}. [{icon} {label}] {r['title']}")
            print(f"   来源：{r['source']} | 相关度：{r['score']}")
            print(f"   {r['snippet'][:150]}...")
            print()

    elif args.command == "add":
        kb.add(args.category, args.title, args.content, args.module, args.severity)

    elif args.command == "ingest":
        kb.ingest(args.file, args.category, args.module)

    elif args.command == "export":
        output = args.output or None
        content = kb.export(args.query, output=output, category=args.category, limit=args.limit)
        if not output:
            print(content)

    elif args.command == "status":
        kb.status()

    elif args.command == "list":
        kb.list_entries(category=args.category, limit=args.limit)

    return 0


if __name__ == "__main__":
    sys.exit(main())
