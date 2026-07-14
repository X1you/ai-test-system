#!/usr/bin/env python3
"""
知识库管理器 - MCP 层集成
通过 MCP 协议访问 Obsidian Vault
"""

import sys
import os
from pathlib import Path
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
from collections import defaultdict
from datetime import datetime
import json

# 添加 scripts 目录到路径
sys.path.insert(0, str(Path(__file__).parent))

from mcp_client import MCPClient, create_mcp_client, KnowledgeItem

# ============================================================================
# 配置
# ============================================================================

HERMES_PYTHON = str(Path.home() / ".hermes" / "hermes-agent" / "venv" / "bin" / "python")
OBSIDIAN_VAULT = str(Path.home() / "Documents" / "test-interview-kb")

# ============================================================================
# 知识库管理器 (MCP 层)
# ============================================================================

class KnowledgeBaseManager:
    """知识库管理器 - 通过 MCP 协议访问 Obsidian Vault"""

    def __init__(self, vault_path: str = OBSIDIAN_VAULT):
        self.mcp_client = MCPClient(vault_path)

    def search(self, query: str, category: str = None, limit: int = 20) -> List[Dict]:
        """检索知识库"""
        return self.mcp_client.search(query, category=category, limit=limit)

    def add(self, item: KnowledgeItem) -> bool:
        """添加知识条目"""
        return self.mcp_client.create_file(item)

    def ingest(self, source_file: str, category: str, module: str = "",
               project: str = "", batch: str = "") -> int:
        """回灌知识 (从 Excel 或 Markdown)

        Excel 格式自动识别:
        - testcases.xlsx 标准12列格式（用例编号|模块|功能点|维度|标题|优先级|前置条件|步骤|数据|预期|备注|结果）
        - 通用3列格式（标题|内容|标签）

        归档结构（历史用例）:
          🏆 历史用例/{项目名}/{批次日期}/TC-001 xxx.md
          如未指定 project/batch，则用文件名和当前日期兜底

        其他分类: 仍按平铺存储（业务规则/坑点等天然是独立条目）
        """
        if not os.path.exists(source_file):
            print(f"❌ 文件不存在: {source_file}")
            return 0

        count = 0

        # 兜底项目名和批次名
        if not project:
            project = Path(source_file).stem  # testcases → testcases
        if not batch:
            batch = datetime.now().strftime("%Y-%m-%d")

        if source_file.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(source_file)
                ws = wb.active

                # 自动检测 Excel 格式
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                is_testcase_format = headers and headers[0] == '用例编号' and '用例标题' in headers

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or len(row) < 3:
                        continue

                    if is_testcase_format:
                        # 标准 testcases.xlsx 12列格式
                        tc_id = str(row[0] or '')
                        tc_module = str(row[1] or '')
                        tc_feature = str(row[2] or '')
                        tc_dimension = str(row[3] or '')
                        tc_title = str(row[4] or '')
                        tc_priority = str(row[5] or '')
                        tc_precondition = str(row[6] or '')
                        tc_steps = str(row[7] or '')
                        tc_test_data = str(row[8] or '')
                        tc_expected = str(row[9] or '')
                        tc_result = str(row[11] or '')

                        if not tc_title:
                            continue

                        # 组装结构化内容
                        content_lines = [
                            f"**用例编号**: {tc_id}",
                            f"**模块**: {tc_module}",
                            f"**功能点**: {tc_feature}",
                            f"**测试维度**: {tc_dimension}",
                            f"**优先级**: {tc_priority}",
                            f"",
                            f"**前置条件**: {tc_precondition}",
                            f"**测试步骤**:",
                            tc_steps,
                            f"",
                            f"**测试数据**: {tc_test_data}",
                            f"**预期结果**: {tc_expected}",
                        ]
                        if tc_result:
                            content_lines.append(f"**执行结果**: {tc_result}")

                        item = KnowledgeItem(
                            id='',
                            title=f"{tc_id} {tc_title}" if tc_id else tc_title,
                            content='\n'.join(content_lines),
                            category=category,
                            module=tc_module or module,
                            tags=[t.strip() for t in [tc_dimension, tc_priority, tc_feature] if t.strip()]
                        )
                        # 历史用例按项目维度归档
                        if category == 'historical-cases':
                            # 把项目和批次信息编码到 module 字段
                            item.module = f"{project}/{batch}/{tc_module or module}"
                        if self.add(item):
                            count += 1
                    else:
                        # 通用3列格式 (title, content, tags)
                        title = str(row[0] or '')
                        content = str(row[1] or '')
                        tags = str(row[2] or '').split(',')

                        if title and content:
                            item = KnowledgeItem(
                                id='',
                                title=title,
                                content=content,
                                category=category,
                                module=module,
                                tags=[t.strip() for t in tags if t.strip()]
                            )
                            if self.add(item):
                                count += 1
            except ImportError:
                print("❌ openpyxl 未安装，无法处理 Excel 文件")
                print("   安装: pip install openpyxl")

        elif source_file.endswith('.md'):
            # Markdown 回灌
            with open(source_file, 'r', encoding='utf-8') as f:
                content = f.read()
                title = Path(source_file).stem

                item = KnowledgeItem(
                    id='',
                    title=title,
                    content=content,
                    category=category,
                    module=module,
                    tags=[category, module]
                )
                if self.add(item):
                    count = 1
        else:
            print(f"❌ 不支持的文件格式: {source_file}")
            return 0

        print(f"✅ 已回灌 {count} 条知识到知识库")
        return count

    def export(self, query: str, output_file: str = "knowledge-context.md") -> str:
        """导出增强上下文 (Markdown 格式)"""
        results = self.search(query, limit=50)

        # 按分类分组
        grouped = defaultdict(list)
        for item in results:
            grouped[item['category']].append(item)

        # 生成 Markdown
        lines = [
            "# 知识库增强上下文",
            "",
            f"> 检索关键词: {query} | 命中 {len(results)} 条相关知识",
            "",
            f"> 来源: Obsidian Vault (MCP 协议)",
            ""
        ]

        for category, items in grouped.items():
            cat_display = {
                'business-rules': '📋 业务规则',
                'historical-cases': '🏆 历史优质用例',
                'pitfalls': '⚠️ 线上坑点',
                'templates': '📝 用例模板',
                'data-dictionary': '📖 数据字典',
                'business-specs': '📘 业务规范',
                'team-standards': '📐 团队规范',
            }.get(category, category)

            lines.append(f"## {cat_display} ({len(items)} 条)")

            for idx, item in enumerate(items, 1):
                lines.append(f"### {idx}. [[{item['filepath']}]]")
                lines.append(f"**模块**: {item.get('module', 'N/A')}")
                if item.get('severity'):
                    lines.append(f"**严重级别**: {item['severity']}")

                # 提取内容片段 (前 500 字)
                content_preview = item['content'][:500]
                if '---' in content_preview:
                    content_preview = content_preview.split('---', 2)[-1].strip()

                lines.append(content_preview + ("..." if len(item['content']) > 500 else ""))
                lines.append("")

        # 写入文件
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))

        print(f"✅ 已导出知识上下文: {output_path}")
        return str(output_path)

    def status(self) -> Dict:
        """知识库概况"""
        return self.mcp_client.status()

    def add_single(self, title: str, content: str, category: str, module: str = "",
                   tags: List[str] = None, severity: str = None) -> bool:
        """添加单条知识"""
        item = KnowledgeItem(
            id='',
            title=title,
            content=content,
            category=category,
            module=module,
            tags=tags or [],
            severity=severity
        )
        return self.add(item)


# ============================================================================
# CLI 入口
# ============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(description="知识库管理器 (MCP 层)")
    subparsers = parser.add_subparsers(dest='command', required=True)

    # search 命令
    search_parser = subparsers.add_parser('search', help='检索知识库')
    search_parser.add_argument('query', help='检索关键词')
    search_parser.add_argument('--category', choices=['business-rules', 'historical-cases', 'pitfalls', 'templates', 'data-dictionary', 'business-specs', 'team-standards'], help='限定分类')
    search_parser.add_argument('--limit', type=int, default=20, help='返回条数')

    # add 命令
    add_parser = subparsers.add_parser('add', help='添加单条知识')
    add_parser.add_argument('--title', required=True, help='标题')
    add_parser.add_argument('--content', required=True, help='内容')
    add_parser.add_argument('--category', required=True, choices=['business-rules', 'historical-cases', 'pitfalls', 'templates', 'data-dictionary', 'business-specs', 'team-standards'])
    add_parser.add_argument('--module', default='', help='所属模块')
    add_parser.add_argument('--tags', default='', help='标签 (逗号分隔)')
    add_parser.add_argument('--severity', choices=['high', 'medium', 'low'], help='严重级别 (仅坑点)')

    # ingest 命令
    ingest_parser = subparsers.add_parser('ingest', help='回灌知识文件')
    ingest_parser.add_argument('source_file', help='源文件 (Excel 或 Markdown)')
    ingest_parser.add_argument('--category', required=True, choices=['business-rules', 'historical-cases', 'pitfalls', 'templates', 'data-dictionary', 'business-specs', 'team-standards'])
    ingest_parser.add_argument('--module', default='', help='所属模块')
    ingest_parser.add_argument('--project', default='', help='项目名 (历史用例归档: 🏆 历史用例/项目名/批次/)')
    ingest_parser.add_argument('--batch', default='', help='批次名/日期 (历史用例归档)')

    # export 命令
    export_parser = subparsers.add_parser('export', help='导出增强上下文')
    export_parser.add_argument('query', help='检索关键词')
    export_parser.add_argument('--output', default='knowledge-context.md', help='输出文件路径')

    # status 命令
    subparsers.add_parser('status', help='知识库概况')

    args = parser.parse_args()

    # 创建管理器
    kb = KnowledgeBaseManager()

    # 执行命令
    if args.command == 'search':
        results = kb.search(args.query, category=args.category, limit=args.limit)
        print(json.dumps(results, ensure_ascii=False, indent=2))

    elif args.command == 'add':
        tags = [t.strip() for t in args.tags.split(',')] if args.tags else []
        kb.add_single(args.title, args.content, args.category, args.module, tags, args.severity)

    elif args.command == 'ingest':
        kb.ingest(args.source_file, args.category, args.module,
                  project=getattr(args, 'project', ''),
                  batch=getattr(args, 'batch', ''))

    elif args.command == 'export':
        kb.export(args.query, args.output)

    elif args.command == 'status':
        summary = kb.status()
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()