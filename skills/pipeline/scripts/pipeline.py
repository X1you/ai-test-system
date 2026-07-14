#!/usr/bin/env python3
"""
测试用例生成全流程 Pipeline 引擎

自动串联 7 个步骤：
  Step 1  需求分析
  Step 2  知识库检索 (RAG)
  Step 3  测试点梳理
  Step 4  生成测试用例
  Step 5  用例评审 + 知识库回灌
  Step 6  [人工执行测试]
  Step 7  生成测试报告 + 知识库回灌

用法:
    python pipeline.py run     <requirements.md> [--output DIR] [--mode auto|semi|step]
    python pipeline.py continue [--output DIR]
    python pipeline.py status  [--output DIR]
    python pipeline.py report  [--output DIR]

依赖:
    openpyxl (在 Hermes venv 中)
"""

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# ═══════════════════════════════════════════════════════════════
# 路径常量
# ═══════════════════════════════════════════════════════════════

HERMES_VENV = Path.home() / ".hermes" / "hermes-agent" / "venv" / "bin" / "python"
SKILLS_DIR = Path.home() / ".hermes" / "skills"

# 脚本路径
SCRIPT_GEN_EXCEL = SKILLS_DIR / "generate-testcases" / "scripts" / "generate_excel.py"
SCRIPT_GEN_XMIND = SKILLS_DIR / "generate-testcases" / "scripts" / "generate_xmind.py"
SCRIPT_GEN_REPORT = SKILLS_DIR / "generate-report" / "scripts" / "generate_report.py"
SCRIPT_KB_MCP = SKILLS_DIR / "knowledge-base" / "scripts" / "kb_manager_mcp.py"
SCRIPT_KB_LOCAL = SKILLS_DIR / "knowledge-base" / "scripts" / "kb_manager.py"

PROJECT_DIR = Path.home() / "Documents" / "ai-test-system"
DEFAULT_KB_DIR = PROJECT_DIR / "knowledge-base"


# ═══════════════════════════════════════════════════════════════
# 工具函数
# ═══════════════════════════════════════════════════════════════

def log(msg, level="INFO"):
    """统一日志输出"""
    icons = {"INFO": "  ", "STEP": "▶ ", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "HUMAN": "👤", "KB": "🧠"}
    icon = icons.get(level, "")
    print(f"{icon} {msg}")


def run_script(script_path: str, args: list, timeout: int = 120) -> dict:
    """运行子脚本，返回结果"""
    cmd = [str(HERMES_VENV), str(script_path)] + args
    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=timeout,
    )
    return {
        "returncode": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
    }


def extract_keywords_from_md(md_path: str, max_keywords: int = 5) -> str:
    """从需求分析 Markdown 中提取关键词"""
    content = Path(md_path).read_text(encoding="utf-8")

    # 提取模块名称
    modules = re.findall(r"模块[一二三四五六七八九十]+[：:]\s*(.+)", content)
    # 提取功能点
    features = re.findall(r"功能点\s*[\d.]+[：:]\s*(.+)", content)

    keywords = modules[:3] + features[:3]
    # 去重，取前 N 个
    seen = set()
    unique = []
    for kw in keywords:
        kw = kw.strip()
        if kw and kw not in seen:
            seen.add(kw)
            unique.append(kw)
        if len(unique) >= max_keywords:
            break

    return " ".join(unique) if unique else "测试"


def count_testpoints(md_path: str) -> int:
    """统计测试点数量"""
    content = Path(md_path).read_text(encoding="utf-8")
    return len(re.findall(r"-\s+测试点\s*[\d.]+[：:]", content))


def count_cases(xlsx_path: str) -> int:
    """统计 Excel 中用例数"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        count = ws.max_row - 1
        wb.close()
        return count
    except ImportError:
        pass
    try:
        result = run_script(str(HERMES_VENV), ["-c", (
            "from openpyxl import load_workbook\n"
            f"wb = load_workbook('{xlsx_path}', data_only=True)\n"
            "ws = wb.active\n"
            "print(ws.max_row - 1)\n"
            "wb.close()\n"
        )], timeout=30)
        if result["returncode"] == 0:
            return int(result["stdout"].strip())
    except Exception:
        pass
    return 0


def check_has_results(xlsx_path: str) -> bool:
    """检查 Excel 是否已填写执行结果（直接 import，不用 subprocess）"""
    try:
        # 尝试直接导入 openpyxl（如果当前环境有）
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, data_only=True)
            ws = wb.active
            result_col = None
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col).value or "").strip()
                if header == "执行结果":
                    result_col = col
                    break
            if not result_col:
                for col in range(1, ws.max_column + 1):
                    header = str(ws.cell(row=1, column=col).value or "").strip()
                    if "执行" in header and "结果" in header:
                        result_col = col
                        break
            if not result_col:
                wb.close()
                return False
            filled = 0
            for row in range(2, ws.max_row + 1):
                val = str(ws.cell(row=row, column=result_col).value or "").strip()
                if val:
                    filled += 1
            wb.close()
            return filled > 0
        except ImportError:
            pass

        # 回退：用 subprocess，但通过环境变量避免编码问题
        result = run_script(str(HERMES_VENV), ["-c", (
            "from openpyxl import load_workbook\n"
            f"wb = load_workbook('{xlsx_path}', data_only=True)\n"
            "ws = wb.active\n"
            "result_col = None\n"
            "for col in range(1, ws.max_column + 1):\n"
            "    header = str(ws.cell(row=1, column=col).value or '').strip()\n"
            "    if header == '\u6267\u884c\u7ed3\u679c':\n"  # 执行结果
            "        result_col = col\n"
            "        break\n"
            "if not result_col:\n"
            "    for col in range(1, ws.max_column + 1):\n"
            "        header = str(ws.cell(row=1, column=col).value or '').strip()\n"
            "        if '\u6267\u884c' in header and '\u7ed3\u679c' in header:\n"
            "            result_col = col\n"
            "            break\n"
            "if not result_col:\n"
            "    print('no_col')\n"
            "else:\n"
            "    filled = 0\n"
            "    for row in range(2, ws.max_row + 1):\n"
            "        val = str(ws.cell(row=row, column=result_col).value or '').strip()\n"
            "        if val:\n"
            "            filled += 1\n"
            "    print(filled)\n"
            "wb.close()\n"
        )], timeout=30)
        if result["returncode"] != 0:
            return False
        val = result["stdout"].strip()
        if val == "no_col":
            return False
        try:
            return int(val) > 0
        except ValueError:
            return False
    except Exception:
        return False
    return False


# ═══════════════════════════════════════════════════════════════
# Pipeline 状态管理
# ═══════════════════════════════════════════════════════════════

STATE_FILE = "_pipeline_state.json"

STEPS = [
    {"id": 1, "name": "需求分析", "output": "requirements_analysis.md"},
    {"id": 2, "name": "知识库检索", "output": "knowledge-context.md"},
    {"id": 3, "name": "测试点梳理", "output": "testpoints.md"},
    {"id": 4, "name": "生成测试用例", "output": "testcases.xlsx"},
    {"id": 5, "name": "用例评审", "output": "test_case_review_report.md"},
    {"id": 6, "name": "执行测试（人工）", "output": "testcases.xlsx"},
    {"id": 7, "name": "生成测试报告", "output": "test_report.md"},
]


def load_state(output_dir: str) -> dict:
    """加载 pipeline 状态"""
    state_path = Path(output_dir) / STATE_FILE
    if state_path.exists():
        return json.loads(state_path.read_text(encoding="utf-8"))
    return {
        "started": datetime.now().isoformat(),
        "completed_steps": [],
        "step_results": {},
        "mode": "semi",
        "requirements_file": "",
        "kb_dir": str(DEFAULT_KB_DIR),
    }


def save_state(output_dir: str, state: dict):
    """保存 pipeline 状态"""
    state["updated"] = datetime.now().isoformat()
    state_path = Path(output_dir) / STATE_FILE
    state_path.parent.mkdir(parents=True, exist_ok=True)
    state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def mark_step_done(output_dir: str, state: dict, step_id: int, result_info: dict = None):
    """标记步骤完成"""
    if step_id not in state["completed_steps"]:
        state["completed_steps"].append(step_id)
    if result_info:
        state["step_results"][str(step_id)] = result_info
    save_state(output_dir, state)


def is_step_done(state: dict, step_id: int) -> bool:
    """检查步骤是否已完成"""
    return step_id in state["completed_steps"]


def check_output_exists(output_dir: str, filename: str) -> bool:
    """检查输出文件是否存在"""
    return (Path(output_dir) / filename).exists()


# ═══════════════════════════════════════════════════════════════
# Pipeline 步骤实现
# ═══════════════════════════════════════════════════════════════

class Pipeline:
    def __init__(self, output_dir: str, kb_dir: str = None):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.kb_dir = kb_dir or str(DEFAULT_KB_DIR)
        self.kb_available = Path(self.kb_dir).exists() and (Path(self.kb_dir) / "index.json").exists()

    def _out(self, filename: str) -> str:
        return str(self.output_dir / filename)

    # ─── Step 1: 需求分析 ───
    def step1_analysis(self, requirements_file: str) -> dict:
        log("Step 1/7: 需求分析", "STEP")
        req_path = Path(requirements_file)
        if not req_path.exists():
            log(f"需求文档不存在: {requirements_file}", "ERR")
            return {"ok": False, "error": "文件不存在"}

        content = req_path.read_text(encoding="utf-8")
        analysis_path = self.output_dir / "requirements_analysis.md"
        clarification_path = self.output_dir / "clarification_needed.md"

        # AI 在实际使用时直接生成这两个文件
        # 脚本模式下：复制需求文档作为输入，标记步骤完成
        # 真正的需求分析由 AI 完成，这里只做占位
        log("  （需求分析由 AI 直接执行，输出 requirements_analysis.md）", "INFO")
        log("  请确保 AI 已生成 requirements_analysis.md", "INFO")

        if analysis_path.exists():
            # 统计模块和功能点
            modules = len(re.findall(r"模块[一二三四五六七八九十]+[：:]", content))
            features = len(re.findall(r"功能点\s*[\d.]+[：:]", content))
            log(f"需求分析完成 — {modules} 模块, {features} 功能点", "OK")
            return {"ok": True, "modules": modules, "features": features}
        else:
            log("requirements_analysis.md 尚未生成", "WARN")
            log("请由 AI 执行需求分析后重试", "INFO")
            return {"ok": False, "error": "等待 AI 生成需求分析"}

    # ─── Step 2: 知识库检索 ───
    def step2_kb_search(self, keywords: str) -> dict:
        log("Step 2/7: 知识库检索 (RAG)", "STEP")

        # 优先使用 MCP 层知识库（kb_manager_mcp.py），fallback 到本地知识库
        kb_script = None
        if SCRIPT_KB_MCP.exists():
            kb_script = SCRIPT_KB_MCP
        elif SCRIPT_KB_LOCAL.exists():
            kb_script = SCRIPT_KB_LOCAL
        else:
            log("知识库脚本未找到，跳过 RAG 检索", "WARN")
            return {"ok": True, "skipped": True, "hits": 0}

        context_path = self._out("knowledge-context.md")
        result = run_script(str(kb_script), [
            "export", keywords,
            "--output", context_path,
        ], timeout=60)

        if result["returncode"] == 0:
            # 统计命中数
            try:
                ctx_content = Path(context_path).read_text(encoding="utf-8")
                hits = ctx_content.count("### ")
            except Exception:
                hits = 0

            if hits > 0:
                log(f"知识库命中 {hits} 条相关知识", "KB")
            else:
                log("知识库未命中相关知识", "KB")
            return {"ok": True, "hits": hits, "skipped": False}
        else:
            log(f"知识库检索出错: {result['stderr'][:100]}", "WARN")
            return {"ok": True, "skipped": True, "hits": 0}

    # ─── Step 3: 测试点梳理 ───
    def step3_testpoints(self) -> dict:
        log("Step 3/7: 测试点梳理", "STEP")
        tp_path = self.output_dir / "testpoints.md"

        if not tp_path.exists():
            log("testpoints.md 尚未生成（由 AI 执行测试点梳理 Skill）", "WARN")
            return {"ok": False, "error": "等待 AI 生成测试点"}

        count = count_testpoints(str(tp_path))
        log(f"测试点梳理完成 — {count} 个测试点", "OK")
        return {"ok": True, "count": count}

    # ─── Step 4: 生成测试用例 ───
    def step4_generate_cases(self, dimensions: str = "basic", formats: str = "excel") -> dict:
        log("Step 4/7: 生成测试用例", "STEP")
        tp_path = self._out("testpoints.md")

        if not Path(tp_path).exists():
            log("testpoints.md 不存在，无法生成用例", "ERR")
            return {"ok": False, "error": "缺少测试点文件"}

        # Excel
        if "excel" in formats:
            xlsx_path = self._out("testcases.xlsx")

            # 如果 Excel 已存在且有执行结果，不覆盖
            if Path(xlsx_path).exists() and check_has_results(xlsx_path):
                case_count = count_cases(xlsx_path)
                log(f"Excel 已存在（含执行结果），跳过重新生成 — {case_count} 条", "OK")
                return {"ok": True, "case_count": case_count, "reused": True}

            result = run_script(str(SCRIPT_GEN_EXCEL), [
                tp_path, "-o", xlsx_path, "-d", dimensions,
            ], timeout=120)
            if result["returncode"] == 0:
                case_count = count_cases(xlsx_path)
                log(f"Excel 用例生成完成 — {case_count} 条", "OK")
            else:
                log(f"Excel 生成失败: {result['stderr'][:200]}", "ERR")
                return {"ok": False, "error": result["stderr"][:200]}

        # XMind
        if "xmind" in formats:
            xmind_path = self._out("testcases.xmind")
            xmind_args = [tp_path, "-o", xmind_path]
            if dimensions != "all":
                xmind_args.extend(["-d", dimensions])
            result = run_script(str(SCRIPT_GEN_XMIND), xmind_args, timeout=120)
            if result["returncode"] == 0:
                log(f"XMind 用例生成完成", "OK")
            else:
                log(f"XMind 生成失败: {result['stderr'][:200]}", "WARN")

        return {"ok": True, "case_count": case_count if "excel" in formats else 0}

    # ─── Step 5: 用例评审 ───
    def step5_review(self) -> dict:
        log("Step 5/7: 用例评审", "STEP")
        review_path = self.output_dir / "test_case_review_report.md"

        if not review_path.exists():
            log("用例评审由 AI 直接执行（test-case-review Skill）", "INFO")
            log("请由 AI 完成评审后继续", "INFO")
            return {"ok": False, "error": "等待 AI 完成评审"}

        log("用例评审完成", "OK")
        return {"ok": True}

    # ─── Step 5+: 知识库回灌 ───
    def step5_kb_ingest(self) -> dict:
        log("知识库回灌：优质用例", "KB")

        xlsx_path = self._out("testcases.xlsx")
        if not Path(xlsx_path).exists():
            return {"ok": True, "skipped": True}

        # 优先使用 MCP 层知识库
        kb_script = None
        if SCRIPT_KB_MCP.exists():
            kb_script = SCRIPT_KB_MCP
        elif SCRIPT_KB_LOCAL.exists():
            kb_script = SCRIPT_KB_LOCAL
        else:
            log("知识库脚本未找到，跳过回灌", "WARN")
            return {"ok": True, "skipped": True}

        result = run_script(str(kb_script), [
            "ingest", xlsx_path,
            "--category", "historical-cases",
        ], timeout=60)

        if result["returncode"] == 0:
            log("优质用例已回灌到知识库", "KB")
            return {"ok": True}
        else:
            log(f"回灌失败: {result['stderr'][:100]}", "WARN")
            return {"ok": True, "skipped": True}

    # ─── Step 6: 人工执行测试 ───
    def step6_human_test(self) -> dict:
        log("Step 6/7: 执行测试（人工）", "HUMAN")
        xlsx_path = self._out("testcases.xlsx")

        log(f"请打开 {xlsx_path} 执行测试", "HUMAN")
        log("在「执行结果」列填写：通过/失败/阻塞/跳过", "HUMAN")
        log("填写完成后，运行: pipeline.py continue --output <dir>", "HUMAN")

        # 检查是否已填写
        if check_has_results(xlsx_path):
            log("检测到执行结果已填写，继续下一步", "OK")
            return {"ok": True}

        return {"ok": False, "error": "等待人工执行测试", "human": True}

    # ─── Step 7: 生成测试报告 ───
    def step7_report(self) -> dict:
        log("Step 7/7: 生成测试报告", "STEP")
        xlsx_path = self._out("testcases.xlsx")
        report_path = self._out("test_report.md")

        result = run_script(str(SCRIPT_GEN_REPORT), [
            xlsx_path, "-o", report_path,
        ], timeout=60)

        if result["returncode"] == 0:
            # 提取通过率
            stdout = result["stdout"]
            pass_match = re.search(r"通过: (\d+) 个 \(([0-9.]+)%\)", stdout)
            if pass_match:
                log(f"测试报告生成完成 — 通过率 {pass_match.group(2)}%", "OK")
            else:
                log("测试报告生成完成", "OK")
            return {"ok": True}
        else:
            log(f"报告生成失败: {result['stderr'][:200]}", "ERR")
            return {"ok": False, "error": result["stderr"][:200]}

    # ─── Step 7+: 知识库回灌坑点 ───
    def step7_kb_ingest(self) -> dict:
        log("知识库回灌：失败分析 + 坑点", "KB")

        report_path = self._out("test_report.md")
        if not Path(report_path).exists():
            return {"ok": True, "skipped": True}

        # 优先使用 MCP 层知识库
        kb_script = None
        if SCRIPT_KB_MCP.exists():
            kb_script = SCRIPT_KB_MCP
        elif SCRIPT_KB_LOCAL.exists():
            kb_script = SCRIPT_KB_LOCAL
        else:
            log("知识库脚本未找到，跳过回灌", "WARN")
            return {"ok": True, "skipped": True}

        result = run_script(str(kb_script), [
            "ingest", report_path,
            "--category", "pitfalls",
        ], timeout=60)

        if result["returncode"] == 0:
            log("失败分析和坑点已回灌到知识库", "KB")
            return {"ok": True}
        else:
            return {"ok": True, "skipped": True}


# ═══════════════════════════════════════════════════════════════
# Pipeline 执行器
# ═══════════════════════════════════════════════════════════════

def run_pipeline(args):
    """执行 pipeline"""
    output_dir = args.output
    mode = args.mode
    requirements_file = args.requirements

    state = load_state(output_dir)
    state["mode"] = mode
    state["requirements_file"] = requirements_file
    state["kb_dir"] = args.kb_dir or str(DEFAULT_KB_DIR)
    save_state(output_dir, state)

    pipe = Pipeline(output_dir, kb_dir=state["kb_dir"])

    print()
    print("═" * 60)
    print(f"  🚀 全流程 Pipeline 启动 — 模式: {mode}")
    print(f"  📂 输出目录: {output_dir}")
    print(f"  📄 需求文档: {requirements_file}")
    print("═" * 60)
    print()

    # ─── Step 1: 需求分析 ───
    if is_step_done(state, 1) and check_output_exists(output_dir, "requirements_analysis.md"):
        log("Step 1 已完成，跳过", "OK")
    else:
        r = pipe.step1_analysis(requirements_file)
        if r["ok"]:
            mark_step_done(output_dir, state, 1, r)
        elif mode == "semi":
            log("⏸️  半自动模式：请由 AI 完成需求分析后运行 continue", "HUMAN")
            return
        else:
            log("需求分析失败，pipeline 终止", "ERR")
            return

    # ─── Step 2: 知识库检索 ───
    if is_step_done(state, 2):
        log("Step 2 已完成，跳过", "OK")
    else:
        keywords = extract_keywords_from_md(pipe._out("requirements_analysis.md"))
        log(f"检索关键词: {keywords}", "KB")
        r = pipe.step2_kb_search(keywords)
        mark_step_done(output_dir, state, 2, r)

    # ─── Step 3: 测试点梳理 ───
    if is_step_done(state, 3) and check_output_exists(output_dir, "testpoints.md"):
        log("Step 3 已完成，跳过", "OK")
    else:
        r = pipe.step3_testpoints()
        if r["ok"]:
            mark_step_done(output_dir, state, 3, r)
        else:
            log("⏸️  请由 AI 完成测试点梳理后运行 continue", "HUMAN")
            return

    # ─── Step 4: 生成测试用例 ───
    if is_step_done(state, 4) and check_output_exists(output_dir, "testcases.xlsx"):
        log("Step 4 已完成，跳过", "OK")
    else:
        r = pipe.step4_generate_cases(dimensions=args.dimensions, formats=args.formats)
        if r["ok"]:
            mark_step_done(output_dir, state, 4, r)
        else:
            log("用例生成失败，pipeline 终止", "ERR")
            return

    # ─── Step 5: 用例评审 ───
    if is_step_done(state, 5) and check_output_exists(output_dir, "test_case_review_report.md"):
        log("Step 5 已完成，跳过", "OK")
    else:
        r = pipe.step5_review()
        if r["ok"]:
            mark_step_done(output_dir, state, 5, r)
        else:
            log("⏸️  请由 AI 完成用例评审后运行 continue", "HUMAN")
            return

    # ─── Step 5+: 知识库回灌 ───
    pipe.step5_kb_ingest()

    # ─── Step 6: 人工执行测试 ───
    xlsx_path = pipe._out("testcases.xlsx")
    has_results = check_has_results(xlsx_path) if Path(xlsx_path).exists() else False
    log(f"检查执行结果: {has_results}", "INFO")
    if has_results:
        log("Step 6 检测到执行结果已填写", "OK")
        if not is_step_done(state, 6):
            mark_step_done(output_dir, state, 6, {"ok": True})
    else:
        r = pipe.step6_human_test()
        if not r["ok"]:
            log("", "INFO")
            log("⏸️  Pipeline 暂停 — 等待人工执行测试", "HUMAN")
            log(f"    文件: {xlsx_path}", "INFO")
            log(f"    填写完成后运行: pipeline.py continue -o {output_dir}", "INFO")
            return

    # ─── Step 7: 生成测试报告 ───
    if is_step_done(state, 7) and check_output_exists(output_dir, "test_report.md"):
        log("Step 7 已完成，跳过", "OK")
    else:
        r = pipe.step7_report()
        if r["ok"]:
            mark_step_done(output_dir, state, 7, r)
        else:
            log("报告生成失败，pipeline 终止", "ERR")
            return

    # ─── Step 7+: 知识库回灌 ───
    pipe.step7_kb_ingest()

    # ─── 汇总 ───
    print()
    print("═" * 60)
    print("  ✅ 全流程 Pipeline 执行完成！")
    print("═" * 60)
    print()
    print_pipeline_summary(output_dir, state)
    print()
    log(f"📁 所有输出文件: {output_dir}/", "INFO")


def continue_pipeline(args):
    """从断点继续"""
    output_dir = args.output
    state = load_state(output_dir)

    if not state.get("requirements_file"):
        log("未找到 pipeline 状态，请先运行 run", "ERR")
        return

    print()
    log(f"▶  继续执行 Pipeline — 已完成步骤: {state['completed_steps']}", "INFO")
    print()

    # 把当前状态当作 run 的参数重新跑，已完成的步骤会跳过
    args.requirements = state["requirements_file"]
    args.kb_dir = state.get("kb_dir")
    args.mode = state.get("mode", "semi")
    run_pipeline(args)


def show_status(args):
    """显示 pipeline 状态"""
    output_dir = args.output
    state = load_state(output_dir)

    print()
    print("═" * 60)
    print("  📊 Pipeline 状态")
    print("═" * 60)
    print()

    log(f"启动时间: {state.get('started', 'N/A')[:19]}", "INFO")
    log(f"最后更新: {state.get('updated', 'N/A')[:19]}", "INFO")
    log(f"执行模式: {state.get('mode', 'N/A')}", "INFO")
    log(f"需求文档: {state.get('requirements_file', 'N/A')}", "INFO")
    print()

    print("  步骤                    │ 状态   │ 输出文件")
    print("  " + "─" * 56)

    for step in STEPS:
        done = step["id"] in state["completed_steps"]
        exists = check_output_exists(output_dir, step["output"])
        icon = "✅" if done else ("📁" if exists else "⬜")
        status = "已完成" if done else ("文件存在" if exists else "待执行")
        print(f"  {icon} Step {step['id']}. {step['name']:<16} │ {status:<6} │ {step['output']}")

    print()
    done_count = len(state["completed_steps"])
    total = len(STEPS)
    log(f"进度: {done_count}/{total} 步 ({done_count/total*100:.0f}%)", "INFO")


def print_pipeline_summary(output_dir: str, state: dict):
    """打印 pipeline 汇总"""
    print(f"  {'步骤':<22} │ {'状态':<6} │ {'输出文件'}")
    print(f"  {'─' * 56}")

    for step in STEPS:
        done = step["id"] in state["completed_steps"]
        icon = "✅" if done else "⬜"
        status = "完成" if done else "待执行"
        print(f"  {icon} {step['id']}. {step['name']:<18} │ {status:<4} │ {step['output']}")


# ═══════════════════════════════════════════════════════════════
# 主函数
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="测试用例生成全流程 Pipeline"
    )
    subparsers = parser.add_subparsers(dest="command")

    default_out = str(Path.cwd() / "pipeline-output")

    # run
    p_run = subparsers.add_parser("run", help="执行全流程")
    p_run.add_argument("requirements", help="需求文档路径")
    p_run.add_argument("-o", "--output", default=default_out, help="输出目录")
    p_run.add_argument("--mode", choices=["auto", "semi", "step"], default="semi",
                       help="执行模式 (默认 semi)")
    p_run.add_argument("--dimensions", default="basic",
                       help="测试维度: all|basic|positive,negative")
    p_run.add_argument("--formats", default="excel", help="输出格式: excel|xmind|excel,xmind")
    p_run.add_argument("--kb-dir", default=None, help="知识库目录")

    # continue
    p_cont = subparsers.add_parser("continue", help="从断点继续")
    p_cont.add_argument("-o", "--output", default=default_out, help="输出目录")
    p_cont.add_argument("--dimensions", default="basic")
    p_cont.add_argument("--formats", default="excel")

    # status
    p_status = subparsers.add_parser("status", help="查看进度")
    p_status.add_argument("-o", "--output", default=default_out, help="输出目录")

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return 1

    if args.command == "run":
        run_pipeline(args)
    elif args.command == "continue":
        continue_pipeline(args)
    elif args.command == "status":
        show_status(args)

    return 0


if __name__ == "__main__":
    sys.exit(main())
