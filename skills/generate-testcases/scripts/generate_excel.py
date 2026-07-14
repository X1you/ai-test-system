#!/usr/bin/env python3
"""
生成 Excel 格式测试用例
读取测试点 Markdown 文件，生成结构化的 .xlsx 文件

用法:
    python generate_excel.py <testpoints.md> [--output testcases.xlsx] [--dimensions all|basic|positive,negative]
"""

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 缺少依赖库 openpyxl", file=sys.stderr)
    print("请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════
# 测试点解析器
# ═══════════════════════════════════════════════════════════════

class TestPointParser:
    """解析测试点 Markdown 文件"""

    def __parse(self, content: str) -> list:
        """解析 Markdown 内容，返回结构化的测试点列表"""
        test_points = []
        current_module = ""
        current_feature = ""
        current_dimension = ""
        current_number = 0

        for line in content.split("\n"):
            line = line.rstrip()

            # 模块
            # ## 模块一：订单创建  或  ## 模块一：订单创建
            m = re.match(r"^##\s+模块[一二三四五六七八九十]+[：:]\s*(.+)", line)
            if m:
                current_module = m.group(1).strip()
                continue

            # 功能点
            # ### 功能点 1.1：下单流程
            m = re.match(r"^###\s+功能点\s*[\d.]+[：:]\s*(.+)", line)
            if m:
                current_feature = m.group(1).strip()
                continue

            # 测试维度
            # #### 测试维度：正向测试
            m = re.match(r"^####\s+测试维度[：:]\s*(.+)", line)
            if m:
                current_dimension = m.group(1).strip()
                continue

            # 测试点
            # - 测试点 1.1.1：正常下单流程
            m = re.match(r"^-\s+测试点\s*[\d.]+[：:]\s*(.+)", line)
            if m:
                current_number += 1
                test_points.append({
                    "module": current_module,
                    "feature": current_feature,
                    "dimension": current_dimension,
                    "title": m.group(1).strip(),
                    "test_data": "",
                    "expected": "",
                    "number": current_number,
                })
                continue

            # 测试数据
            m = re.match(r"^\s+-\s+测试数据[：:]\s*(.+)", line)
            if m and test_points:
                test_points[-1]["test_data"] = m.group(1).strip()
                continue

            # 预期结果
            m = re.match(r"^\s+-\s+预期结果[：:]\s*(.+)", line)
            if m and test_points:
                test_points[-1]["expected"] = m.group(1).strip()
                continue

        return test_points

    def parse_file(self, file_path: str) -> list:
        """解析文件"""
        content = Path(file_path).read_text(encoding="utf-8")
        return self.__parse(content)


# ═══════════════════════════════════════════════════════════════
# 用例生成器
# ═════════════════════════════════════════════════════　　　　　　

class TestCaseGenerator:
    """根据测试点生成测试用例"""

    DIMENSION_MAP = {
        "正向测试": "positive",
        "负向测试": "negative",
        "边界测试": "boundary",
        "异常测试": "exception",
        "性能测试": "performance",
        "安全测试": "security",
    }

    def __generate_steps(self, tp: dict) -> str:
        """根据测试点生成具体的测试步骤"""
        title = tp["title"]
        dimension = tp["dimension"]
        module = tp["module"]
        feature = tp["feature"]

        # ── 正向测试 ──
        if "正向" in dimension:
            # 通用正向步骤模式
            steps = []
            steps.append("1. 准备测试环境，确认前置条件满足")
            # 如果是"登录/输入/填写"类操作
            if any(k in title for k in ["登录", "注册", "输入", "填写"]):
                steps.append("2. 输入必要的测试数据（详见测试数据字段）")
                steps.append("3. 点击确认/提交按钮")
            elif any(k in title for k in ["查询", "列表", "搜索", "筛选"]):
                steps.append("2. 进入查询页面/模块")
                steps.append("3. 输入查询条件（如有）")
                steps.append("4. 执行查询操作")
            elif any(k in title for k in ["导出", "下载", "导入", "上传"]):
                steps.append("2. 选择待操作的文件/数据")
                steps.append("3. 执行导入/导出操作")
                steps.append("4. 确认操作完成提示")
            elif any(k in title for k in ["创建", "新增", "添加", "提交"]):
                steps.append("2. 填写/选择必要的信息")
                steps.append("3. 点击提交/保存按钮")
                steps.append("4. 确认操作成功，检查返回结果")
            elif any(k in title for k in ["编辑", "修改", "更新", "变更"]):
                steps.append("2. 进入编辑模式")
                steps.append("3. 修改目标字段")
                steps.append("4. 保存修改")
            elif any(k in title for k in ["删除", "取消", "作废"]):
                steps.append("2. 定位目标数据/操作对象")
                steps.append("3. 执行删除/取消操作")
                steps.append("4. 确认删除/取消结果")
            elif any(k in title for k in ["校验", "验证", "检查", "对比"]):
                steps.append("2. 准备对比/校验的基准数据")
                steps.append("3. 执行校验操作")
                steps.append("4. 比对实际结果与预期是否一致")
            elif "集成" in title or "串联" in title or "顺序" in title:
                steps.append("2. 按顺序执行各步骤/操作")
                steps.append("3. 检查每一步的中间结果")
                steps.append("4. 确认最终结果与预期一致")
            elif any(k in title for k in ["超长", "大量", "大批量"]):
                steps.append("2. 准备超长/大量测试数据")
                steps.append("3. 执行操作")
                steps.append("4. 检查系统是否正常处理")
            else:
                steps.append("2. 执行目标操作")
                steps.append("3. 观察操作结果")
            return "\n".join(steps)

        # ── 负向测试 ──
        if "负向" in dimension:
            steps = ["1. 准备测试环境，确保系统处于可测试状态"]
            if any(k in title for k in ["为空", "空", "缺失", "缺少", "未填写"]):
                steps.append("2. 保持必填字段为空/不输入数据")
                steps.append("3. 提交/确认操作")
                steps.append("4. 检查系统是否给出正确的提示信息")
            elif any(k in title for k in ["非法", "无效", "错误", "异常字符", "特殊字符"]):
                steps.append(f"2. 在输入字段中输入非法/无效数据: {tp.get('test_data', '非法数据')}")
                steps.append("3. 提交/确认操作")
                steps.append("4. 检查系统是否拒绝并给出提示")
            elif any(k in title for k in ["未授权", "越权", "无权限", "未登录"]):
                steps.append("2. 使用无权限账号/未登录状态下尝试操作")
                steps.append("3. 观察系统是否阻止操作并提示")
            elif any(k in title for k in ["不存在", "不匹配", "错误", "不一致"]):
                steps.append(f"2. 使用不存在的/错误的测试数据: {tp.get('test_data', '无效数据')}")
                steps.append("3. 执行操作")
                steps.append("4. 检查系统是否给出正确的错误提示")
            elif "格式" in title:
                steps.append(f"2. 输入不符合格式要求的数据: {tp.get('test_data', '格式错误数据')}")
                steps.append("3. 提交操作")
                steps.append("4. 检查系统是否提示格式错误")
            else:
                steps.append(f"2. 准备非法条件/数据: {tp.get('test_data', '非法数据')}")
                steps.append("3. 执行操作")
                steps.append("4. 验证系统正确处理异常情况")
            return "\n".join(steps)

        # ── 边界测试 ──
        if "边界" in dimension:
            steps = ["1. 确定目标字段的边界值"]
            # 从标题中判断边界类型
            if any(k in title for k in ["最小值", "最小", "下限", "刚好", "恰好"]):
                steps.append(f"2. 设置测试数据为边界最小值")
            elif any(k in title for k in ["最大值", "最大", "上限", "超出", "超过"]):
                steps.append("2. 设置测试数据为边界最大值（或略超）")
            elif "超长" in title or "长度" in title:
                steps.append("2. 构造长度在边界值附近（±1）的测试数据")
            elif "数量" in title:
                steps.append("2. 设置数量为边界值（如刚好 N 个）")
            elif "金额" in title:
                steps.append(f"2. 设置金额为边界值（如 0.01 / 999999.99）")
            else:
                steps.append(f"2. 设置边界测试数据: {tp.get('test_data', '边界数据')}")
            if tp.get('test_data'):
                steps.append(f"   - 测试数据: {tp['test_data']}")
            steps.append("3. 执行操作")
            steps.append("4. 检查系统是否正确处理边界情况")
            return "\n".join(steps)

        # ── 异常测试 ──
        if "异常" in dimension:
            steps = ["1. 准备正常的测试环境"]
            if any(k in title for k in ["超时", "timeout", "超时"]):
                steps.append("2. 模拟网络超时/服务无响应场景")
                steps.append("3. 执行操作")
                steps.append("4. 观察系统是否给出超时提示")
            elif any(k in title for k in ["网络", "断网", "中断", "断开"]):
                steps.append("2. 执行操作过程中断开网络连接")
                steps.append("3. 检查系统是否有重试/恢复机制")
            elif any(k in title for k in ["并发", "同时", "并行"]):
                steps.append("2. 使用多线程/多进程模拟并发请求")
                steps.append("3. 检查并发处理结果的一致性")
            elif any(k in title for k in ["编码", "字符集"]):
                steps.append(f"2. 使用异常编码/字符集的数据: {tp.get('test_data', '非标编码数据')}")
                steps.append("3. 执行操作")
                steps.append("4. 检查是否正确处理编码问题")
            elif any(k in title for k in ["不可读", "不存在", "损坏", "异常"]):
                steps.append(f"2. 模拟异常条件: {tp.get('test_data', '异常条件')}")
                steps.append("3. 执行操作")
                steps.append("4. 验证系统的异常处理机制")
            else:
                steps.append(f"2. 模拟异常场景: {tp.get('test_data', '异常场景')}")
                steps.append("3. 执行操作")
                steps.append("4. 观察系统响应是否合理")
            return "\n".join(steps)

        # ── 性能测试 ──
        if "性能" in dimension:
            steps = ["1. 准备性能测试工具和监控环境"]
            if "并发" in title or "高并发" in title or "大量" in title:
                steps.append("2. 使用性能测试工具（如 JMeter/Locust）模拟并发用户")
                steps.append(f"3. 按测试数据设置并发数量和脚本参数")
                steps.append("4. 执行测试，记录响应时间、TPS、错误率等指标")
                steps.append("5. 与性能基线对比，判断是否达标")
            elif "响应时间" in title or "延迟" in title:
                steps.append("2. 单次执行目标操作")
                steps.append("3. 记录从请求发起到收到响应的时间")
                steps.append("4. 重复执行 N 次（N>=10），计算平均响应时间")
                steps.append("5. 判断是否满足性能要求")
            elif "效率" in title or "生成" in title:
                steps.append("2. 准备较大规模输入数据")
                steps.append("3. 执行目标操作并计时")
                steps.append("4. 记录完成耗时，判断是否满足时间要求")
            else:
                steps.append(f"2. 执行目标操作: {title}")
                steps.append("3. 记录性能指标数据")
            return "\n".join(steps)

        # ── 安全测试 ──
        if "安全" in dimension:
            steps = ["1. 安全测试准备"]
            if any(k in title for k in ["越权", "未授权", "权限"]):
                steps.append("2. 使用低权限用户 A 的凭证登录")
                steps.append("3. 尝试访问/操作高权限用户 B 的数据或功能")
                steps.append("4. 验证系统是否阻止越权访问并给出提示")
            elif any(k in title for k in ["注入", "SQL", "XSS", "脚本"]):
                steps.append(f"2. 在输入框/参数中注入恶意载荷: {tp.get('test_data', 'SQL/XSS 注入 payload')}")
                steps.append("3. 提交请求")
                steps.append("4. 检查是否被拦截或过滤，不泄露原始错误信息")
            elif any(k in title for k in ["篡改", "伪造", "篡改"]):
                steps.append("2. 使用抓包工具（如 Burp Suite/Charles）拦截请求")
                steps.append("3. 篡改关键参数（如金额、用户ID、token）")
                steps.append("4. 发送篡改后的请求")
                steps.append("5. 验证服务端是否校验了数据的合法性")
            elif any(k in title for k in ["敏感", "泄露", "加密"]):
                steps.append("2. 检查 API 响应和页面渲染是否包含敏感信息")
                steps.append("3. 检查敏感字段在传输和存储时是否加密")
                steps.append("4. 验证日志中是否过滤了敏感数据")
            elif any(k in title for k in ["认证", "身份", "Token", "token"]):
                steps.append("2. 使用无效/过期的 token 或认证信息访问接口")
                steps.append("3. 验证系统是否返回 401/403")
                steps.append("4. 验证系统是否正确处理认证异常")
            else:
                steps.append(f"2. 执行安全测试: {title}")
                steps.append("3. 验证系统的安全防护机制")
            return "\n".join(steps)

        # 兜底
        return f"1. 准备测试环境\n2. 执行操作: {title}\n3. 验证结果与预期一致"

    def __assign_priority(self, tp: dict) -> str:
        """分配优先级 - 基于模块、维度和测试类型"""
        title = tp["title"]
        dimension = tp["dimension"]
        module = tp["module"]

        # P0: 核心功能的正向测试、关键安全测试
        if "正向" in dimension:
            # 核心操作类关键字（任何系统都适用的）
            core_kw = ["登录", "注册", "创建", "新增", "提交", "支付",
                       "下单", "核心", "主流程", "关键", "基础"]
            if any(k in title for k in core_kw):
                return "P0"
            # 校验/验证类通常较重要
            if any(k in title for k in ["校验", "验证", "完整性"]):
                return "P0"
            # 集成/串联测试
            if any(k in title for k in ["集成", "串联", "全流程"]):
                return "P0"
            # 其余正向
            return "P1"

        # 安全测试 - 高风险场景 P0
        if "安全" in dimension:
            high_risk = ["越权", "篡改", "注入", "泄露", "认证", "敏感"]
            if any(k in title for k in high_risk):
                return "P0"
            return "P1"

        # 异常测试 - 关键异常 P0
        if "异常" in dimension:
            critical = ["并发", "支付", "核心", "关键", "数据丢失"]
            if any(k in title for k in critical):
                return "P0"
            return "P1"

        # 边界测试 - 边界场景通常是 P1
        if "边界" in dimension:
            return "P1"

        # 负向测试 - 通常是 P1
        if "负向" in dimension:
            # 涉及安全/权限的负向测试可提级
            if any(k in title for k in ["越权", "未授权", "未登录", "注入"]):
                return "P0"
            return "P1"

        # 性能测试 - 通常是 P2
        if "性能" in dimension:
            # 但如果是核心模块的性能则 P1
            if any(k in title for k in ["核心", "主流程", "并发", "高并发"]):
                return "P1"
            return "P2"

        return "P1"

    def __generate_precondition(self, tp: dict) -> str:
        """生成前置条件"""
        title = tp["title"]
        dimension = tp["dimension"]
        module = tp["module"]

        # 根据模块和维度生成前置条件
        if any(k in title for k in ["登录", "注册", "认证"]):
            return "用户处于登录/注册页面"

        if any(k in title for k in ["查询", "列表", "搜索", "筛选", "搜索"]):
            return "用户已登录，系统中有可查询的数据"

        if any(k in title for k in ["创建", "新增", "添加", "提交", "上传"]):
            return "用户已登录，具备创建/新增权限"

        if any(k in title for k in ["编辑", "修改", "更新", "变更"]):
            return "用户已登录，有可编辑的数据记录"

        if any(k in title for k in ["删除", "取消", "作废"]):
            return "用户已登录，有待删除/取消的数据记录"

        if any(k in title for k in ["导出", "下载"]):
            return "用户已登录，有可导出的数据"

        if any(k in title for k in ["校验", "验证", "检查", "对比"]):
            return "基准数据和待校验数据已准备就绪"

        if any(k in title for k in ["集成", "串联"]):
            return "前置环节已完成，相关服务运行正常"

        if any(k in title for k in ["性能", "并发", "效率"]):
            return "性能测试环境已就绪，监控工具已配置"

        # 异常/负向/安全
        if "异常" in dimension or "安全" in dimension or "负向" in dimension:
            return "测试环境已就绪，准备好测试工具和数据"

        if "边界" in dimension:
            return "测试环境已就绪，确定好边界值参数"

        # 默认
        return "用户已登录系统，测试环境正常可用"

    def generate(self, test_points: list) -> list:
        """生成测试用例"""
        test_cases = []
        for i, tp in enumerate(test_points, 1):
            test_cases.append({
                "id": f"TC-{i:03d}",
                "module": tp["module"],
                "feature": tp["feature"],
                "dimension": tp["dimension"],
                "title": tp["title"],
                "priority": self.__assign_priority(tp),
                "precondition": self.__generate_precondition(tp),
                "steps": self.__generate_steps(tp),
                "test_data": tp["test_data"],
                "预留": tp["expected"],
            })
        return test_cases

    # 维度关键词映射（英文 -> 中文）
    DIMENSION_ALIASES = {
        "positive": ["正向"],
        "negative": ["负向"],
        "boundary": ["边界"],
        "exception": ["异常"],
        "performance": ["性能"],
        "security": ["安全"],
        "basic": ["正向", "负向", "边界", "异常"],
    }

    def filter_by_dimensions(self, test_points: list, dimensions: str) -> list:
        """按测试维度过滤"""
        if dimensions == "all":
            return test_points

        # 解析关键词
        keywords = []
        for part in dimensions.split(","):
            part = part.strip()
            # 先尝试英文别名
            if part in self.DIMENSION_ALIASES:
                keywords.extend(self.DIMENSION_ALIASES[part])
            else:
                keywords.append(part)

        return [tp for tp in test_points
                if any(k in tp["dimension"] for k in keywords)]


# ═══════════════════════════════════════════════════════════════
# Excel 写入器
# ═════════════════════════════════　　　　　　　　　　　　　　　　

class ExcelWriter:
    """写入 Excel 文件，带格式化"""

    # 表头定义
    HEADERS = [
        ("用例编号", 12),
        ("所属模块", 16),
        ("功能点", 20),
        ("测试维度", 12),
        ("用例标题", 30),
        ("优先级", 10),
        ("前置条件", 25),
        ("测试步骤", 45),
        ("测试数据", 25),
        ("预期结果", 40),
        ("备注", 15),
        ("执行结果", 12),
    ]

    # 优先级颜色
    PRIORITY_FILLS = {
        "P0": PatternFill(start_color="FFCDD2", fill_type="solid"),  # 浅红
        "P1": PatternFill(start_color="FFE0B2", fill_type="solid"),  # 浅橙
        "P2": PatternFill(start_color="C8E6C9", fill_type="solid"),  # 浅绿
    }

    def __write(self, test_cases: list, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # ── 表头 ──
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F4F4F", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col, (header, width) in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[1].height = 30

        # ── 数据行 ──
        data_font = Font(name="微软雅黑", size=10)
        for row_idx, tc in enumerate(test_cases, 2):
            row_data = [
                tc["id"], tc["module"], tc["feature"], tc["dimension"],
                tc["title"], tc["priority"], tc["precondition"],
                tc["steps"], tc["test_data"], tc["预留"], "", ""
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border

            # 优先级着色
            if tc["priority"] in self.PRIORITY_FILLS:
                ws.cell(row=row_idx, column=6).fill = self.PRIORITY_FILLS[tc["priority"]]

            # 行高根据步骤数调整
            step_count = tc["steps"].count("\n") + 1
            ws.row_dimensions[row_idx].height = max(30, step_count * 18)

        # ── 冻结首行 ──
        ws.freeze_panes = "A2"

        # ── 自动筛选 ──
        last_col = get_column_letter(len(self.HEADERS))
        ws.auto_filter.ref = f"A1:{last_col}{len(test_cases)+1}"

        wb.save(output_path)

    def write(self, test_cases: list, output_path: str):
        self.__write(test_cases, output_path)


# ═════════════════════════════════════════ 计数助手 ═══════════════════════════

def count_stats(test_cases: list) -> dict:
    """统计用例分布"""
    stats = {
        "total": len(test_cases),
        "modules": set(),
        "features": set(),
        "dimensions": {},
        "priorities": {},
    }
    for tc in test_cases:
        stats["modules"].add(tc["module"])
        stats["features"].add(tc["feature"])
        stats["dimensions"][tc["dimension"]] = stats["dimensions"].get(tc["dimension"], 0) + 1
        stats["priorities"][tc["priority"]] = stats["priorities"].get(tc["priority"], 0) + 1
    return stats


# ═══════════════ SkillScript ┐ ═══════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="生成 Excel 格式测试用例")
    parser.add_argument("input", help="测试点 Markdown 文件路径")
    parser.add_argument("-o", "--output", default="testcases.xlsx", help="输出文件路径")
    parser.add_argument(
        "-d", "--dimensions",
        default="all",
        help="测试维度过滤: all|basic|positive,negative,boundary,exception,performance,security"
    )
    args = parser.parse_args()

    # 1. 解析测试点
    print(f"📖 读取测试点文件: {args.input}")
    test_point_parser = TestPointParser()
    test_points = test_point_parser.parse_file(args.input)

    if not test_points:
        print("⚠️  未解析到任何测试点，请检查文件格式", file=sys.stderr)
        return 1

    print(f"✅ 解析到 {len(test_points)} 个测试点")

    # 2. 过滤
    if args.dimensions != "all":
        test_points = TestCaseGenerator().filter_by_dimensions(test_points, args.dimensions)
        print(f"🔍 过滤后剩余 {len(test_points)} 个测试点")

    # 3. 生成用例
    print("🔨 生成测试用例...")
    generator = TestCaseGenerator()
    test_cases = generator.generate(test_points)

    # 4. 统计
    stats = count_stats(test_cases)

    print(f"\n✅ 测试用例生成完成！")
    print(f"📊 统计信息：")
    print(f"  - 测试模块：{len(stats['modules'])} 个")
    print(f"  - 功能点：{len(stats['features'])} 个")
    print(f"  - 用例总数：{stats['total']} 个")
    print(f"\n📊 测试维度分布：")
    for dim, count in stats["dimensions"].items():
        print(f"  - {dim}: {count} 个")
    print(f"\n📊 优先级分布：")
    for pri in ["P0", "P1", "P2"]:
        count = stats["priorities"].get(pri, 0)
        print(f"  - {pri}: {count} 个")

    # 5. 写入 Excel
    print(f"\n📁 写入文件: {args.output}")
    excel_writer = ExcelWriter()
    excel_writer.write(test_cases, args.output)

    print(f"\n✅ 文件已保存: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())